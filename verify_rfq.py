#!/usr/bin/env python3
"""verify_rfq.py — independent recompute of the auto-rfq-banana headline numbers.

Pattern source: supplier-pricing/verify_impact.py — sibling Python script that
recomputes the on-screen totals via a deliberately-different code path. The
"red / green" test for the math when the methodology is questioned.

Usage:
    python3 verify_rfq.py /path/to/multi_year_export.xlsx

Prints:
  - Item count (deduplicated)
  - PO count (distinct)
  - Total spend (all-time)
  - Spend in 12 / 24 / 36-month windows
  - Items active in each window
  - Date range + years-span

Compare these numbers against the KPI tiles in the app. Mismatch > a few % =
something to investigate (column-mapping mistake, parse bug, or a real edge
case the in-app extractor handles differently).

Independent code paths:
  - Uses openpyxl directly with no shared helpers from app.py
  - Walks rows linearly (not by item bucket)
  - Re-derives "now" from the data (max date) — same anchor as the app
"""

from __future__ import annotations

import sys
import re
from collections import defaultdict
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Same alias dict shape as app.py — but rewritten by hand here so a typo there
# doesn't quietly propagate. Keep these patterns in sync with EXPORT_ALIASES.
ALIASES = {
    "item":  ["item #", "item number", "andersen item", "item num"],
    "eam":   ["eam part number", "eam part", "eam pn", "eam #"],
    # Supplier's own catalog SKU — fallback dedup key when item / eam are blank
    # (McMaster cXML orders typically leave the Andersen-side fields empty).
    # Must NOT match "Manufacturer Part Number" — auto_map runs in dict order
    # so we don't include "manufacturer part" in this list.
    "part":  ["part number", "supplier part", "supplier sku", "supplier auxiliary part number"],
    "date":  ["order date", "po date", "transaction date"],
    "qty":   ["quantity", "qty ordered", "order qty", "qty"],
    "price": ["unit price", "price per unit", "price each", "price"],
    "po":    ["po number", "po #", "purchase order"],
}


def _is_blanky(s):
    """Treat 'N/A' and friends as effectively blank (McMaster pattern)."""
    if s is None:
        return True
    t = str(s).strip().upper()
    return t in ("", "N/A", "#N/A", "NA", "-", "—", "(BLANK)", "NULL", "NONE")


def auto_map(headers):
    """Two-pass: exact-equality first, then substring fallback. Without exact-
    first, 'part number' substring-matches 'Supplier Auxiliary Part Number'
    before reaching the literal 'Part Number' column. McMaster bug 2026-04-26."""
    out = {}
    norm = [(h or "").strip().lower() for h in headers]

    # Pass 1: exact equality (skip MFG cols when matching "part")
    for field, pats in ALIASES.items():
        for pat in pats:
            for i, h in enumerate(norm):
                if not h or i in out.values():
                    continue
                if field == "part" and "manufacturer" in h:
                    continue
                if pat == h:
                    out[field] = i
                    break
            if field in out:
                break

    # Pass 2: substring fallback
    for field, pats in ALIASES.items():
        if field in out:
            continue
        for pat in pats:
            for i, h in enumerate(norm):
                if not h or i in out.values():
                    continue
                if field == "part" and "manufacturer" in h:
                    continue
                if pat in h:
                    out[field] = i
                    break
            if field in out:
                break
    return out


def parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s.split(" ")[0] if "T" not in s else s.split("T")[0], fmt)
        except ValueError:
            continue
    return None


def safe_float(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def norm_pn(s):
    if s is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def main(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    headers = [str(h or "").strip() for h in headers]
    m = auto_map(headers)

    print(f"Source:    {path}")
    print(f"Sheet:     {wb.sheetnames[0]}")
    print(f"Mapping:   {m}")
    print(f"Columns:   {len(headers)}")
    print()

    items = defaultdict(lambda: {"spend": 0.0, "qty": 0.0, "first": None, "last": None})
    pos = set()
    rows_total = 0
    rows_used = 0
    annual = defaultdict(float)

    rows = list(rows_iter)
    for row in rows:
        rows_total += 1
        if row is None:
            continue
        def g(field):
            i = m.get(field)
            return None if (i is None or i >= len(row)) else row[i]
        item_raw = g("item") or g("eam") or g("part")
        if _is_blanky(item_raw):
            item_raw = None
        if not item_raw:
            continue
        d = parse_date(g("date"))
        q = safe_float(g("qty"))
        p = safe_float(g("price"))
        if d is None or q is None or p is None:
            continue
        key = norm_pn(item_raw)
        ext = q * p
        rec = items[key]
        rec["spend"] += ext
        rec["qty"] += q
        if rec["first"] is None or d < rec["first"]:
            rec["first"] = d
        if rec["last"] is None or d > rec["last"]:
            rec["last"] = d
        po = g("po")
        if po:
            pos.add(str(po).strip())
        annual[d.year] += ext
        rows_used += 1

    if not items:
        print("No usable rows. Check column mapping.")
        return

    last_dt = max(rec["last"] for rec in items.values())
    first_dt = min(rec["first"] for rec in items.values())
    years_span = (last_dt - first_dt).days / 365.25

    cutoffs = {w: last_dt - timedelta(days=int(w * 365 / 12)) for w in (12, 24, 36)}
    spend_w = {w: 0.0 for w in (12, 24, 36)}
    items_w = {w: 0 for w in (12, 24, 36)}
    # Re-walk for window math (counted at LINE level, but item-counted here)
    for key, rec in items.items():
        # Need per-line dates for accurate window aggregation; rebuild cheaply:
        pass
    # Cleaner: re-iterate the source for window sums (tradeoff: I/O vs memory)
    wb2 = load_workbook(path, data_only=True, read_only=True)
    ws2 = wb2[wb2.sheetnames[0]]
    it = ws2.iter_rows(values_only=True)
    next(it)
    item_active = {w: set() for w in (12, 24, 36)}
    for row in it:
        if row is None:
            continue
        def g(field):
            i = m.get(field)
            return None if (i is None or i >= len(row)) else row[i]
        item_raw = g("item") or g("eam") or g("part")
        if _is_blanky(item_raw):
            item_raw = None
        if not item_raw:
            continue
        d = parse_date(g("date"))
        q = safe_float(g("qty"))
        p = safe_float(g("price"))
        if d is None or q is None or p is None:
            continue
        ext = q * p
        for w in (12, 24, 36):
            if d >= cutoffs[w]:
                spend_w[w] += ext
                item_active[w].add(norm_pn(item_raw))

    print("=" * 60)
    print("HEADLINE NUMBERS (verify against the app's KPI tiles)")
    print("=" * 60)
    print(f"  Items (deduped):     {len(items):>12,}")
    print(f"  Distinct POs:        {len(pos):>12,}")
    print(f"  Order lines:         {rows_used:>12,}  (rows skipped: {rows_total - rows_used:,})")
    print(f"  Total spend:         ${sum(r['spend'] for r in items.values()):>14,.2f}")
    print(f"  Date range:          {first_dt.date()} → {last_dt.date()}  ({years_span:.1f} yr)")
    print()
    print(f"  12-mo spend:         ${spend_w[12]:>14,.2f}   active items: {len(item_active[12]):>5,}")
    print(f"  24-mo spend:         ${spend_w[24]:>14,.2f}   active items: {len(item_active[24]):>5,}")
    print(f"  36-mo spend:         ${spend_w[36]:>14,.2f}   active items: {len(item_active[36]):>5,}")
    print()
    print("Annual spend by year:")
    for y in sorted(annual):
        print(f"  {y}: ${annual[y]:>14,.2f}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 verify_rfq.py /path/to/multi_year_export.xlsx", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1])
