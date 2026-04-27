#!/usr/bin/env python3
"""verify_isolation.py — independent assertion that supplier-bound files are
strictly isolated.

Pattern source: verify_rfq.py (same red/green sibling-script approach).

Given a folder of award letter xlsx files (one per supplier), this script:
  1. Identifies the intended supplier for each file (from filename or cover sheet).
  2. Scans every cell of every sheet for any OTHER supplier name.
  3. Refuses with non-zero exit if any cross-supplier leakage is detected.

Run before sending any batch of award letters. Independent of app.py logic
on purpose — this is the cross-check that the in-app isolation guard
actually worked.

Usage:
    python3 verify_isolation.py /path/to/award_letters_folder/

Exit 0  = all clean, safe to send
Exit 1  = isolation violation found, do not send
Exit 2  = usage error
"""

from __future__ import annotations

import sys
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def _supplier_from_filename(stem: str) -> str | None:
    """Award letter filename convention: AwardLetter_<supplier>_<...>.xlsx
    or RFQ-...-<supplier>.xlsx. Try a few patterns."""
    for prefix in ("AwardLetter_", "Award_", "Award-"):
        if stem.startswith(prefix):
            rest = stem[len(prefix):]
            return rest.split("_")[0].split("-")[0]
    # Fall back: take everything before first underscore or dash
    parts = re.split(r"[_-]", stem)
    if parts:
        return parts[0]
    return None


def _supplier_from_cover(path: Path) -> str | None:
    """Look for 'Awarded to: <supplier>' on the first sheet."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(max_row=10, values_only=True):
            for v in row:
                if v and isinstance(v, str) and v.strip().lower().startswith("awarded to:"):
                    return v.split(":", 1)[1].strip()
        wb.close()
    except Exception:
        pass
    return None


def _is_internal_audience(path: Path) -> bool:
    """Internal-audience files (per the engine convention) carry an INTERNAL
    banner on the first sheet. Skip isolation checks on those — they're
    SUPPOSED to mention every supplier."""
    if path.stem.upper().startswith("INTERNAL"):
        return True
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(max_row=5, values_only=True):
            for v in row:
                if v and isinstance(v, str) and "INTERNAL" in v.upper() and "NEVER FORWARD" in v.upper():
                    wb.close()
                    return True
        wb.close()
    except Exception:
        pass
    return False


def scan_for_other_suppliers(path: Path, intended_supplier: str, all_suppliers: list) -> list:
    """Return a list of (sheet, row, col, value, foreign_supplier) tuples
    for any cell containing a supplier name OTHER than the intended one."""
    violations = []
    others = [s for s in all_suppliers if s and s != intended_supplier]
    if not others:
        return []
    wb = load_workbook(path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for c_idx, v in enumerate(row, start=1):
                if v is None:
                    continue
                sv = str(v)
                if not sv.strip():
                    continue
                for other in others:
                    # Word-boundary-ish substring
                    if other in sv:
                        violations.append((sheet_name, r_idx, c_idx, sv[:80], other))
    wb.close()
    return violations


def main(folder_path: str):
    folder = Path(folder_path)
    if not folder.is_dir():
        print(f"Not a directory: {folder_path}", file=sys.stderr)
        sys.exit(2)
    files = sorted(folder.glob("*.xlsx"))
    if not files:
        print(f"No .xlsx files in {folder_path}", file=sys.stderr)
        sys.exit(2)

    # First pass: collect intended supplier per file. Skip internal-audience
    # files (their whole purpose is to contain all supplier names).
    file_suppliers: dict = {}
    skipped_internal = []
    for path in files:
        if _is_internal_audience(path):
            skipped_internal.append(path)
            continue
        sup = _supplier_from_cover(path) or _supplier_from_filename(path.stem)
        if not sup:
            print(f"WARN: could not detect intended supplier for {path.name}", file=sys.stderr)
            continue
        file_suppliers[path] = sup

    all_suppliers = sorted(set(file_suppliers.values()))
    print(f"Detected {len(all_suppliers)} suppliers across {len(file_suppliers)} supplier-bound files:")
    for s in all_suppliers:
        n = sum(1 for sup in file_suppliers.values() if sup == s)
        print(f"  · {s}  ({n} file{'s' if n != 1 else ''})")
    if skipped_internal:
        print(f"\nSkipped {len(skipped_internal)} internal-audience file(s):")
        for p in skipped_internal:
            print(f"  · {p.name}  (INTERNAL — intentionally contains all suppliers)")
    print()

    # Second pass: scan each file for foreign supplier names
    total_violations = 0
    for path, sup in file_suppliers.items():
        viols = scan_for_other_suppliers(path, sup, all_suppliers)
        if viols:
            total_violations += len(viols)
            print(f"❌ {path.name} (intended for {sup}) — {len(viols)} cell(s) contain other supplier names:")
            for sheet, r, c, val, other in viols[:10]:
                print(f"     [{sheet}!R{r}C{c}] mentions {other!r}: {val[:60]}")
            if len(viols) > 10:
                print(f"     ... and {len(viols) - 10} more")
            print()
        else:
            print(f"✅ {path.name} (intended for {sup}) — clean")

    print()
    print("=" * 60)
    if total_violations == 0:
        print("✅ ALL CLEAR — every award letter contains only its intended supplier")
        sys.exit(0)
    else:
        print(f"❌ ISOLATION FAILURE — {total_violations} violation(s) across {sum(1 for p in file_suppliers if scan_for_other_suppliers(p, file_suppliers[p], all_suppliers))} files")
        print("DO NOT SEND these files until the violations are fixed.")
        sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 verify_isolation.py /path/to/award_letters_folder/", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1])
