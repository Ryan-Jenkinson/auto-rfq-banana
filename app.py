"""auto-rfq-banana — Python engine (Pyodide-loaded at boot).

Module name: app_engine (we register it under that name so app.js can `from app_engine import ...`).

Stage A (Phase 1): parse a multi-year supplier export and build a deduplicated
candidate RFQ item list with multi-window aggregations + annual breakdowns.

Pattern source: supplier-pricing/app.py — alias system, MFG canonicalization,
normalization helpers. Stripped of per-cycle / matching logic — this app
inverts the model (build the master from history, not match against it).
"""

from __future__ import annotations

import io
import re
import sys
import json
import math
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell


# ---------------------------------------------------------------------------
# Normalization helpers (port from supplier-pricing/app.py)
# ---------------------------------------------------------------------------

def norm_pn(s) -> str:
    """Normalize a part number for keying — strip non-alphanumerics, uppercase."""
    if s is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def norm_text(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


def safe_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").replace("$", "").strip()
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


def parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Excel date strings come in many flavors; try the common ones
    for fmt in (
        "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y",
        "%d/%m/%Y", "%Y/%m/%d", "%m-%d-%Y",
    ):
        try:
            return datetime.strptime(s.split(" ")[0] if fmt == "%Y-%m-%d" else s, fmt)
        except ValueError:
            continue
    return None


# Manufacturer aliases — verbatim port from supplier-pricing
MFG_ALIASES = {
    "AB": "ALLEN-BRADLEY", "ALLENBRADLEY": "ALLEN-BRADLEY",
    "TB": "THOMAS AND BETTS", "T&B": "THOMAS AND BETTS",
    "GE": "GENERAL ELECTRIC",
    "SQ D": "SQUARE D", "SQD": "SQUARE D",
    "HUB": "HUBBELL", "HUBBL": "HUBBELL",
    "PAN": "PANDUIT", "PND": "PANDUIT",
    "EATN": "EATON",
    "RKWL": "ROCKWELL", "ROCK": "ROCKWELL", "RA": "ROCKWELL",
    "FEST": "FESTO",
    "SMC": "SMC",
    "OMRON": "OMRON",
    "SIEMENS": "SIEMENS",
    "SCHNEIDER": "SCHNEIDER",
    "NUM": "NUMATICS",  # NUMATICS often appears as "NUM" prefix in JHF data
}


def canon_mfg(s) -> str:
    if not s:
        return ""
    raw = re.sub(r"[^A-Z0-9 &-]", "", str(s).upper()).strip()
    if raw in MFG_ALIASES:
        return MFG_ALIASES[raw]
    # Try first token
    first = raw.split()[0] if raw.split() else ""
    if first in MFG_ALIASES:
        return MFG_ALIASES[first]
    return raw


# ---------------------------------------------------------------------------
# Column alias system — for the multi-year export
# ---------------------------------------------------------------------------
# Each value is a list of substring patterns (case-insensitive). First match wins.
# Patterns are MULTI-WORD or distinctive — never a bare common word that could
# false-positive (per supplier-pricing dev-lessons memory item #3).

EXPORT_ALIASES = {
    "item_num":    ["item #", "item number", "andersen item", "item num", "item id"],
    "eam_pn":      ["eam part number", "eam part", "eam pn", "eam part #", "eam #"],
    "description": ["item name", "detailed item", "description", "item description", "long description", "desc"],
    "mfg_name":    ["manufacturer name", "manufacturer", "mfg name", "mfr name", "make"],
    "mfg_pn":      ["manufacturer part #", "manufacturer part number", "mfg part #", "mfr part", "mpn", "manufacturer pn", "mfg pn"],
    # part_number = supplier's own catalog SKU (e.g. McMaster's "5709A45").
    # MUST come AFTER mfg_pn so "Manufacturer Part Number" gets claimed first.
    # Used as a fallback dedup key when item_num + eam_pn are both blank
    # (McMaster orders via cXML mostly leave those Andersen-side fields empty).
    "part_number": ["part number", "supplier part number", "supplier sku", "supplier auxiliary part number", "catalog number", "catalog #"],
    "order_date":  ["order date", "po date", "transaction date", "date ordered", "po creation date"],
    "qty":         ["quantity", "qty ordered", "qty received", "order qty", "qty"],
    "unit_price":  ["unit price", "price per unit", "unit cost", "price each", "current price", "price"],
    "po_number":   ["po number", "po #", "purchase order", "po num", "po no"],
    "uom":         ["unit of measure", "uom", "unit measure", "u/m", "purch uom"],
    "commodity":   ["commodity", "category", "product category", "spend category"],
    "supplier":    ["supplier name", "supplier", "vendor name", "vendor"],
}


def _is_blanky(s) -> bool:
    """Treat 'N/A', '#N/A', '-', '—', '(blank)', etc. as effectively blank.
    McMaster orders set Manufacturer / MFG Part Number to literal 'N/A'."""
    if s is None:
        return True
    t = str(s).strip().upper()
    return t in ("", "N/A", "#N/A", "NA", "-", "—", "(BLANK)", "NULL", "NONE")


# ---------------------------------------------------------------------------
# UOM normalization
#
# Suppliers spell the same UOM many ways — EA/Each/EACH/each. Map them to a
# canonical short form for comparison + display. Risky changes (Box→Each,
# Roll→Feet, Pair→Each) are flagged but never auto-converted without an
# explicit conversion factor.
# ---------------------------------------------------------------------------

UOM_CANONICAL = {
    # Each
    "EA": "EA", "EACH": "EA", "EACHES": "EA", "EAS": "EA",
    "PCE": "EA", "PIECE": "EA", "PIECES": "EA", "PC": "EA", "UNIT": "EA",
    # Box
    "BX": "BX", "BOX": "BX", "BOXES": "BX",
    # Pack
    "PK": "PK", "PACK": "PK", "PACKAGE": "PK", "PACKAGES": "PK", "PKG": "PK",
    # Foot
    "FT": "FT", "FOOT": "FT", "FEET": "FT",
    # Inch
    "IN": "IN", "INCH": "IN", "INCHES": "IN",
    # Pound
    "LB": "LB", "LBS": "LB", "POUND": "LB", "POUNDS": "LB",
    # Kilogram
    "KG": "KG", "KILO": "KG", "KILOGRAM": "KG", "KILOGRAMS": "KG",
    # Gallon / quart / pint / ounce
    "GAL": "GAL", "GALLON": "GAL", "GALLONS": "GAL",
    "QT": "QT", "QUART": "QT", "QUARTS": "QT",
    "PT": "PT", "PINT": "PT", "PINTS": "PT",
    "OZ": "OZ", "OUNCE": "OZ", "OUNCES": "OZ",
    # Roll / case / dozen / pair / set / kit
    "RL": "RL", "ROLL": "RL", "ROLLS": "RL",
    "CS": "CS", "CASE": "CS", "CASES": "CS",
    "DZ": "DZ", "DOZ": "DZ", "DOZEN": "DZ",
    "PR": "PR", "PAIR": "PR", "PAIRS": "PR",
    "ST": "SET", "SET": "SET", "SETS": "SET",
    "KT": "KIT", "KIT": "KIT", "KITS": "KIT",
    # Metric length
    "M": "M", "METER": "M", "METERS": "M", "METRE": "M",
    "CM": "CM", "CENTIMETER": "CM", "CENTIMETERS": "CM",
    "MM": "MM", "MILLIMETER": "MM", "MILLIMETERS": "MM",
    # Bag / bundle
    "BG": "BG", "BAG": "BG", "BAGS": "BG",
    "BD": "BD", "BUNDLE": "BD", "BUNDLES": "BD",
}


def canon_uom(s) -> str:
    """Return the canonical UOM short form. Empty string if unknown/blank."""
    if not s:
        return ""
    t = str(s).strip().upper()
    if not t or _is_blanky(t):
        return ""
    # Strip trailing periods (E.A. → EA)
    t = t.rstrip(".").rstrip("/")
    return UOM_CANONICAL.get(t, t)


# UOM pairs that are inherently risky to compare or convert without a
# conversion factor. Direction matters: ("BX", "EA") means treating Box as
# Each is the wrong move. We flag these on the per-item record so the UI can
# surface the concern; never auto-convert.
RISKY_UOM_CHANGES = {
    ("BX", "EA"), ("EA", "BX"),
    ("PK", "EA"), ("EA", "PK"),
    ("CS", "EA"), ("EA", "CS"),
    ("RL", "FT"), ("FT", "RL"),
    ("RL", "IN"), ("IN", "RL"),
    ("PR", "EA"), ("EA", "PR"),
    ("DZ", "EA"), ("EA", "DZ"),
    ("GAL", "OZ"), ("OZ", "GAL"),
    ("GAL", "QT"), ("QT", "GAL"),
    ("LB", "EA"), ("EA", "LB"),
    ("BG", "EA"), ("EA", "BG"),
    ("KIT", "EA"), ("EA", "KIT"),
    ("SET", "EA"), ("EA", "SET"),
}


def is_risky_uom_change(from_uom: str, to_uom: str) -> bool:
    a = canon_uom(from_uom)
    b = canon_uom(to_uom)
    if not a or not b or a == b:
        return False
    return (a, b) in RISKY_UOM_CHANGES


# ---------------------------------------------------------------------------
# Description-pattern flags
#
# Items whose description suggests they don't belong in a standard RFQ:
# services, freight, custom one-offs, etc. Surface these as flags so the
# user can drop them from the candidate list intentionally.
# ---------------------------------------------------------------------------

DESC_PATTERN_FLAGS = {
    "service":  ["service", "labor", "installation install", "tech support", "consulting"],
    "freight":  ["freight", "shipping charge", "delivery charge", "expedite"],
    "tariff":   ["tariff", "duty", "customs charge"],
    "custom":   ["custom", "made-to-order", "made to order", "special order", "bespoke", "non-standard"],
    "repair":   ["repair", "rebuild", "refurbish", "remanufactured"],
    "rental":   ["rental", "lease", "loaner"],
    "misc":     ["misc", "miscellaneous", "general supplies", "various", "assorted"],
    "obsolete": ["obsolete", "discontinued", "no longer available", "end of life", "eol"],
    "generic":  [],   # placeholder — handled separately by length check
}


def description_pattern_flags(desc: str) -> list:
    """Return a list of pattern category names that fired for this description.
    Used for both Step 3 curation and the readiness scoring."""
    if not desc:
        return ["generic"]
    d = str(desc).strip().lower()
    if not d:
        return ["generic"]
    flags = []
    for category, patterns in DESC_PATTERN_FLAGS.items():
        for pat in patterns:
            # Word-boundary-ish match: flank with spaces or string edges
            if pat in d:
                flags.append(category)
                break
    # Generic: very short descriptions like "Cord Grip" / "Brackets"
    # or descriptions with no spaces (single-word items)
    if "generic" not in flags and (len(d) < 12 or " " not in d):
        flags.append("generic")
    return flags


def auto_map_export(headers: list) -> dict:
    """Given a list of header strings, return {field → header_idx} where matched.

    Two-pass: exact-equality first across all (field, pattern) pairs, THEN
    substring fallback for fields that didn't get an exact hit. Without this,
    a substring like "part number" matching "Supplier Auxiliary Part Number"
    (mostly empty col) wins over the literal "Part Number" column that's
    actually populated. McMaster bug 2026-04-26."""
    out = {}
    norm_headers = [norm_text(h).lower() if h else "" for h in headers]

    # Pass 1: exact equality
    for field, patterns in EXPORT_ALIASES.items():
        for pat in patterns:
            pat_l = pat.lower()
            for i, h in enumerate(norm_headers):
                if not h or i in out.values():
                    continue
                if h == pat_l:
                    out[field] = i
                    break
            if field in out:
                break

    # Pass 2: substring fallback for unmapped fields
    for field, patterns in EXPORT_ALIASES.items():
        if field in out:
            continue
        for pat in patterns:
            pat_l = pat.lower()
            for i, h in enumerate(norm_headers):
                if not h or i in out.values():
                    continue
                if pat_l in h:
                    out[field] = i
                    break
            if field in out:
                break
    return out


# ---------------------------------------------------------------------------
# Workbook inspection (cheap header peek)
# ---------------------------------------------------------------------------

def inspect_workbook(file_bytes) -> dict:
    if not isinstance(file_bytes, (bytes, bytearray)):
        file_bytes = bytes(file_bytes)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    # Pick the first non-empty sheet (most exports have a single data sheet)
    sheet_name = None
    headers = []
    row_count = 0
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            continue
        headers = [norm_text(c) for c in first]
        sheet_name = name
        # Count rows quickly via max_row attr
        row_count = max(0, (ws.max_row or 1) - 1)
        break
    wb.close()
    return {
        "sheet_name": sheet_name or "",
        "headers": headers,
        "row_count": row_count,
    }


# ---------------------------------------------------------------------------
# Module-level state (so xlsx generator can read post-extraction)
# ---------------------------------------------------------------------------
_STATE: dict = {
    "items": [],
    "kpis": {},
    "annual_spend": [],
    "supplier_name": "",
    "source_file_stem": "",
    "windows": (12, 24, 36),
    # Per-item outlier exclusions for the per-item history modal. Maps the
    # item_num (display key) → list of 0-based indices into the ASCENDING-
    # date-sorted po_lines for that item. Excluded indices are dropped from
    # the trend/R²/median/expected-today recompute when the modal opens.
    # Persisted via serialize_state / restore_state.
    "item_exclusions": {},
    # Per-item supplier locks — analyst-confirmed pinning of an item's
    # award to a specific supplier after visually auditing a bid. Maps
    # item_num → {"supplier": str, "reason": str, "locked_at": isoformat}.
    # Locks are evaluated AFTER scenario `overrides` but BEFORE strategy
    # logic in `_evaluate_scenario`, so the same lock applies across every
    # scenario the user runs. A lock with no matching priced bid is
    # remembered but does not force an award (recorded as a warning).
    "item_locks": {},
    # Excluded-line review log — every time the analyst unticks a
    # suspicious priced line in the per-item modal, an entry is appended
    # with the line's full data PLUS the pre-exclusion median + avg of
    # the OTHER priced lines for that item, so a downstream "show me what
    # I removed and why" review (or auditor request) can reconstruct the
    # decision. Persisted via serialize_state / restore_state.
    # Schema per entry: see _append_exclusion_log_entries.
    "exclusion_log": [],
}


# ---------------------------------------------------------------------------
# Extraction — multi-year export → deduped item list with windowed aggs
# ---------------------------------------------------------------------------

def _add(d: dict, k, v):
    d[k] = d.get(k, 0) + v


def _extract_rows(file_bytes, mapping: dict) -> list:
    """Stream rows from the workbook, returning a list of dicts (parsed)."""
    if not isinstance(file_bytes, (bytes, bytearray)):
        file_bytes = bytes(file_bytes)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Resolve column indexes once
    def col(field):
        i = mapping.get(field)
        return i if (i is not None and i >= 0) else None

    c_item = col("item_num")
    c_eam = col("eam_pn")
    c_desc = col("description")
    c_mfg = col("mfg_name")
    c_mpn = col("mfg_pn")
    c_part = col("part_number")
    c_date = col("order_date")
    c_qty = col("qty")
    c_price = col("unit_price")
    c_po = col("po_number")
    c_uom = col("uom")
    c_comm = col("commodity")
    c_sup = col("supplier")

    def _clean(v):
        """norm_text but also coerce literal 'N/A'-style values to ''."""
        s = norm_text(v)
        return "" if _is_blanky(s) else s

    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row is None:
            continue
        def g(i):
            return None if (i is None or i >= len(row)) else row[i]
        item = _clean(g(c_item)) if c_item is not None else ""
        eam = _clean(g(c_eam)) if c_eam is not None else ""
        part = _clean(g(c_part)) if c_part is not None else ""
        # Dedup-key fallback: Andersen Item # → EAM Part # → supplier Part Number.
        # Distributors that order via cXML (McMaster, Grainger PunchOut, etc.)
        # often leave the Andersen-side fields blank because the part lives in
        # the supplier's catalog, not Andersen's item-master.
        key_raw = item or eam or part
        if not key_raw:
            continue
        rows.append({
            "key": norm_pn(key_raw),
            "item_num": item or eam or part,
            "eam_pn": eam,
            "part_number": part,
            "description": _clean(g(c_desc)) if c_desc is not None else "",
            "mfg_name": _clean(g(c_mfg)) if c_mfg is not None else "",
            "mfg_pn": _clean(g(c_mpn)) if c_mpn is not None else "",
            "order_date": parse_date(g(c_date)) if c_date is not None else None,
            "qty": safe_float(g(c_qty)) if c_qty is not None else None,
            "unit_price": safe_float(g(c_price)) if c_price is not None else None,
            "po": _clean(g(c_po)) if c_po is not None else "",
            "uom": _clean(g(c_uom)) if c_uom is not None else "",
            "commodity": _clean(g(c_comm)) if c_comm is not None else "",
            "supplier": _clean(g(c_sup)) if c_sup is not None else "",
        })
    wb.close()
    return rows


def extract_rfq_list(file_bytes, mapping: dict) -> dict:
    """Build the candidate RFQ item list from a multi-year export.

    Returns:
        {
          "kpis": {...},
          "items": [{item_num, description, mfg_name, mfg_pn, qty_12mo, spend_12mo, ...}],
          "annual_spend": [{year, spend, item_count}, ...]
        }
    """
    rows = _extract_rows(file_bytes, mapping)
    if not rows:
        return {"kpis": {}, "items": [], "annual_spend": []}

    # Anchor "now" at the most recent order date in the data — not real today,
    # because the export might be a few weeks stale.
    valid_dates = [r["order_date"] for r in rows if r["order_date"]]
    now = max(valid_dates) if valid_dates else datetime.now()
    cutoffs = {
        12: now - timedelta(days=365),
        24: now - timedelta(days=730),
        36: now - timedelta(days=1095),
    }

    items: dict = defaultdict(lambda: {
        "key": "",
        "item_num": "",
        "eam_pn": "",
        "part_number": "",
        "description": "",
        "mfg_name_counts": Counter(),
        "mfg_pn_counts": Counter(),
        "uom_counts": Counter(),
        "commodity_counts": Counter(),
        "po_set": set(),
        # Per-item PO line list, used by the per-item drill-down modal.
        # Each entry: (date_iso, qty, unit_price, line_total, po, uom)
        "po_lines": [],
        "qty_12mo": 0.0, "spend_12mo": 0.0,
        "qty_24mo": 0.0, "spend_24mo": 0.0,
        "qty_36mo": 0.0, "spend_36mo": 0.0,
        "qty_all": 0.0,  "spend_all": 0.0,
        "first_order": None, "last_order": None,
        "last_unit_price": None,
        "last_order_dt": None,
    })

    annual: dict = defaultdict(lambda: {"spend": 0.0, "items": set(), "qty": 0.0})
    supplier_counter = Counter()
    line_count = 0

    for r in rows:
        if r["qty"] is None or r["unit_price"] is None or r["order_date"] is None:
            continue
        line_count += 1
        ext = r["qty"] * r["unit_price"]
        k = r["key"]
        rec = items[k]
        rec["key"] = k
        if not rec["item_num"]:
            rec["item_num"] = r["item_num"]
            rec["eam_pn"] = r["eam_pn"]
            rec["part_number"] = r["part_number"]
        # Description: take the first non-empty seen (descriptions vary slightly across years)
        if not rec["description"] and r["description"]:
            rec["description"] = r["description"]
        if r["mfg_name"]:
            rec["mfg_name_counts"][r["mfg_name"]] += 1
        if r["mfg_pn"]:
            rec["mfg_pn_counts"][r["mfg_pn"]] += 1
        if r["uom"]:
            rec["uom_counts"][r["uom"]] += 1
        if r["commodity"]:
            rec["commodity_counts"][r["commodity"]] += 1
        if r["po"]:
            rec["po_set"].add(r["po"])

        rec["po_lines"].append((
            r["order_date"].date().isoformat() if r["order_date"] else "",
            r["qty"],
            r["unit_price"],
            ext,
            r["po"],
            r["uom"],
        ))

        rec["qty_all"] += r["qty"]
        rec["spend_all"] += ext

        d = r["order_date"]
        if rec["first_order"] is None or d < rec["first_order"]:
            rec["first_order"] = d
        if rec["last_order_dt"] is None or d > rec["last_order_dt"]:
            rec["last_order_dt"] = d
            rec["last_unit_price"] = r["unit_price"]

        for w in (12, 24, 36):
            if d >= cutoffs[w]:
                rec[f"qty_{w}mo"] += r["qty"]
                rec[f"spend_{w}mo"] += ext

        # Annual rollup
        y = d.year
        annual[y]["spend"] += ext
        annual[y]["qty"] += r["qty"]
        annual[y]["items"].add(k)

        if r["supplier"]:
            supplier_counter[r["supplier"]] += 1

    # Flatten items
    out_items = []
    for k, rec in items.items():
        mfg_name = rec["mfg_name_counts"].most_common(1)[0][0] if rec["mfg_name_counts"] else ""
        mfg_pn = rec["mfg_pn_counts"].most_common(1)[0][0] if rec["mfg_pn_counts"] else ""
        commodity = rec["commodity_counts"].most_common(1)[0][0] if rec["commodity_counts"] else ""
        uoms_raw = rec["uom_counts"].most_common()
        uom_raw = uoms_raw[0][0] if uoms_raw else ""
        # Canonicalize before deciding "mixed" — Each/EA/EACH all map to EA so
        # they shouldn't trip the mixed-UOM warning.
        canon_counts = Counter()
        for u, n in uoms_raw:
            cu = canon_uom(u)
            if cu:
                canon_counts[cu] += n
        canon_top = canon_counts.most_common()
        uom = canon_top[0][0] if canon_top else canon_uom(uom_raw)
        uom_mixed = len(canon_top) > 1
        # Description-pattern flags — "freight", "obsolete", "generic", etc.
        desc_flags = description_pattern_flags(rec["description"])

        # Last unit price = EXACT unit price of the most recent priced order line.
        # No median, no smoothing — RFQ reporting must be to-the-penny.
        # Tiebreak when multiple lines share the latest date: highest PO #
        # (later transaction), then iteration order.
        priced_lines = [l for l in rec["po_lines"] if l[2] is not None]
        if priced_lines:
            latest = max(priced_lines, key=lambda l: (l[0] or "", l[4] or ""))
            current_price = latest[2]
        else:
            current_price = rec["last_unit_price"]

        # Window spends restated at the current (representative) price so the
        # per-row table math is internally consistent: qty × price = total.
        # Per ryan: "5 bananas × $5 each = $25". KPI totals at the top stay
        # as historical actuals — those are computed from rec["spend_*"].
        out_items.append({
            "key": rec["key"],
            "item_num": rec["item_num"],
            "eam_pn": rec["eam_pn"],
            "part_number": rec["part_number"],
            "description": rec["description"],
            "mfg_name": mfg_name,
            "mfg_pn": mfg_pn,
            "commodity": commodity,
            "uom": uom,
            "uom_raw": uom_raw,
            "uom_mixed": uom_mixed,
            "uom_distinct": [u for u, _ in canon_top],
            "desc_flags": desc_flags,
            "po_count": len(rec["po_set"]),
            "qty_12mo": rec["qty_12mo"],
            "qty_24mo": rec["qty_24mo"],
            "qty_36mo": rec["qty_36mo"],
            "qty_all": rec["qty_all"],
            # Spend columns = qty × current price (internally consistent, RFQ-relevant)
            "spend_12mo": (rec["qty_12mo"] or 0) * (current_price or 0),
            "spend_24mo": (rec["qty_24mo"] or 0) * (current_price or 0),
            "spend_36mo": (rec["qty_36mo"] or 0) * (current_price or 0),
            "spend_all":  (rec["qty_all"]  or 0) * (current_price or 0),
            # Historical actuals retained for KPI totals + reference
            "spend_12mo_actual": rec["spend_12mo"],
            "spend_24mo_actual": rec["spend_24mo"],
            "spend_36mo_actual": rec["spend_36mo"],
            "spend_all_actual":  rec["spend_all"],
            "last_unit_price": current_price,
            "first_order": rec["first_order"].date().isoformat() if rec["first_order"] else None,
            "last_order": rec["last_order_dt"].date().isoformat() if rec["last_order_dt"] else None,
            # Default include: any qty in the 24-month window
            "included": rec["qty_24mo"] > 0,
        })

    # KPIs — use HISTORICAL actuals (spend_X_actual). Per-row table shows
    # qty × current_price, but file-level totals reconcile to source.
    items_24mo = sum(1 for it in out_items if it["qty_24mo"] > 0)
    items_12mo = sum(1 for it in out_items if it["qty_12mo"] > 0)
    items_36mo = sum(1 for it in out_items if it["qty_36mo"] > 0)
    total_spend = sum(it["spend_all_actual"] for it in out_items)
    spend_12mo = sum(it["spend_12mo_actual"] for it in out_items)
    spend_24mo = sum(it["spend_24mo_actual"] for it in out_items)
    spend_36mo = sum(it["spend_36mo_actual"] for it in out_items)
    all_pos = set()
    for rec in items.values():
        all_pos |= rec["po_set"]

    first_dt = min((rec["first_order"] for rec in items.values() if rec["first_order"]), default=None)
    last_dt = max((rec["last_order_dt"] for rec in items.values() if rec["last_order_dt"]), default=None)
    years_span = (last_dt - first_dt).days / 365.25 if (first_dt and last_dt) else 0.0

    kpis = {
        "item_count": len(out_items),
        "total_spend": total_spend,
        "po_count": len(all_pos),
        "line_count": line_count,
        "first_order": first_dt.date().isoformat() if first_dt else "—",
        "last_order": last_dt.date().isoformat() if last_dt else "—",
        "years_span": years_span,
        "spend_12mo": spend_12mo,
        "spend_24mo": spend_24mo,
        "spend_36mo": spend_36mo,
        "items_12mo": items_12mo,
        "items_24mo": items_24mo,
        "items_36mo": items_36mo,
    }

    annual_out = sorted(
        [{"year": y, "spend": v["spend"], "qty": v["qty"], "item_count": len(v["items"])} for y, v in annual.items()],
        key=lambda d: d["year"],
    )

    # Score each item + compute file-level difficulty rating
    _STATE["data_anchor_date"] = now.date().isoformat() if now else None
    score_items_in_place(out_items, anchor_date=now)
    add_demand_concern_flags(out_items, anchor_date=now)
    difficulty = compute_difficulty_rating(out_items, kpis)
    # Persist difficulty as a timestamped snapshot for period-end reporting
    record_difficulty_snapshot(difficulty)

    # Persist for the xlsx generator + drill-down modal
    _STATE["items"] = out_items
    _STATE["kpis"] = kpis
    _STATE["annual_spend"] = annual_out
    _STATE["supplier_name"] = supplier_counter.most_common(1)[0][0] if supplier_counter else ""
    _STATE["difficulty"] = difficulty
    # Per-item PO lines are kept Python-side (not shipped to JS in the items
    # array) to keep the JS payload small. The modal pulls them on demand.
    _STATE["po_lines_by_key"] = {k: list(rec["po_lines"]) for k, rec in items.items()}

    # If we're reloading a saved session, exclusions were restored ahead of
    # the extract — replay them now so every aggregate (last_unit_price,
    # qty_*, spend_*, KPIs) reflects the analyst's prior outlier curation.
    # Without this, a saved exclusion would change the per-item modal view
    # but the RFQ-list table, outbound xlsx, and downstream comparison
    # would all read the stale (uncleaned) numbers.
    if _STATE.get("item_exclusions"):
        apply_all_item_exclusions_to_aggregates(rebuild_kpis=True)

    log_event(
        "extract_rfq_list",
        f"{kpis['item_count']:,} items / ${kpis['total_spend']:,.0f} total / difficulty {difficulty.get('score','?')}/100 {difficulty.get('level','')}",
        related=_STATE.get("supplier_name", ""),
    )
    return {"kpis": kpis, "items": out_items, "annual_spend": annual_out, "difficulty": difficulty}


# ---------------------------------------------------------------------------
# Inclusion scoring + file difficulty rating
# ---------------------------------------------------------------------------

# Tunable scoring weights — total to 1.0
SCORE_WEIGHTS = {
    "spend":     0.40,   # 24-mo spend (capped at the threshold below)
    "recency":   0.25,   # days since last order, decays over a year
    "frequency": 0.25,   # distinct PO count
    "data":      0.10,   # data-quality (MFG + UOM)
}
# Calibration: full score at these levels
SCORE_SPEND_FULL_24MO = 1000.0     # $1k+ spend in 24mo = full spend points
SCORE_RECENCY_RECENT_DAYS = 90     # within 90d = full recency points
SCORE_RECENCY_DEAD_DAYS = 540      # 18mo+ since last order = zero recency
SCORE_FREQUENCY_FULL = 6           # 6+ POs = full frequency points

# Tier thresholds
TIER_STRONG = 70
TIER_MODERATE = 45
TIER_WEAK = 25


def score_item(it: dict, anchor_date) -> dict:
    """Score one item 0-100, return {score, tier, reasons, default_include}.

    Anchor_date should be the dataset's max order date (so 'recency' is
    measured against the data, not wall-clock — same anchor as the windows).
    """
    # Use historical actual (not the qty × current_price projection) — scoring
    # is about "did this item have meaningful real spend"
    spend_24 = it.get("spend_24mo_actual") or it.get("spend_24mo") or 0.0
    spend_score = min(1.0, spend_24 / SCORE_SPEND_FULL_24MO)

    # Recency: parse last_order ISO; days since vs anchor
    recency_score = 0.0
    days_since_last = None
    last_order_iso = it.get("last_order")
    if last_order_iso and anchor_date:
        try:
            last_dt = datetime.fromisoformat(last_order_iso)
            days_since_last = (anchor_date - last_dt).days
            if days_since_last <= SCORE_RECENCY_RECENT_DAYS:
                recency_score = 1.0
            elif days_since_last >= SCORE_RECENCY_DEAD_DAYS:
                recency_score = 0.0
            else:
                # Linear decay between recent and dead
                rng = SCORE_RECENCY_DEAD_DAYS - SCORE_RECENCY_RECENT_DAYS
                recency_score = max(0.0, 1.0 - (days_since_last - SCORE_RECENCY_RECENT_DAYS) / rng)
        except (ValueError, TypeError):
            pass

    po_count = it.get("po_count") or 0
    frequency_score = min(1.0, po_count / SCORE_FREQUENCY_FULL)

    # Data quality: MFG present + UOM consistent
    has_mfg = bool(it.get("mfg_name"))
    uom_clean = bool(it.get("uom")) and not it.get("uom_mixed")
    data_score = 0.0
    if has_mfg: data_score += 0.5
    if uom_clean: data_score += 0.5

    # Weighted sum
    total = (
        SCORE_WEIGHTS["spend"]     * spend_score +
        SCORE_WEIGHTS["recency"]   * recency_score +
        SCORE_WEIGHTS["frequency"] * frequency_score +
        SCORE_WEIGHTS["data"]      * data_score
    )
    score = round(total * 100)

    # Tier
    if score >= TIER_STRONG:
        tier = "STRONG"
    elif score >= TIER_MODERATE:
        tier = "MODERATE"
    elif score >= TIER_WEAK:
        tier = "WEAK"
    else:
        tier = "SKIP"

    # Human-readable reasons (lead with the negatives — that's why scoring exists)
    reasons = []
    if spend_24 < 100:
        reasons.append(f"Low 24-mo spend (${spend_24:,.0f})")
    elif spend_24 < SCORE_SPEND_FULL_24MO / 2:
        reasons.append(f"Modest 24-mo spend (${spend_24:,.0f})")
    if days_since_last is None:
        reasons.append("No order date")
    elif days_since_last > SCORE_RECENCY_DEAD_DAYS:
        reasons.append(f"No order in {days_since_last // 30}mo")
    elif days_since_last > 365:
        reasons.append(f"Last order {days_since_last // 30}mo ago")
    if po_count <= 1:
        reasons.append(f"Only {po_count} PO")
    elif po_count <= 2:
        reasons.append(f"Only {po_count} POs")
    if not has_mfg:
        reasons.append("MFG blank")
    if it.get("uom_mixed"):
        reasons.append("UOM mixed")
    if not reasons and tier in ("STRONG", "MODERATE"):
        reasons.append("Healthy spend + recent + recurring")

    return {
        "score": score,
        "tier": tier,
        "reasons": reasons,
        "default_include": tier in ("STRONG", "MODERATE"),
        "days_since_last_order": days_since_last,
    }


def add_demand_concern_flags(items: list, anchor_date) -> None:
    """Flag items whose demand pattern is suspicious for RFQ. Per the brief:
    one-time spikes, dropped-off demand, surging demand, single-order
    dominance, etc. These are surfaced as `demand_flags` on each item record.
    """
    for it in items:
        flags = []
        q12 = it.get("qty_12mo") or 0
        q24 = it.get("qty_24mo") or 0
        q36 = it.get("qty_36mo") or 0
        q_all = it.get("qty_all") or 0
        po_count = it.get("po_count") or 0
        last_iso = it.get("last_order")

        # 12-mo zero but older usage exists
        if q12 == 0 and q_all > 0:
            flags.append("DORMANT_12MO")

        # Demand drop > 50% (compare 12mo annualized to prior-12-to-24mo annualized)
        prior_12_24 = q24 - q12  # months 13-24
        if prior_12_24 > 0:
            change = (q12 - prior_12_24) / prior_12_24
            if change <= -0.5:
                flags.append("DEMAND_DROP_50")
            elif change >= 0.5:
                flags.append("DEMAND_SURGE_50")

        # One order dominates >80% of usage
        if po_count == 1 and q_all > 0:
            flags.append("SINGLE_ORDER")
        elif po_count > 0 and q_all > 0:
            # Approximate: if total qty / po_count is wildly unbalanced
            avg_per_po = q_all / po_count
            # If the 24mo qty is concentrated (one PO at 80%+ of 24mo) — heuristic
            # only available with PO line detail, deferred
            pass

        # Last order > 12 months ago — already in scoring as recency penalty,
        # but we surface it explicitly here too
        if last_iso and anchor_date:
            try:
                last_dt = datetime.fromisoformat(last_iso)
                days_since = (anchor_date - last_dt).days
                if days_since > 365:
                    flags.append("STALE_OVER_12MO")
            except (ValueError, TypeError):
                pass

        # Order count low overall
        if 0 < po_count <= 2:
            flags.append("FEW_ORDERS")

        it["demand_flags"] = flags


def score_items_in_place(items: list, anchor_date) -> None:
    """Mutate each item dict in-place, adding score/tier/reasons.

    Note: this does NOT override the existing `included` flag. The tier
    is exposed so the UI can render it (chip, filter), but the user
    explicitly opts in to "auto-exclude WEAK/SKIP" via a toggle (TODO).
    Avoids surprising the user with a mass-exclusion on refresh.
    """
    for it in items:
        s = score_item(it, anchor_date)
        it["score"] = s["score"]
        it["tier"] = s["tier"]
        it["score_reasons"] = s["reasons"]
        it["days_since_last_order"] = s["days_since_last_order"]
        # Conservative: keep the original `included` default (spend_24mo > 0).
        # The tier is informational; user can apply tier-based defaults later.


def compute_difficulty_rating(items: list, kpis: dict) -> dict:
    """Roll per-item scores up to a file-level difficulty rating.

    Outputs a 0-100 difficulty score (HIGHER = HARDER to RFQ — opposite of
    item score, which is item quality), a level label, and the contributing
    signals. Snapshot intended to be persisted with a timestamp so
    period-end reports can chart difficulty trending over time as more
    files run through the tool.

    Phase 2.5 will add a bid-feedback signal (% of outlier bids retroactively
    increases the difficulty after-the-fact).
    """
    if not items:
        return {"score": 0, "level": "EMPTY", "signals": {}, "summary": "no items"}

    n = len(items)
    by_tier = Counter(it.get("tier", "SKIP") for it in items)
    pct_weak_or_skip = 100.0 * (by_tier.get("WEAK", 0) + by_tier.get("SKIP", 0)) / n
    pct_strong = 100.0 * by_tier.get("STRONG", 0) / n
    pct_no_mfg = 100.0 * sum(1 for it in items if not it.get("mfg_name")) / n
    pct_uom_mixed = 100.0 * sum(1 for it in items if it.get("uom_mixed")) / n
    pct_no_desc = 100.0 * sum(1 for it in items if not it.get("description") or len((it.get("description") or "").strip()) < 10) / n
    avg_score = sum(it.get("score", 0) for it in items) / n

    # Spend concentration — what % of items capture 80% of spend?
    sorted_spend = sorted([it.get("spend_all", 0) for it in items], reverse=True)
    total = sum(sorted_spend) or 1.0
    cum = 0.0
    items_for_80 = 0
    for i, sp in enumerate(sorted_spend):
        cum += sp
        items_for_80 = i + 1
        if cum / total >= 0.80:
            break
    pct_items_for_80 = 100.0 * items_for_80 / n

    # Difficulty score — weighted aggregate (higher = harder)
    # Each signal contributes 0-100 to the total.
    sig_weak_tail = pct_weak_or_skip                                   # already 0-100
    sig_data_quality = (pct_no_mfg + pct_uom_mixed + pct_no_desc) / 3  # 0-100
    sig_long_tail = max(0, min(100, pct_items_for_80 * 1.5))           # bias toward long-tail = harder
    sig_low_avg = max(0, 100 - avg_score)                              # inverse of average item quality

    diff_raw = (
        0.30 * sig_weak_tail +
        0.30 * sig_data_quality +
        0.20 * sig_long_tail +
        0.20 * sig_low_avg
    )
    diff_score = round(diff_raw)

    if diff_score >= 70:
        level = "VERY DIFFICULT"
    elif diff_score >= 50:
        level = "DIFFICULT"
    elif diff_score >= 30:
        level = "MODERATE"
    else:
        level = "EASY"

    # Build a one-line summary the UI can show under the label
    summary_parts = []
    if pct_weak_or_skip > 40:
        summary_parts.append(f"{pct_weak_or_skip:.0f}% items in WEAK/SKIP tier")
    if pct_no_mfg > 25:
        summary_parts.append(f"{pct_no_mfg:.0f}% missing MFG")
    if pct_no_desc > 15:
        summary_parts.append(f"{pct_no_desc:.0f}% generic/short descriptions")
    if pct_items_for_80 > 25:
        summary_parts.append(f"long tail — top {pct_items_for_80:.0f}% of items hold 80% of spend")
    if not summary_parts:
        summary_parts.append("clean data, healthy spend distribution")

    return {
        "score": diff_score,
        "level": level,
        "summary": " · ".join(summary_parts),
        "snapshot_at": datetime.now().isoformat(),
        "signals": {
            "n_items": n,
            "by_tier": dict(by_tier),
            "pct_weak_or_skip": round(pct_weak_or_skip, 1),
            "pct_strong": round(pct_strong, 1),
            "pct_no_mfg": round(pct_no_mfg, 1),
            "pct_uom_mixed": round(pct_uom_mixed, 1),
            "pct_no_desc": round(pct_no_desc, 1),
            "items_for_80pct_spend": items_for_80,
            "pct_items_for_80": round(pct_items_for_80, 1),
            "avg_score": round(avg_score, 1),
        },
    }


# ---------------------------------------------------------------------------
# Per-item history (drill-down modal)
# ---------------------------------------------------------------------------

def _linear_fit(xs, ys):
    """Simple least-squares linear regression. Returns (slope, intercept, r2).
    No numpy — we're in Pyodide and want to keep the dep surface minimal."""
    n = len(xs)
    if n < 2:
        return (0.0, ys[0] if ys else 0.0, 0.0)
    mean_x = sum(xs) / n
    mean_y = sum(ys) / n
    num = 0.0
    den = 0.0
    for x, y in zip(xs, ys):
        num += (x - mean_x) * (y - mean_y)
        den += (x - mean_x) ** 2
    slope = num / den if den else 0.0
    intercept = mean_y - slope * mean_x
    # R-squared
    ss_tot = sum((y - mean_y) ** 2 for y in ys)
    ss_res = 0.0
    for x, y in zip(xs, ys):
        pred = slope * x + intercept
        ss_res += (y - pred) ** 2
    r2 = 1.0 - (ss_res / ss_tot) if ss_tot else 0.0
    return (slope, intercept, r2)


def _median(values):
    if not values:
        return None
    s = sorted(values)
    mid = len(s) // 2
    return (s[mid] + s[~mid]) / 2.0


def get_item_history(item_num: str) -> dict:
    """Return the per-item drill-down payload for the modal.

    Builds the analytical view that backs the per-item history modal:
      * Every PO line for this item, ascending by date, each tagged with its
        canonical 0-based ``line_idx`` and an ``excluded`` flag pulled from
        ``_STATE["item_exclusions"][item_num]``.
      * A linear-trend fit (slope / intercept / R²) computed on the
        NON-EXCLUDED priced lines only, so analysts can untick obvious
        outliers in the modal table and watch the trend redraw cleanly.
      * An "expected price today" extrapolation of that cleaned trend at the
        dataset's anchor date — the "we last ordered 11 months ago at $12.40
        / trend says ~$13.10 today" reference Ryan asked for.
      * 90-day-median spike detection (same recency-tiebreak rule as the
        table's LAST $/ea) — also computed against the cleaned line set.
      * A per-supplier ``bids`` overlay listing every priced quote we have
        for this item across all loaded supplier bid files. Each bid is
        scored for distance from the cleaned trend's expected-today value:
            - ``ratio`` = bid_price / expected_today  (1.0 = on the line)
            - ``possible_typo`` = True when the bid is ≥60% below the trend
              (ratio ≤ 0.4) — the "way below the line" pattern Ryan flagged
              as the canonical typo signature.

    The shape is JSON-safe (numbers, strings, bools) so the JS renderer can
    consume it via runPythonAsync + json.dumps.

    Args:
        item_num: Display item number (the same field shown in the RFQ list
            table). Internal lookup key is ``norm_pn(item_num)``.

    Returns:
        dict with: item_num, description, mfg_name, mfg_pn, uom, uom_mixed,
        po_lines (each augmented with line_idx + excluded), summary, trend,
        bids (overlay markers), exclusions (the in-effect index list).
        Or ``{"error": "..."}`` if no history exists.

    Gotchas:
        * ``line_idx`` is into the ASCENDING-date-sorted list. The JS table
          renders newest-first but must pass back this canonical index.
        * If the user excludes EVERY priced line, trend fields fall back to
          None and the chart shows raw points only (no trend line).
        * ``latest_unit_price`` here is the cleaned-set's latest priced line.
          The to-the-penny LAST $/ea used by the RFQ-list table elsewhere is
          unrelated — Ryan's "no rounding/medians" rule is for that table,
          not for this analytical modal.
    """
    if not item_num:
        return {"error": "item_num required"}
    key = norm_pn(item_num)
    po_lines = _STATE.get("po_lines_by_key", {}).get(key, [])
    if not po_lines:
        return {"error": f"no history for {item_num}"}

    # Find the matching item record for context
    items = _STATE.get("items", [])
    item = next((it for it in items if it["item_num"] == item_num), None)

    # Sort by date ascending — this is the canonical order. line_idx values
    # in _STATE["item_exclusions"] index into THIS list.
    sorted_lines = sorted(po_lines, key=lambda r: r[0] or "")
    excluded_set = set(
        _STATE.get("item_exclusions", {}).get(item_num, []) or []
    )
    line_dicts = [
        {
            "line_idx": idx,
            "date": d,
            "qty": q,
            "unit_price": p,
            "line_total": lt,
            "po": po,
            "uom": uom,
            "excluded": idx in excluded_set,
        }
        for idx, (d, q, p, lt, po, uom) in enumerate(sorted_lines)
    ]

    # Linear regression on (days_since_first_order, unit_price) — CLEANED
    # set only. first_dt anchors x=0 to the earliest INCLUDED priced line so
    # the slope+intercept render correctly when the user excludes the very
    # first orders as outliers.
    first_dt = None
    xs, ys = [], []
    cleaned_line_dicts = [ln for ln in line_dicts if not ln["excluded"]]
    for ln in cleaned_line_dicts:
        if not ln["date"] or ln["unit_price"] is None:
            continue
        try:
            d = datetime.fromisoformat(ln["date"])
        except ValueError:
            continue
        if first_dt is None:
            first_dt = d
        xs.append((d - first_dt).days)
        ys.append(float(ln["unit_price"]))

    if len(xs) >= 2:
        slope, intercept, r2 = _linear_fit(xs, ys)
    else:
        # No fittable trend — surface None so the chart skips the line and
        # the callout falls back to the confidence-reason text.
        slope, intercept, r2 = (None, None, None)
    last_x = xs[-1] if xs else 0
    last_price = ys[-1] if ys else None

    # Expected price at the dataset anchor date — uses the cleaned-trend fit.
    expected_today = None
    days_since_last = None
    anchor = _STATE.get("data_anchor_date")
    if anchor and first_dt and slope is not None and intercept is not None:
        try:
            anchor_dt = datetime.fromisoformat(anchor)
            anchor_x = (anchor_dt - first_dt).days
            expected_today = slope * anchor_x + intercept
            days_since_last = anchor_x - last_x
        except ValueError:
            pass

    # Confidence label based on R² + sample size of the CLEANED set.
    excl_suffix = f" (after excluding {len(excluded_set)})" if excluded_set else ""
    if len(xs) < 3:
        confidence = "low"
        confidence_reason = f"only {len(xs)} priced order line(s){excl_suffix}"
    elif r2 is None:
        confidence = "low"
        confidence_reason = "trend not fittable on cleaned set"
    elif r2 < 0.2:
        confidence = "low"
        confidence_reason = f"R² {r2:.2f} — prices noisy, trend may not be meaningful"
    elif r2 < 0.6:
        confidence = "medium"
        confidence_reason = f"R² {r2:.2f}"
    else:
        confidence = "high"
        confidence_reason = f"R² {r2:.2f}"

    # 90-day median price (analytical reference — separate from the to-the-penny
    # last_unit_price). Used as the "is the latest line a price spike" baseline
    # on the chart. Falls back to median of last 10 lines if <3 within 90 days.
    # Computed on cleaned lines only so a single bad outlier doesn't dominate.
    median_90d = None
    median_window_label = None
    if cleaned_line_dicts:
        try:
            recent_dt = datetime.fromisoformat(cleaned_line_dicts[-1]["date"])
            cutoff_90 = recent_dt - timedelta(days=90)
            recent_prices = [
                ln["unit_price"] for ln in cleaned_line_dicts
                if ln["date"] and ln["unit_price"] is not None
                and datetime.fromisoformat(ln["date"]) >= cutoff_90
            ]
            if len(recent_prices) >= 3:
                median_90d = _median(recent_prices)
                median_window_label = f"90-day median ({len(recent_prices)} lines)"
            else:
                fallback = sorted(
                    cleaned_line_dicts, key=lambda x: x["date"], reverse=True
                )[:10]
                fp = [ln["unit_price"] for ln in fallback if ln["unit_price"] is not None]
                if fp:
                    median_90d = _median(fp)
                    median_window_label = f"median of last {len(fp)} lines"
        except (ValueError, TypeError):
            pass

    # Spike detection — compare latest CLEANED unit price to median_90d.
    # Selection rule mirrors the table's LAST $/ea: max by (date, po) so
    # ties on the most recent date pick the highest PO # (later transaction).
    priced_cleaned = [
        (sorted_lines[idx][0], sorted_lines[idx][4], sorted_lines[idx][2])
        for idx in range(len(sorted_lines))
        if idx not in excluded_set and sorted_lines[idx][2] is not None
    ]
    if priced_cleaned:
        latest_tuple = max(priced_cleaned, key=lambda x: (x[0] or "", x[1] or ""))
        latest_unit_price = latest_tuple[2]
    else:
        latest_unit_price = None
    spike = None
    if latest_unit_price is not None and median_90d and median_90d > 0:
        ratio = latest_unit_price / median_90d
        pct_diff = (ratio - 1.0) * 100.0
        if ratio >= 1.5 or ratio <= 0.5:
            direction = "above" if ratio > 1 else "below"
            spike = {
                "is_spike": True,
                "ratio": ratio,
                "pct_diff": pct_diff,
                "message": f"Latest price ${latest_unit_price:,.2f} is {abs(pct_diff):.0f}% {direction} {median_window_label} (${median_90d:,.2f})",
            }
        else:
            spike = {
                "is_spike": False,
                "ratio": ratio,
                "pct_diff": pct_diff,
                "message": f"Latest ${latest_unit_price:,.2f} vs {median_window_label} ${median_90d:,.2f} ({pct_diff:+.0f}%)",
            }

    # ------------------------------------------------------------------
    # Supplier-bid overlay — every priced quote we have for this rfq_key,
    # scored against the cleaned trend's expected-today price. Powers the
    # right-edge horizontal markers on the per-item chart and the
    # POSSIBLE_TYPO flag for "way below the line" bids.
    # ------------------------------------------------------------------
    locks = _STATE.get("item_locks", {}) or {}
    locked_record = locks.get(item_num)
    locked_supplier = locked_record.get("supplier") if locked_record else None
    bids_overlay = _build_item_bid_overlay(
        key, expected_today, latest_unit_price, locked_supplier
    )

    return {
        "item_num": item_num,
        "description": item["description"] if item else "",
        "mfg_name": item["mfg_name"] if item else "",
        "mfg_pn": item["mfg_pn"] if item else "",
        "uom": item["uom"] if item else "",
        "uom_mixed": item["uom_mixed"] if item else False,
        "po_lines": line_dicts,
        "n_total_lines": len(line_dicts),
        "n_excluded": sum(1 for ln in line_dicts if ln["excluded"]),
        "n_priced_after_exclusion": len(xs),
        "exclusions": sorted(excluded_set),
        "lock": locked_record,
        "summary": {
            "po_count": item["po_count"] if item else 0,
            "qty_12mo": item["qty_12mo"] if item else 0,
            "spend_12mo": item["spend_12mo"] if item else 0,
            "qty_24mo": item["qty_24mo"] if item else 0,
            "spend_24mo": item["spend_24mo"] if item else 0,
            "qty_36mo": item["qty_36mo"] if item else 0,
            "spend_36mo": item["spend_36mo"] if item else 0,
            "qty_all": item["qty_all"] if item else 0,
            "spend_all": item["spend_all"] if item else 0,
            "last_order": item["last_order"] if item else None,
            "first_order": item["first_order"] if item else None,
            "last_unit_price": last_price,
        },
        "trend": {
            "slope_per_day": slope,
            "intercept": intercept,
            "r2": r2,
            "confidence": confidence,
            "confidence_reason": confidence_reason,
            "expected_today": expected_today,
            "days_since_last_order": days_since_last,
            "anchor_date": anchor,
            "latest_unit_price": latest_unit_price,
            "median_90d": median_90d,
            "median_window_label": median_window_label,
            "spike": spike,
        },
        "bids": bids_overlay,
    }


# Threshold for the POSSIBLE_TYPO flag on the per-item bid overlay. A bid
# whose ratio against expected_today (or, when no trend, against the latest
# cleaned price) is ≤ this value is flagged as "way below the line" — the
# canonical typo signature on supplier bid files.
ITEM_OVERLAY_TYPO_RATIO_MAX = 0.4


# Supplier-bid status priority for the per-item overlay's canonical pick.
# PRICED beats UOM_DISC beats SUBSTITUTE beats anything else. NEED_INFO and
# NO_BID don't carry a usable price so they're never canonical (filtered out
# upstream). Within the same status the canonical pick is the LOWEST price
# — it's the most aggressive offer the supplier put on the table for this
# item and matches what the analyst is likeliest to be asked to evaluate.
_BID_STATUS_PRIORITY = {
    "PRICED": 0,
    "UOM_DISC": 1,
    "SUBSTITUTE": 2,
}


def _build_item_bid_overlay(rfq_key: str, expected_today, latest_price,
                            locked_supplier: str = None) -> list:
    """Collect ONE canonical priced supplier bid per supplier for one item,
    tagged with a distance metric vs. the cleaned trend.

    Walks ``_STATE["bids"][supplier]["bids"]`` for every loaded supplier and
    matches on the canonical ``rfq_key`` (already normalized via norm_pn).
    Bids with status PRICED, UOM_DISC, and SUBSTITUTE that carry an
    effective_price > 0 are eligible. NO_BID and NEED_INFO are skipped.

    Why dedupe per supplier:
        Suppliers routinely return MULTIPLE lines for the same PN (qty break
        tiers, alt SKUs they offered, or literal duplicate rows in their
        template). Without dedup, a single item explodes into 20-30 chart
        markers. The comparison matrix and scenario engine already collapse
        per-(supplier, rfq_key) to one bid; this overlay matches.

    Canonical pick (per supplier):
        1. Lowest ``_BID_STATUS_PRIORITY`` (PRICED beats UOM_DISC beats SUBSTITUTE)
        2. Within ties, the lowest effective_price (the supplier's most
           aggressive offer — closest to what the analyst will evaluate).
        Each canonical bid carries:
          * ``alt_quotes`` — the OTHER prices this supplier returned for
            the same item (sorted ascending; status preserved)
          * ``n_alt_quotes`` — count of those alternates
        so the UI can surface "+N alt quotes" without flooding the screen.

    Distance scoring against the reference price:
      * If a non-None ``expected_today`` is available, use it as the
        reference and report ``ratio = price / expected_today`` plus
        ``pct_diff = (ratio - 1) * 100``.
      * Otherwise fall back to ``latest_price`` (no trend means no
        extrapolation, so ``ratio`` is informational only).
      * ``possible_typo`` is True iff a reference exists AND
        ``ratio <= ITEM_OVERLAY_TYPO_RATIO_MAX`` (≥60% below the line).

    Lock awareness:
      * Each entry carries an ``is_locked`` flag set when its supplier
        equals ``locked_supplier``. The UI renders the locked bid
        distinctly and disables cross-supplier lock buttons.

    Returns a list sorted ascending by canonical price.
    """
    out = []
    bids_by_supplier = _STATE.get("bids", {}) or {}
    if not rfq_key:
        return out
    reference = expected_today if expected_today is not None else latest_price
    annotations = _STATE.get("uom_annotations") or {}
    for supplier, parsed in bids_by_supplier.items():
        if not isinstance(parsed, dict):
            continue
        # Collect every eligible bid record for this supplier + item.
        candidates = []
        for b in parsed.get("bids", []) or []:
            if b.get("rfq_key") != rfq_key:
                continue
            price = b.get("effective_price")
            if price is None or price <= 0:
                continue
            status = b.get("status") or ""
            if status not in _BID_STATUS_PRIORITY:
                continue
            candidates.append((b, status, float(price)))
        if not candidates:
            continue
        # Sort: status priority asc, then price asc. First entry is canonical.
        candidates.sort(key=lambda c: (_BID_STATUS_PRIORITY[c[1]], c[2]))
        canonical_b, canonical_status, canonical_price = candidates[0]
        # The other quotes — surface as alt_quotes for the UI to show. Trim
        # to a reasonable cap so a pathological 200-row supplier file doesn't
        # blow the JSON payload.
        alt_quotes = []
        for b2, st2, pr2 in candidates[1:21]:  # cap at 20 alts
            alt_quotes.append({
                "price": pr2,
                "status": st2,
                "qty": b2.get("qty"),
                "uom": b2.get("uom") or "",
                "alt_part": b2.get("alt_part") or "",
                "notes": (b2.get("notes") or "")[:120],
            })
        n_alt = len(candidates) - 1

        ratio = (canonical_price / reference) if (reference and reference > 0) else None
        pct_diff = ((ratio - 1.0) * 100.0) if ratio is not None else None
        possible_typo = (ratio is not None and ratio <= ITEM_OVERLAY_TYPO_RATIO_MAX)
        # In-effect UOM annotation for this (item, supplier) — drives the
        # 🚩 flag-UOM button state in the modal. None means unflagged.
        ann = annotations.get(_annotation_key(rfq_key, supplier))
        uom_status = ann.get("status") if ann else None
        out.append({
            "supplier": supplier,
            "price": canonical_price,
            "qty": canonical_b.get("qty"),
            "uom": canonical_b.get("uom") or "",
            "status": canonical_status,
            "alt_part": canonical_b.get("alt_part") or "",
            "notes": (canonical_b.get("notes") or "")[:200],
            "ratio": ratio,
            "pct_diff": pct_diff,
            "possible_typo": possible_typo,
            "reference": ("trend" if expected_today is not None
                          else ("latest" if latest_price is not None else None)),
            "is_locked": (locked_supplier is not None
                          and supplier == locked_supplier),
            "uom_status": uom_status,                    # None / "needs_review" / "resolved" / "skipped"
            "uom_factor": ann.get("factor") if ann else None,
            "uom_note": ann.get("note") if ann else "",
            "n_alt_quotes": n_alt,
            "alt_quotes": alt_quotes,
        })
    out.sort(key=lambda r: r["price"])
    return out


def set_item_lock(item_num: str, supplier: str, reason: str = "") -> dict:
    """Pin one item's award to a specific supplier across every scenario.

    Locks express analyst intent: 'I have visually audited Supplier X's bid
    for this item and confirmed it. Award it to them regardless of who's
    cheaper — the cheap one might be a typo I don't trust.'

    Stored as ``_STATE["item_locks"][item_num] = {"supplier", "reason",
    "locked_at"}``. Evaluated in ``_evaluate_scenario`` AFTER explicit
    scenario overrides but BEFORE strategy logic, so the same lock applies
    across every saved scenario for this RFQ event.

    A lock for an item whose locked supplier never bid (or bid NO_BID) is
    persisted but does NOT force an award — strategy logic falls through.
    The ``decision_basis`` notes the lock as a warning when this happens.

    Args:
        item_num: Display item number (the same key the per-item modal opens with).
        supplier: Exact supplier name (must match a key in _STATE["bids"]
            for the lock to actually steer scenario evaluation).
        reason: Free-form analyst note (e.g., "audited 4/29 — Grainger
            confirmed UOM is each, not box").

    Returns:
        ``{"item_num", "supplier", "reason", "locked_at"}`` — the canonical
        record after the write.
    """
    if not item_num:
        return {"error": "item_num required"}
    if not supplier:
        return {"error": "supplier required (use clear_item_lock to remove)"}
    locks = _STATE.setdefault("item_locks", {})
    record = {
        "supplier": supplier,
        "reason": reason or "",
        "locked_at": datetime.now().isoformat(),
    }
    locks[item_num] = record
    log_event(
        "item_lock_set",
        f"{item_num} → {supplier}" + (f" ({reason})" if reason else ""),
        item_num,
    )
    return {"item_num": item_num, **record}


def clear_item_lock(item_num: str) -> dict:
    """Remove a previously-set per-item supplier lock.

    Returns ``{"item_num", "cleared": True/False}`` — False when no lock
    existed for this item (no-op).
    """
    if not item_num:
        return {"error": "item_num required"}
    locks = _STATE.setdefault("item_locks", {})
    had = item_num in locks
    if had:
        prev = locks.pop(item_num)
        log_event(
            "item_lock_clear",
            f"{item_num} (was → {prev.get('supplier')})",
            item_num,
        )
    return {"item_num": item_num, "cleared": had}


def list_item_locks() -> dict:
    """Return all in-effect item locks: ``{item_num: lock_record}``."""
    return dict(_STATE.get("item_locks", {}) or {})


def set_item_exclusions(item_num: str, excluded_indices) -> dict:
    """Persist the user's outlier-line selections for one item — and propagate.

    The exclusion isn't just a per-modal view filter: ``last_unit_price``,
    every windowed ``qty_*`` / ``spend_*`` / ``spend_*_actual``, ``po_count``,
    ``first_order``, ``last_order``, and the ``included`` default for this
    item are re-derived from the cleaned ``po_lines_by_key`` set so the
    RFQ-list table, outbound RFQ xlsx, and the comparison-stage historical
    baseline + scenario savings all read the cleaned numbers. File-level
    KPIs are rebuilt too so the headline tiles stay reconciled.

    This is the load-bearing piece of the "no rounding in RFQ math" rule:
    the rule was always about not blurring REAL prices (medians, smoothing,
    averaging across periods); analyst-confirmed outlier removal is the
    opposite — it's removing data errors so the to-the-penny number is
    actually meaningful. Each exclusion change writes to the audit log so
    the decision trail is preserved (Decision Log will pick this up too).

    Stores the cleaned ``excluded_indices`` list under
    ``_STATE["item_exclusions"][item_num]``. Empty lists clear the entry so
    a fully-uncrossed item doesn't leave dead state in saves.

    Args:
        item_num: Display item number (matches the key used by the table).
        excluded_indices: Iterable of 0-based indices into the
            ascending-date-sorted po_lines for this item. Non-int entries
            and duplicates are coerced/dedup'd.

    Returns:
        ``{"item_num", "n_excluded", "exclusions", "item": <updated record>,
        "kpis": <rebuilt headline KPIs>}`` — the JS uses ``item`` to patch
        ``_rfqResult.items`` in place and re-render the RFQ table, and
        ``kpis`` to refresh the headline tiles.
    """
    if not item_num:
        return {"error": "item_num required"}
    cleaned = []
    seen = set()
    for v in (excluded_indices or []):
        try:
            iv = int(v)
        except (TypeError, ValueError):
            continue
        if iv < 0 or iv in seen:
            continue
        seen.add(iv)
        cleaned.append(iv)
    cleaned.sort()
    excl_map = _STATE.setdefault("item_exclusions", {})
    prior = list(excl_map.get(item_num, []))
    prior_set = set(prior)
    cleaned_set = set(cleaned)
    newly_excluded = sorted(cleaned_set - prior_set)
    newly_unexcluded = sorted(prior_set - cleaned_set)
    if cleaned:
        excl_map[item_num] = cleaned
    else:
        excl_map.pop(item_num, None)

    # Append to the exclusion review log BEFORE we recompute aggregates —
    # the "before" metrics (median + avg of the other priced lines) need
    # the pre-change snapshot to be meaningful.
    if newly_excluded:
        _append_exclusion_log_entries(item_num, newly_excluded, prior_set)
    if newly_unexcluded:
        _append_unexclusion_log_entries(item_num, newly_unexcluded)

    # Propagate to the canonical item aggregates + the file-level KPIs.
    updated_item = _recompute_item_aggregates_for(item_num)
    rebuilt_kpis = _rebuild_kpis_from_items()

    log_event(
        "item_exclusions_set",
        f"{item_num}: {len(cleaned)} line(s) excluded — last_unit_price now ${updated_item.get('last_unit_price') or 0:.2f}"
        if updated_item else f"{item_num}: {len(cleaned)} line(s)",
        item_num,
    )
    return {
        "item_num": item_num,
        "n_excluded": len(cleaned),
        "exclusions": cleaned,
        "item": updated_item,
        "kpis": rebuilt_kpis,
    }


def _append_exclusion_log_entries(item_num: str, newly_excluded_indices: list,
                                   prior_excluded_set: set) -> None:
    """Snapshot each newly-excluded line + the pre-exclusion baseline.

    Called from ``set_item_exclusions`` BEFORE the aggregates are recomputed
    so the "before" median + avg are computed from the line set as it
    looked just before this exclusion was applied. Entry shape is shared
    across the auto-rfq-banana / supplier-pricing / tariff-impact data-
    quality logs — a downstream tool can `concat` all three apps' logs
    into one audit packet for review (cross-app master record).

    Per-entry fields:
        timestamp        ISO datetime when the analyst unticked it
        app_source       "auto-rfq-banana"  (constant — for cross-app concat)
        event_type       "exclusion"        (constant; "unexclusion" path
                                              is in _append_unexclusion_log_entries)
        rfq_id           the in-flight RFQ session id (the save manager's
                          stable per-session id), for grouping
        supplier_name    incumbent supplier on this multi-year export
        item_num         the display item number
        description, mfg_name, mfg_pn, uom — full item context
        line_idx         0-based position in the ASCENDING-date-sorted
                          po_lines for the item
        line_date        the excluded order date
        line_qty         the excluded order qty
        line_unit_price  the excluded $/ea (the suspicious one)
        line_total       qty × unit_price for that line
        line_po          PO# carrying this line
        line_uom         per-line UOM (may differ from item-canonical UOM)
        median_before    median of OTHER priced lines for this item BEFORE
                          this exclusion was applied (incl. older exclusions)
        avg_before       same but mean
        n_other_lines_before  count of those other priced lines
        ratio_to_median  unit_price / median_before (1.0 = on the line;
                          values >1 = above; values <1 = below)
        pct_diff_median  (ratio - 1) * 100, signed
        pct_diff_avg     same vs avg_before, signed
        notes            free-form (default empty; UI may expand later
                          to capture analyst rationale)
    """
    log = _STATE.setdefault("exclusion_log", [])
    items = _STATE.get("items", []) or []
    item = next((it for it in items if it.get("item_num") == item_num), None)
    description = (item.get("description") if item else "") or ""
    mfg_name    = (item.get("mfg_name")    if item else "") or ""
    mfg_pn      = (item.get("mfg_pn")      if item else "") or ""
    item_uom    = (item.get("uom")         if item else "") or ""

    key = item.get("key") if item else norm_pn(item_num)
    raw_lines = _STATE.get("po_lines_by_key", {}).get(key, [])
    if not raw_lines:
        return
    sorted_lines = sorted(raw_lines, key=lambda r: r[0] or "")

    # "Before" baseline = OTHER priced lines that weren't already excluded
    # ahead of this batch. Excludes the lines being newly excluded so the
    # baseline isn't polluted by them.
    other_priced = []
    for idx, ln in enumerate(sorted_lines):
        if idx in prior_excluded_set:
            continue
        if idx in newly_excluded_indices:
            continue
        if ln[2] is None:  # unit_price
            continue
        other_priced.append(float(ln[2]))
    if other_priced:
        median_before = _median(other_priced)
        avg_before = sum(other_priced) / len(other_priced)
    else:
        median_before = None
        avg_before = None

    rfq_id = _STATE.get("rfq_id") or ""
    supplier_name = _STATE.get("supplier_name", "") or ""
    ts = datetime.now().isoformat()

    for idx in newly_excluded_indices:
        if idx >= len(sorted_lines):
            continue
        date_iso, qty, unit_price, line_total, po, uom = sorted_lines[idx]
        if unit_price is None:
            continue
        ratio = (unit_price / median_before) if (median_before and median_before > 0) else None
        pct_med = ((ratio - 1.0) * 100.0) if ratio is not None else None
        pct_avg = (((unit_price / avg_before) - 1.0) * 100.0) if (avg_before and avg_before > 0) else None
        log.append({
            "timestamp": ts,
            "app_source": "auto-rfq-banana",
            "event_type": "exclusion",
            "rfq_id": rfq_id,
            "supplier_name": supplier_name,
            "item_num": item_num,
            "description": description,
            "mfg_name": mfg_name,
            "mfg_pn": mfg_pn,
            "uom": item_uom,
            "line_idx": idx,
            "line_date": date_iso or "",
            "line_qty": qty,
            "line_unit_price": float(unit_price),
            "line_total": float(line_total) if line_total is not None else None,
            "line_po": po or "",
            "line_uom": uom or "",
            "median_before": median_before,
            "avg_before": avg_before,
            "n_other_lines_before": len(other_priced),
            "ratio_to_median": ratio,
            "pct_diff_median": pct_med,
            "pct_diff_avg": pct_avg,
            "notes": "",
        })


def _append_unexclusion_log_entries(item_num: str, newly_unexcluded_indices: list) -> None:
    """Mirror of _append_exclusion_log_entries for re-included lines.

    Records that the analyst changed their mind about a previously-excluded
    line. Same schema as exclusion entries (so downstream filters can group
    by event_type) but only fills the line snapshot — "before" baselines
    aren't meaningful for an unexclusion (the line is going BACK into the
    set, not coming OUT of it).
    """
    log = _STATE.setdefault("exclusion_log", [])
    items = _STATE.get("items", []) or []
    item = next((it for it in items if it.get("item_num") == item_num), None)
    description = (item.get("description") if item else "") or ""
    mfg_name    = (item.get("mfg_name")    if item else "") or ""
    mfg_pn      = (item.get("mfg_pn")      if item else "") or ""
    item_uom    = (item.get("uom")         if item else "") or ""
    key = item.get("key") if item else norm_pn(item_num)
    raw_lines = _STATE.get("po_lines_by_key", {}).get(key, [])
    if not raw_lines:
        return
    sorted_lines = sorted(raw_lines, key=lambda r: r[0] or "")
    rfq_id = _STATE.get("rfq_id") or ""
    supplier_name = _STATE.get("supplier_name", "") or ""
    ts = datetime.now().isoformat()
    for idx in newly_unexcluded_indices:
        if idx >= len(sorted_lines):
            continue
        date_iso, qty, unit_price, line_total, po, uom = sorted_lines[idx]
        log.append({
            "timestamp": ts,
            "app_source": "auto-rfq-banana",
            "event_type": "unexclusion",
            "rfq_id": rfq_id,
            "supplier_name": supplier_name,
            "item_num": item_num,
            "description": description,
            "mfg_name": mfg_name,
            "mfg_pn": mfg_pn,
            "uom": item_uom,
            "line_idx": idx,
            "line_date": date_iso or "",
            "line_qty": qty,
            "line_unit_price": float(unit_price) if unit_price is not None else None,
            "line_total": float(line_total) if line_total is not None else None,
            "line_po": po or "",
            "line_uom": uom or "",
            "median_before": None,
            "avg_before": None,
            "n_other_lines_before": None,
            "ratio_to_median": None,
            "pct_diff_median": None,
            "pct_diff_avg": None,
            "notes": "re-included after prior exclusion",
        })


def flag_uom_suspected(item_num: str, supplier: str, note: str = "") -> dict:
    """Flag (item, supplier) as a suspected UOM mismatch from the per-item modal.

    Workflow handoff: the analyst spots an outlier in the per-item history
    chart, excludes it via the USE checkbox, and notices that one or two
    supplier bids only line up with the cleaned trend if their UOM is off
    (e.g., supplier quoted "per box of 10" against an "each" history).
    Hitting the 🚩 button next to that supplier's bid card calls into here:

      1. Writes a `needs_review` UOM annotation via `set_uom_annotation`
         with factor=None — same place the UOM Resolution Queue surfaces
         needs-review rows in step 4. So when the analyst opens that
         queue, this supplier is already pre-filled at the top.
      2. Appends an entry to the master data-quality event log with
         event_type="uom_suspected" so the cross-app audit packet shows
         the lineage from "outlier excluded" → "supplier flagged" →
         (later) "UOM corrected with factor X" → "NORMALIZED savings".

    Args:
        item_num: display item number (the same key the modal opens with).
        supplier: exact supplier name (matches a key in _STATE["bids"]).
        note:     optional analyst memo. Default mentions the modal source.

    Returns:
        ``{"item_num", "supplier", "annotation": <full annotation dict>}``
        — caller refreshes the modal to render the flag-button as set.
    """
    if not item_num or not supplier:
        return {"error": "item_num and supplier required"}
    items = _STATE.get("items", []) or []
    item = next((it for it in items if it.get("item_num") == item_num), None)
    item_key = item.get("key") if item else norm_pn(item_num)
    hist_uom = (item.get("uom") if item else "") or ""
    bid_uom = ""
    parsed = (_STATE.get("bids", {}) or {}).get(supplier) or {}
    for b in parsed.get("bids", []) or []:
        if b.get("rfq_key") == item_key:
            bid_uom = b.get("uom") or ""
            break
    note_full = note or "flagged from per-item modal — suspect UOM mismatch with cleaned history"
    annotation = set_uom_annotation(
        item_key=item_key,
        supplier=supplier,
        factor=None,
        hist_uom=hist_uom,
        bid_uom=bid_uom,
        note=note_full,
        status="needs_review",
        set_by="per_item_modal",
        direction="auto_detect",
    )
    # Master data-quality log entry — cross-app audit packet sees it.
    log = _STATE.setdefault("exclusion_log", [])
    log.append({
        "timestamp": datetime.now().isoformat(),
        "app_source": "auto-rfq-banana",
        "event_type": "uom_suspected",
        "rfq_id": _STATE.get("rfq_id") or "",
        "supplier_name": supplier,
        "item_num": item_num,
        "description": (item.get("description") if item else "") or "",
        "mfg_name": (item.get("mfg_name") if item else "") or "",
        "mfg_pn": (item.get("mfg_pn") if item else "") or "",
        "uom": hist_uom,
        "line_idx": None,
        "line_date": "",
        "line_qty": None,
        "line_unit_price": None,
        "line_total": None,
        "line_po": "",
        "line_uom": bid_uom,
        "median_before": None,
        "avg_before": None,
        "n_other_lines_before": None,
        "ratio_to_median": None,
        "pct_diff_median": None,
        "pct_diff_avg": None,
        "notes": note_full,
    })
    return {"item_num": item_num, "supplier": supplier, "annotation": annotation}


def clear_uom_suspected_flag(item_num: str, supplier: str) -> dict:
    """Clear the per-item-modal-set UOM suspected flag for (item, supplier).

    Calls `clear_uom_annotation` and appends an event_type="uom_unflagged"
    entry to the master log so the trail shows the analyst changed their
    mind. Doesn't touch annotations that are status="resolved" or
    "skipped" — only the ones flagged from the modal as needs_review.
    """
    if not item_num or not supplier:
        return {"error": "item_num and supplier required"}
    items = _STATE.get("items", []) or []
    item = next((it for it in items if it.get("item_num") == item_num), None)
    item_key = item.get("key") if item else norm_pn(item_num)
    existing = get_uom_annotation(item_key, supplier) if "get_uom_annotation" in globals() else (_STATE.get("uom_annotations") or {}).get(_annotation_key(item_key, supplier))
    cleared = False
    if existing and existing.get("status") == "needs_review":
        ann_map = _STATE.get("uom_annotations") or {}
        ann_map.pop(_annotation_key(item_key, supplier), None)
        cleared = True
    log_event("uom_annotation_cleared", f"per-item modal cleared UOM flag for {supplier}", related=item_num)
    if cleared:
        _STATE.setdefault("exclusion_log", []).append({
            "timestamp": datetime.now().isoformat(),
            "app_source": "auto-rfq-banana",
            "event_type": "uom_unflagged",
            "rfq_id": _STATE.get("rfq_id") or "",
            "supplier_name": supplier,
            "item_num": item_num,
            "description": (item.get("description") if item else "") or "",
            "mfg_name": (item.get("mfg_name") if item else "") or "",
            "mfg_pn": (item.get("mfg_pn") if item else "") or "",
            "uom": (item.get("uom") if item else "") or "",
            "line_idx": None, "line_date": "", "line_qty": None,
            "line_unit_price": None, "line_total": None, "line_po": "", "line_uom": "",
            "median_before": None, "avg_before": None, "n_other_lines_before": None,
            "ratio_to_median": None, "pct_diff_median": None, "pct_diff_avg": None,
            "notes": "UOM-suspected flag cleared from per-item modal",
        })
    return {"item_num": item_num, "supplier": supplier, "cleared": cleared}


def list_exclusion_log() -> list:
    """Return the exclusion review log (newest entries first)."""
    return list(reversed(_STATE.get("exclusion_log", []) or []))


def get_exclusion_log_summary() -> dict:
    """Headline counts for the dashboard banner — drives the 'X excluded
    lines this session — review/export?' reminder shown above the
    Generate-outbound and Compare-bids actions."""
    log = _STATE.get("exclusion_log", []) or []
    n_excl = sum(1 for e in log if e.get("event_type") == "exclusion")
    n_unexcl = sum(1 for e in log if e.get("event_type") == "unexclusion")
    distinct_items = sorted({e.get("item_num") for e in log if e.get("event_type") == "exclusion" and e.get("item_num")})
    total_dollars_removed = 0.0
    for e in log:
        if e.get("event_type") != "exclusion":
            continue
        lt = e.get("line_total")
        if lt is not None:
            total_dollars_removed += float(lt)
    return {
        "n_exclusions": n_excl,
        "n_unexclusions": n_unexcl,
        "n_distinct_items": len(distinct_items),
        "total_dollars_removed": total_dollars_removed,
    }


def gen_exclusion_log_xlsx() -> bytes:
    """Build a master-record xlsx of every exclusion / unexclusion event.

    Format is intentionally cross-app: an ``app_source`` column tags every
    row so this xlsx can be `pd.concat`'d with the equivalent log from
    supplier-pricing and tariff-impact (when those apps grow the same
    feature) to produce one audit packet for a colleague / auditor to
    review across all the data-quality calls made this period.

    Banner row 1: 'INTERNAL — DATA QUALITY EVENT LOG'.
    Header row 2: every field documented in _append_exclusion_log_entries.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Exclusion Log"

    # Banner — same convention as internal-only summary xlsx
    ws.append(["INTERNAL — DATA QUALITY EVENT LOG · NEVER FORWARD"])
    ws.append([
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"App: auto-rfq-banana",
        f"Supplier: {_STATE.get('supplier_name', '')}",
        f"Anchor date: {_STATE.get('data_anchor_date', '')}",
    ])
    ws.append([])  # spacer

    headers = [
        "timestamp", "app_source", "event_type", "rfq_id",
        "supplier_name", "item_num", "description", "mfg_name", "mfg_pn", "uom",
        "line_idx", "line_date", "line_qty", "line_unit_price", "line_total",
        "line_po", "line_uom",
        "median_before", "avg_before", "n_other_lines_before",
        "ratio_to_median", "pct_diff_median", "pct_diff_avg",
        "notes",
    ]
    ws.append(headers)
    header_row_idx = ws.max_row
    for c in ws[header_row_idx]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3A2917")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Banner row formatting
    ws["A1"].font = Font(bold=True, color="FF6B6B", size=14)

    log = _STATE.get("exclusion_log", []) or []
    for e in log:
        ws.append([e.get(k) for k in headers])

    # Number formats on the numeric columns
    money_cols = {"line_unit_price": "N", "line_total": "O",
                  "median_before": "R", "avg_before": "S"}
    for fld, col in money_cols.items():
        for row in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws[f"{col}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0.0000'

    pct_cols = {"pct_diff_median": "V", "pct_diff_avg": "W"}
    for fld, col in pct_cols.items():
        for row in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws[f"{col}{row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '+0.0"%";-0.0"%";0.0"%"'

    autosize(ws, min_w=10, max_w=40)
    ws.freeze_panes = f"A{header_row_idx + 1}"

    log_event(
        "exclusion_log_export",
        f"{len(log)} entries exported to xlsx",
        related=_STATE.get("supplier_name", ""),
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _recompute_item_aggregates_for(item_num: str) -> dict:
    """Re-derive one item's aggregates from its cleaned po_lines.

    Walks ``_STATE["po_lines_by_key"][key]`` for the matching item, drops
    the indices listed in ``_STATE["item_exclusions"][item_num]`` (indices
    are positions in the ASCENDING-date-sorted line list, same convention
    as ``get_item_history``), and writes back:

        last_unit_price (current_price)  — most recent priced line in the
                                            cleaned set, max by (date, po)
        qty_12mo / qty_24mo / qty_36mo / qty_all
        spend_12mo / 24mo / 36mo / all          (qty × current_price)
        spend_12mo_actual / 24mo_actual / etc   (sum of cleaned line totals)
        po_count                                (distinct PO# count, cleaned)
        first_order, last_order                 (iso strings, cleaned)
        included                                (default = qty_24mo > 0)

    Returns the updated item dict (the same object that lives in
    ``_STATE["items"]`` — JS-safe), or ``{}`` if the item isn't found.
    """
    items = _STATE.get("items", [])
    item = next((it for it in items if it.get("item_num") == item_num), None)
    if not item:
        return {}
    key = item.get("key") or norm_pn(item_num)
    raw_lines = _STATE.get("po_lines_by_key", {}).get(key, [])
    if not raw_lines:
        return item
    sorted_lines = sorted(raw_lines, key=lambda r: r[0] or "")
    excluded = set(_STATE.get("item_exclusions", {}).get(item_num, []) or [])
    cleaned_lines = [ln for idx, ln in enumerate(sorted_lines) if idx not in excluded]

    # Window cutoffs anchored to the data, not the wall clock — same rule
    # as extract_rfq_list (the dataset's max order date).
    anchor_iso = _STATE.get("data_anchor_date")
    anchor_dt = None
    if anchor_iso:
        try:
            anchor_dt = datetime.fromisoformat(anchor_iso)
        except ValueError:
            anchor_dt = None
    if anchor_dt is None:
        anchor_dt = datetime.now()
    cutoffs = {w: anchor_dt - timedelta(days=w * 30) for w in (12, 24, 36)}

    qty_12 = qty_24 = qty_36 = qty_all = 0.0
    spend_12 = spend_24 = spend_36 = spend_all = 0.0
    first_dt = None
    last_dt = None
    last_unit_price = None
    last_po_for_tiebreak = None
    pos = set()

    for (date_iso, qty, unit_price, line_total, po, uom) in cleaned_lines:
        if qty is None or unit_price is None:
            # Still count toward po_count / first/last but skip aggregates
            if po: pos.add(po)
            try:
                d = datetime.fromisoformat(date_iso) if date_iso else None
            except ValueError:
                d = None
            if d:
                if first_dt is None or d < first_dt: first_dt = d
                if last_dt  is None or d > last_dt:  last_dt = d
            continue
        try:
            d = datetime.fromisoformat(date_iso) if date_iso else None
        except ValueError:
            d = None
        ext = float(line_total) if line_total is not None else (qty * unit_price)
        qty_all += qty
        spend_all += ext
        if d:
            if first_dt is None or d < first_dt: first_dt = d
            if last_dt is None or d > last_dt or (d == last_dt and (po or "") > (last_po_for_tiebreak or "")):
                last_dt = d
                last_unit_price = unit_price
                last_po_for_tiebreak = po
            for w in (12, 24, 36):
                if d >= cutoffs[w]:
                    if w == 12: qty_12 += qty; spend_12 += ext
                    if w == 24: qty_24 += qty; spend_24 += ext
                    if w == 36: qty_36 += qty; spend_36 += ext
        if po: pos.add(po)

    current_price = last_unit_price if last_unit_price is not None else item.get("last_unit_price")

    item["po_count"] = len(pos)
    item["qty_12mo"] = qty_12
    item["qty_24mo"] = qty_24
    item["qty_36mo"] = qty_36
    item["qty_all"]  = qty_all
    # Spend columns: qty × current_price (internally consistent for the table).
    cp = current_price or 0
    item["spend_12mo"] = qty_12 * cp
    item["spend_24mo"] = qty_24 * cp
    item["spend_36mo"] = qty_36 * cp
    item["spend_all"]  = qty_all * cp
    # Historical actuals: sum of line totals on the CLEANED set so file-level
    # KPIs reconcile to the source minus the analyst's removed errors.
    item["spend_12mo_actual"] = spend_12
    item["spend_24mo_actual"] = spend_24
    item["spend_36mo_actual"] = spend_36
    item["spend_all_actual"]  = spend_all
    item["last_unit_price"]   = current_price
    item["first_order"] = first_dt.date().isoformat() if first_dt else None
    item["last_order"]  = last_dt.date().isoformat() if last_dt else None
    # `included` is a soft default; preserve any analyst-set deviation by
    # only flipping it when it equals what the default WAS at extract time.
    # Cheapest correct heuristic: leave `included` alone — the user controls
    # it via the per-row checkbox / smart-trim, and we don't second-guess.
    return item


def _rebuild_kpis_from_items() -> dict:
    """Recompute file-level headline KPIs from the current ``_STATE["items"]``.

    Used after exclusions change so the top-of-page KPI tiles (Items / 12mo /
    24mo / 36mo / Total spend) and active-items counts stay reconciled to
    the cleaned per-item aggregates. ``annual_spend`` stays as the original
    extract — exclusions are too rare to merit re-rolling the per-year
    bars; if needed later we can re-derive from po_lines_by_key.

    Returns the new KPI dict (also written to ``_STATE["kpis"]``).
    """
    items = _STATE.get("items", []) or []
    if not items:
        return _STATE.get("kpis", {})
    prev = dict(_STATE.get("kpis", {}))
    items_24mo = sum(1 for it in items if (it.get("qty_24mo") or 0) > 0)
    items_12mo = sum(1 for it in items if (it.get("qty_12mo") or 0) > 0)
    items_36mo = sum(1 for it in items if (it.get("qty_36mo") or 0) > 0)
    total_spend = sum(it.get("spend_all_actual") or 0 for it in items)
    spend_12mo = sum(it.get("spend_12mo_actual") or 0 for it in items)
    spend_24mo = sum(it.get("spend_24mo_actual") or 0 for it in items)
    spend_36mo = sum(it.get("spend_36mo_actual") or 0 for it in items)
    kpis = {
        **prev,
        "item_count":  len(items),
        "items_12mo":  items_12mo,
        "items_24mo":  items_24mo,
        "items_36mo":  items_36mo,
        "total_spend": total_spend,
        "spend_12mo":  spend_12mo,
        "spend_24mo":  spend_24mo,
        "spend_36mo":  spend_36mo,
    }
    _STATE["kpis"] = kpis
    return kpis


def apply_all_item_exclusions_to_aggregates(rebuild_kpis: bool = True) -> dict:
    """Replay every saved exclusion against the current items list.

    Called at the tail of ``extract_rfq_list`` when a previously-saved
    session is reloaded — items are rebuilt from scratch, but exclusions
    were restored ahead of the extract via ``restore_state``, so we need
    to re-derive aggregates for every affected item.

    Returns ``{"n_items_updated": int, "kpis": <updated>}``.
    """
    excl = _STATE.get("item_exclusions", {}) or {}
    n = 0
    for item_num in list(excl.keys()):
        rec = _recompute_item_aggregates_for(item_num)
        if rec:
            n += 1
    new_kpis = _rebuild_kpis_from_items() if rebuild_kpis else _STATE.get("kpis", {})
    return {"n_items_updated": n, "kpis": new_kpis}


# ---------------------------------------------------------------------------
# xlsx generator — candidate RFQ list (internal-audience; not yet
# the supplier-bound RFQ workbook — that comes in Phase 2).
# ---------------------------------------------------------------------------

def autosize(ws, min_w: int = 10, max_w: int = 60):
    """Per-column autosize that skips MergedCell anchors (which lack column_letter).
    Verbatim pattern from supplier-pricing/app.py (dev-lessons #4)."""
    by_col: dict = defaultdict(list)
    for row in ws.iter_rows():
        for c in row:
            if c.__class__.__name__ == "MergedCell":
                continue
            col_letter = c.column_letter
            v = c.value
            if v is None:
                continue
            by_col[col_letter].append(len(str(v)))
    for letter, lens in by_col.items():
        ml = max(lens) if lens else 0
        ws.column_dimensions[letter].width = min(max(ml + 2, min_w), max_w)


def gen_candidate_rfq_list_xlsx(included_keys=None) -> bytes:
    """Build an xlsx of the candidate RFQ list. Internal audience.

    If `included_keys` is provided (list of item_num strings), filter to those.
    Otherwise emit every item flagged `included=True` from the last extract.
    """
    items = _STATE.get("items", [])
    kpis = _STATE.get("kpis", {})
    annual = _STATE.get("annual_spend", [])

    if included_keys is not None:
        keep = set(str(k) for k in included_keys)
        items = [it for it in items if it["item_num"] in keep]

    wb = Workbook()

    # ---- Sheet 1: Candidate RFQ list ----
    ws = wb.active
    ws.title = "RFQ Candidate List"
    headers = [
        "Item #", "Description", "Manufacturer", "MFG Part #", "Commodity", "UOM", "UOM mixed?",
        "12mo qty", "12mo spend", "24mo qty", "24mo spend", "36mo qty", "36mo spend",
        "All-time qty", "All-time spend", "Last unit price", "First order", "Last order",
        "Distinct POs",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3A2917")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for it in sorted(items, key=lambda x: x["spend_24mo"] or 0, reverse=True):
        ws.append([
            it["item_num"], it["description"], it["mfg_name"], it["mfg_pn"],
            it["commodity"], it["uom"], "YES" if it["uom_mixed"] else "",
            it["qty_12mo"], it["spend_12mo"],
            it["qty_24mo"], it["spend_24mo"],
            it["qty_36mo"], it["spend_36mo"],
            it["qty_all"], it["spend_all"],
            it["last_unit_price"],
            it["first_order"], it["last_order"], it["po_count"],
        ])

    # Number formats
    for col_letter in ("H", "J", "L", "N"):  # qty cols
        for row in ws.iter_rows(min_row=2, min_col=ord(col_letter)-64, max_col=ord(col_letter)-64):
            for c in row:
                c.number_format = "#,##0"
    for col_letter in ("I", "K", "M", "O", "P"):  # money cols
        for row in ws.iter_rows(min_row=2, min_col=ord(col_letter)-64, max_col=ord(col_letter)-64):
            for c in row:
                c.number_format = "$#,##0.00"

    autosize(ws)
    ws.freeze_panes = "A2"

    # ---- Sheet 2: Summary ----
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Auto RFQ Banana — extraction summary", ""])
    ws2.append(["", ""])
    ws2.append(["Items in candidate list", len(items)])
    ws2.append(["Source supplier", _STATE.get("supplier_name", "")])
    ws2.append(["Date range", f"{kpis.get('first_order','—')} → {kpis.get('last_order','—')}"])
    ws2.append(["Years of history", round(kpis.get("years_span", 0), 1)])
    ws2.append(["Total POs", kpis.get("po_count", 0)])
    ws2.append(["Total order lines", kpis.get("line_count", 0)])
    ws2.append(["", ""])
    ws2.append(["Window", "Total spend", "Items active"])
    ws2.append(["12 months", kpis.get("spend_12mo", 0), kpis.get("items_12mo", 0)])
    ws2.append(["24 months", kpis.get("spend_24mo", 0), kpis.get("items_24mo", 0)])
    ws2.append(["36 months", kpis.get("spend_36mo", 0), kpis.get("items_36mo", 0)])
    ws2.append(["All time",  kpis.get("total_spend", 0), kpis.get("item_count", 0)])
    for r in (10, 11, 12, 13, 14):
        c = ws2.cell(row=r, column=2)
        c.number_format = "$#,##0"
    ws2["A1"].font = Font(bold=True, size=14)
    autosize(ws2)

    # ---- Sheet 3: Annual spend ----
    ws3 = wb.create_sheet("Annual Spend")
    ws3.append(["Year", "Spend", "Quantity", "Distinct items"])
    for c in ws3[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="3A2917")
    for d in annual:
        ws3.append([d["year"], d["spend"], d["qty"], d["item_count"]])
    for row in ws3.iter_rows(min_row=2, min_col=2, max_col=2):
        for c in row:
            c.number_format = "$#,##0"
    for row in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
        for c in row:
            c.number_format = "#,##0"
    autosize(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Returned-bid intake — parses each supplier's response file.
#
# Three observed file shapes in the McMaster RFQ:
#   1. "Our format" (Grainger primary, MSC, Fastenal): banner row at top,
#      real headers at row 7 (1-indexed). Cols 1-10:
#        1=Commodity 2=Item 3=EAM Part # 4=Part Number (= McMaster anchor)
#        5=Item # 6=Manufacturer Name 7=Qty 8=UOM 9=Quoted Price 10=Notes
#      Fastenal extends to 20 cols with substitute-part / verified-pricing data.
#   2. "Their format" (Grainger secondary): completely restructured. NOT
#      auto-handled in v1 — falls back to manual mapping (TODO).
#
# No-bid signals (treat as not quoted, not a real $0 line):
#   - Quoted Price is 0 / empty / None
#   - Notes contains "need more information" / "n/a" / "discontinued" / "obsolete"
# ---------------------------------------------------------------------------

# Status enum-ish strings used in the comparison matrix
BID_STATUS_PRICED      = "PRICED"        # quoted with a real price
BID_STATUS_NO_BID      = "NO_BID"        # supplier explicitly declined or left blank
BID_STATUS_NEED_INFO   = "NEED_INFO"     # supplier asked for clarification
BID_STATUS_UOM_DISC    = "UOM_DISC"      # priced but with a UOM-mismatch warning
BID_STATUS_SUBSTITUTE  = "SUBSTITUTE"    # priced with an alternate part offered

# Substring-safe no-bid markers — these phrases are long enough that they can't
# accidentally match other words via substring search.
NO_BID_LONG_MARKERS = (
    "need more information", "need more info", "need info", "more info",
    "discontinued", "obsolete", "no quote", "no bid",
    "unable to quote", "cannot quote", "not available", "not stocked",
    "no longer available", "to be quoted",
)
# Short markers ("na", "tbd", "n/a") that need WORD-BOUNDARY matching to avoid
# false positives — "na" was matching "fasteNAl" in supplier notes, causing
# every UOM-discrepancy line to be misclassified as NO_BID. Discovered during
# the demo-existing-rfq workflow on the real Fastenal bid file (2026-04-29).
import re as _re
_NO_BID_SHORT_RE = _re.compile(r"\b(?:n/?a|tbd)\b", _re.IGNORECASE)


def _matches_no_bid(notes_text: str) -> bool:
    """True if the notes text contains an explicit no-bid signal.

    Long markers (multi-word phrases) use substring matching — they can't
    accidentally match other words. Short ambiguous markers (na, tbd, n/a)
    use word-boundary regex to avoid matching them inside larger words like
    "fasteNAl".
    """
    if not notes_text:
        return False
    t = str(notes_text).strip().lower()
    if not t:
        return False
    if _NO_BID_SHORT_RE.search(t):
        return True
    return any(m in t for m in NO_BID_LONG_MARKERS)


# Backwards-compatible alias for any external code that imports the constant.
NO_BID_NOTE_MARKERS = NO_BID_LONG_MARKERS + ("n/a", "na", "tbd")


def _detect_our_format(ws) -> int:
    """Look for the canonical header row in our-format response files.
    Returns the 1-indexed row number where headers live, or 0 if not detected.

    The signature: at least 4 cells in the row that EXACTLY match expected
    header names (after strip + lowercase). Banner rows that contain the
    phrase 'quoted price' embedded in long instruction text don't match
    because we require exact cell equality, not substring.
    """
    expected = {
        "commodity", "item", "eam part number", "part number", "partnumber",
        "item #", "manufacturer name", "qty", "uom", "quoted price", "notes",
    }
    for r_idx in range(1, 16):
        try:
            row = next(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True), None)
        except Exception:
            continue
        if not row:
            continue
        cells = [str(c or "").strip().lower() for c in row]
        hits = sum(1 for c in cells if c in expected)
        if hits >= 4:
            return r_idx
    return 0


def _our_format_columns(header_row: tuple) -> dict:
    """Map the 'our format' headers to logical fields.
    Handles the 11-col primary layout AND the 20-col Fastenal extension."""
    cols = {}
    for i, raw in enumerate(header_row):
        if raw is None:
            continue
        h = str(raw).strip().lower()
        if not h:
            continue
        # Core columns (assigned only once — first match wins for ambiguous names)
        if h == "commodity" and "commodity" not in cols:
            cols["commodity"] = i
        elif h == "item" and "description" not in cols:
            cols["description"] = i
        elif h == "eam part number" and "eam_pn" not in cols:
            cols["eam_pn"] = i
        elif h in ("part number", "partnumber") and "part_number" not in cols:
            cols["part_number"] = i        # = McMaster anchor
        elif h == "item #" and "item_num" not in cols:
            cols["item_num"] = i
        elif h == "manufacturer name" and "mfg_name" not in cols:
            cols["mfg_name"] = i
        elif h == "qty" and "qty" not in cols:
            cols["qty"] = i
        elif h == "uom" and "uom" not in cols:
            cols["uom"] = i
        elif h == "quoted price" and "quoted_price" not in cols:
            cols["quoted_price"] = i
        elif h == "notes" and "notes" not in cols:
            cols["notes"] = i
        elif h == "verified pricing" and "verified_price" not in cols:
            cols["verified_price"] = i     # Fastenal extension
        elif h.startswith("exact ") and "part" in h and "exact_part" not in cols:
            cols["exact_part"] = i         # Fastenal: their exact catalog part
        elif h.startswith("exact ") and "description" in h and "exact_desc" not in cols:
            cols["exact_desc"] = i
        elif h.startswith("sub ") and "part" in h and "sub_part" not in cols:
            cols["sub_part"] = i           # Fastenal: substitute offered
        elif h.startswith("sub ") and "description" in h and "sub_desc" not in cols:
            cols["sub_desc"] = i
        elif h == "sell price/uom" and "sub_price" not in cols:
            cols["sub_price"] = i
        elif h == "notes" and "notes_2" not in cols:
            cols["notes_2"] = i            # Fastenal's second notes col
    return cols


def parse_supplier_bid(file_bytes, supplier_name: str = "") -> dict:
    """Parse one supplier's returned-bid xlsx. Auto-detects 'our format'
    (banner + headers around row 7). Returns:

        {
          "supplier": str,
          "bids": [{rfq_key, item_num, part_number, mfg_name, description,
                    qty, uom, quoted_price, notes, status, alt_part,
                    verified_price, has_substitute, ...}, ...],
          "summary": {
              "n_lines": int, "n_priced": int, "n_no_bid": int,
              "n_need_info": int, "n_uom_disc": int, "n_substitute": int,
              "total_quoted_value": float,
          },
          "format": "our_format" | "unknown",
        }
    """
    if not isinstance(file_bytes, (bytes, bytearray)):
        file_bytes = bytes(file_bytes)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row_idx = _detect_our_format(ws)
    if not header_row_idx:
        wb.close()
        return {"supplier": supplier_name, "bids": [], "format": "unknown",
                "summary": {"n_lines": 0, "n_priced": 0, "n_no_bid": 0,
                            "n_need_info": 0, "n_uom_disc": 0, "n_substitute": 0,
                            "total_quoted_value": 0.0},
                "error": "Could not auto-detect header row. Use manual column mapping."}

    header_row = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
    cols = _our_format_columns(header_row)

    def g(row, field):
        i = cols.get(field)
        return None if (i is None or i >= len(row)) else row[i]

    bids = []
    n_priced = n_no_bid = n_need_info = n_uom_disc = n_substitute = 0
    total_value = 0.0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None:
            continue
        # Anchor key: McMaster Part Number (the column we sent — same anchor
        # the supplier filled prices against). Fall back to Item # then EAM.
        part = norm_text(g(row, "part_number"))
        item = norm_text(g(row, "item_num"))
        eam  = norm_text(g(row, "eam_pn"))
        if _is_blanky(part): part = ""
        if _is_blanky(item): item = ""
        if _is_blanky(eam):  eam  = ""
        key_raw = part or item or eam
        if not key_raw:
            continue
        rfq_key = norm_pn(key_raw)

        qty = safe_float(g(row, "qty"))
        quoted = safe_float(g(row, "quoted_price"))
        verified = safe_float(g(row, "verified_price"))
        notes = norm_text(g(row, "notes")) or ""
        notes_2 = norm_text(g(row, "notes_2")) or ""
        if notes_2 and notes_2 not in notes:
            notes = (notes + " | " + notes_2).strip(" |")
        sub_part = norm_text(g(row, "sub_part")) or ""
        sub_desc = norm_text(g(row, "sub_desc")) or ""
        sub_price = safe_float(g(row, "sub_price"))
        exact_part = norm_text(g(row, "exact_part")) or ""
        exact_desc = norm_text(g(row, "exact_desc")) or ""

        # Status determination — order matters
        is_no_bid = (
            (quoted is None or quoted == 0) and verified is None
        ) or _matches_no_bid(notes)
        has_uom_warning = "uom" in notes.lower() and ("disc" in notes.lower() or "mismatch" in notes.lower() or "differ" in notes.lower())
        has_substitute_offered = bool(sub_part) and not is_no_bid

        if is_no_bid:
            if any(m in notes.lower() for m in ("need more info", "need info", "more info", "tbd", "to be quoted")):
                status = BID_STATUS_NEED_INFO
                n_need_info += 1
            else:
                status = BID_STATUS_NO_BID
                n_no_bid += 1
        elif has_uom_warning:
            status = BID_STATUS_UOM_DISC
            n_uom_disc += 1
            n_priced += 1
        elif has_substitute_offered:
            status = BID_STATUS_SUBSTITUTE
            n_substitute += 1
            n_priced += 1
        else:
            status = BID_STATUS_PRICED
            n_priced += 1

        # Effective price — verified takes precedence (it's the "we re-confirmed"
        # value Fastenal added) but we keep the raw quoted value for transparency
        effective_price = verified if (verified is not None and verified > 0) else quoted
        if effective_price is not None and qty is not None and effective_price > 0:
            total_value += effective_price * qty

        bids.append({
            "rfq_key": rfq_key,
            "item_num": item,
            "part_number": part,
            "eam_pn": eam,
            "mfg_name": norm_text(g(row, "mfg_name")) or "",
            "description": norm_text(g(row, "description")) or "",
            "commodity": norm_text(g(row, "commodity")) or "",
            "qty": qty,
            "uom": norm_text(g(row, "uom")) or "",
            "quoted_price": quoted,
            "verified_price": verified,
            "effective_price": effective_price,
            "notes": notes,
            "status": status,
            # Fastenal-style extensions (empty for other suppliers)
            "exact_part": exact_part,
            "exact_desc": exact_desc,
            "sub_part": sub_part,
            "sub_desc": sub_desc,
            "sub_price": sub_price,
            "has_substitute": has_substitute_offered,
        })

    wb.close()

    return {
        "supplier": supplier_name or "(unnamed)",
        "format": "our_format",
        "header_row": header_row_idx,
        "columns_detected": cols,
        "bids": bids,
        "summary": {
            "n_lines": len(bids),
            "n_priced": n_priced,
            "n_no_bid": n_no_bid,
            "n_need_info": n_need_info,
            "n_uom_disc": n_uom_disc,
            "n_substitute": n_substitute,
            "total_quoted_value": total_value,
        },
    }


def ingest_supplier_bid(file_bytes, supplier_name: str) -> dict:
    """Parse and persist a supplier bid into _STATE. Returns the parse result.

    Replaces any prior bid for the same supplier_name (re-uploads overwrite).
    """
    parsed = parse_supplier_bid(file_bytes, supplier_name)
    if "bids" not in _STATE or not isinstance(_STATE.get("bids"), dict):
        _STATE["bids"] = {}
    _STATE["bids"][supplier_name] = parsed
    s = parsed.get("summary", {})
    log_event(
        "ingest_supplier_bid",
        f"{s.get('n_lines',0):,} lines / {s.get('n_priced',0):,} priced / ${s.get('total_quoted_value',0):,.0f} quoted",
        related=supplier_name,
    )
    return parsed


def list_supplier_bids() -> list:
    """Summary of all loaded bids."""
    bids = _STATE.get("bids", {})
    return [
        {
            "supplier": s,
            "format": p.get("format"),
            "summary": p.get("summary", {}),
        }
        for s, p in bids.items()
    ]


def _norm_uom(u) -> str:
    """Normalize a UOM string for equality checks. Empty string if missing."""
    if not u:
        return ""
    s = str(u).strip().upper()
    # Common aliases — Each, EACH, Ea., EA all normalize to "EA"
    s = s.replace(".", "").replace("EACH", "EA")
    s = s.replace("PACKAGE", "PK").replace("PACK", "PK")
    s = s.replace("FOOT", "FT").replace("FEET", "FT")
    s = s.replace("INCH", "IN").replace("INCHES", "IN")
    return s


def compute_clean_savings_summary_DEPRECATED_BIDLINE_BASED() -> dict:
    """Compute "clean" savings metrics that exclude bids polluted by data-quality
    issues. Operates on the current _STATE bids + items.

    Three savings tiers per supplier, each progressively stricter:

    - **raw**: every bid with a price > 0, including UOM_DISC and SUBSTITUTE
      statuses. This is what the current dashboard shows, but it's polluted
      when supplier UOM differs from history UOM (per-each Fastenal price
      compared against per-package McMaster anchor inflates apparent savings
      to absurd values).

    - **clean**: bids with status == PRICED only. Excludes UOM_DISC,
      NO_BID, NEED_INFO, SUBSTITUTE. Catches the explicitly-flagged UOM
      mismatches but misses implicit ones (Grainger / MSC don't always note
      UOM in the response — they just quote in their own units).

    - **strict**: clean + the bid's UOM matches the history's UOM (after
      normalization). Catches implicit UOM mismatches the supplier didn't
      annotate. This is the most defensible savings number.

    Returns:
        {
          "by_supplier": [
            {"supplier", "raw_savings", "clean_savings", "strict_savings",
             "n_raw", "n_clean", "n_strict",
             "n_excluded_by_status", "n_excluded_by_uom"},
            ...
          ],
          "totals": {"raw": float, "clean": float, "strict": float},
        }

    Negative savings = bid is more expensive than history. Positive savings
    = bid is cheaper.

    Note this is ADDITIVE — it doesn't modify any existing pipeline output.
    Designed to be surfaced alongside the existing comparison matrix in the
    dashboard so users can see both the raw and corrected numbers and judge
    which to use for their decision.
    """
    items = _STATE.get("items", [])
    bids_by_supplier = _STATE.get("bids", {}) or {}

    # Build a per-item lookup for history qty + price + UOM
    items_by_key = {}
    for it in items:
        items_by_key[it["key"]] = {
            "qty_24mo": it.get("qty_24mo") or 0,
            "last_unit_price": it.get("last_unit_price") or 0,
            "uom": _norm_uom(it.get("uom")),
        }

    by_supplier = []
    totals = {"raw": 0.0, "clean": 0.0, "strict": 0.0}

    for sup, parsed in bids_by_supplier.items():
        raw_save = clean_save = strict_save = 0.0
        n_raw = n_clean = n_strict = 0
        n_excl_status = n_excl_uom = 0

        for b in parsed.get("bids", []):
            bid_price = b.get("effective_price") or 0
            if not bid_price or bid_price <= 0:
                continue
            it = items_by_key.get(b.get("rfq_key"))
            if not it: continue
            hist_price = it["last_unit_price"]
            qty = it["qty_24mo"]
            if not (hist_price > 0 and qty > 0): continue

            # RAW — all priced bids contribute
            line_savings = (hist_price - bid_price) * qty
            raw_save += line_savings; n_raw += 1

            # CLEAN — status-filtered
            if b.get("status") != BID_STATUS_PRICED:
                n_excl_status += 1
                continue
            clean_save += line_savings; n_clean += 1

            # STRICT — additionally require UOM match (when both knowable)
            bid_uom = _norm_uom(b.get("uom"))
            hist_uom = it["uom"]
            if bid_uom and hist_uom and bid_uom != hist_uom:
                n_excl_uom += 1
                continue
            strict_save += line_savings; n_strict += 1

        by_supplier.append({
            "supplier": sup,
            "raw_savings": round(raw_save, 2),
            "clean_savings": round(clean_save, 2),
            "strict_savings": round(strict_save, 2),
            "n_raw": n_raw,
            "n_clean": n_clean,
            "n_strict": n_strict,
            "n_excluded_by_status": n_excl_status,
            "n_excluded_by_uom": n_excl_uom,
        })
        totals["raw"] += raw_save
        totals["clean"] += clean_save
        totals["strict"] += strict_save

    by_supplier.sort(key=lambda d: -d["strict_savings"])
    return {
        "by_supplier": by_supplier,
        "totals": {k: round(v, 2) for k, v in totals.items()},
    }


def compute_clean_savings_summary() -> dict:
    """Compute "clean" savings metrics that exclude bids polluted by data-quality
    issues, plus a NORMALIZED tier that re-includes UOM-mismatched items where
    the analyst has supplied a manual conversion factor (or where one was
    auto-extracted from the supplier's notes).

    Four tiers per supplier, each progressively stricter then re-broadened:

    - RAW        — every priced bid (current dashboard behavior; polluted by
                   UOM mismatches)
    - CLEAN      — excludes UOM_DISC / NO_BID / NEED_INFO / SUBSTITUTE statuses
    - STRICT     — CLEAN + bid UOM equals history UOM after normalization
    - NORMALIZED — STRICT + UOM-mismatched items where a "resolved" annotation
                   exists in _STATE["uom_annotations"]. Bid price is adjusted
                   by the annotation's factor before computing savings, so a
                   per-each Fastenal bid of $0.0123 with factor=50 contributes
                   $0.615 (the per-package equivalent) for comparison against
                   a per-package McMaster anchor of e.g. $0.620.

    Returns:
        {
          "by_supplier": [
            {"supplier", "raw_savings", "clean_savings", "strict_savings",
             "normalized_savings",
             "n_raw", "n_clean", "n_strict", "n_normalized",
             "n_excluded_by_status", "n_excluded_by_uom",
             "n_resolved_by_annotation",   # how many of the n_excluded_by_uom
                                            # were recovered by an annotation
            },
            ...
          ],
          "totals": {"raw": float, "clean": float, "strict": float, "normalized": float},
        }

    NORMALIZED is a SUPERSET of STRICT — it includes everything in STRICT plus
    the recovered annotated items. Use NORMALIZED for the "after the analyst
    has done UOM resolution" headline.
    """
    matrix = compute_comparison_matrix()
    rows = matrix.get("rows", [])
    suppliers = matrix.get("suppliers", [])
    annotations = _STATE.get("uom_annotations") or {}

    by_supplier_dict = {sup: {
        "supplier": sup,
        "raw_savings": 0.0, "clean_savings": 0.0,
        "strict_savings": 0.0, "normalized_savings": 0.0,
        "n_raw": 0, "n_clean": 0, "n_strict": 0, "n_normalized": 0,
        "n_excluded_by_status": 0, "n_excluded_by_uom": 0,
        "n_resolved_by_annotation": 0,
    } for sup in suppliers}

    totals = {"raw": 0.0, "clean": 0.0, "strict": 0.0, "normalized": 0.0}

    for r in rows:
        hist_price = r.get("last_unit_price") or 0
        qty = r.get("qty_24mo") or 0
        if hist_price <= 0 or qty <= 0:
            continue
        hist_uom = _norm_uom(r.get("uom"))
        bids = r.get("bids") or {}
        for sup, b in bids.items():
            if not isinstance(b, dict): continue
            bid_price = b.get("price") or 0
            if not bid_price or bid_price <= 0: continue
            line_savings = (hist_price - bid_price) * qty

            agg = by_supplier_dict[sup]
            # RAW
            agg["raw_savings"] += line_savings
            agg["n_raw"] += 1
            totals["raw"] += line_savings

            status = b.get("status")
            bid_uom = _norm_uom((b.get("raw") or {}).get("uom"))
            uom_match = (not bid_uom or not hist_uom or bid_uom == hist_uom)
            ann = annotations.get(_annotation_key(r.get("rfq_key"), sup))

            # CLEAN — status-filtered
            if status != BID_STATUS_PRICED:
                agg["n_excluded_by_status"] += 1
                # Status-excluded items mostly get NO normalized recovery —
                # NO_BID/NEED_INFO/SUBSTITUTE means there isn't a clean priced
                # number to begin with. EXCEPT: UOM_DISC items DO have valid
                # prices (just flagged for unit warning); analyst-annotated
                # ones can still flow into NORMALIZED.
                if status == BID_STATUS_UOM_DISC and ann and ann.get("status") == "resolved":
                    adjusted_bid, _ = _apply_annotation_to_price(bid_price, hist_price, ann)
                    if adjusted_bid is not None:
                        adjusted_savings = (hist_price - adjusted_bid) * qty
                        agg["normalized_savings"] += adjusted_savings
                        agg["n_normalized"] += 1
                        agg["n_resolved_by_annotation"] += 1
                        totals["normalized"] += adjusted_savings
                continue
            agg["clean_savings"] += line_savings
            agg["n_clean"] += 1
            totals["clean"] += line_savings

            # STRICT — clean + UOM match
            if uom_match:
                agg["strict_savings"] += line_savings
                agg["n_strict"] += 1
                totals["strict"] += line_savings
                # And also counted in NORMALIZED (STRICT + annotation recoveries)
                agg["normalized_savings"] += line_savings
                agg["n_normalized"] += 1
                totals["normalized"] += line_savings
            else:
                # Silent UOM mismatch (status=PRICED but units differ) — out of
                # STRICT. Try to recover via annotation.
                agg["n_excluded_by_uom"] += 1
                if ann and ann.get("status") == "resolved":
                    adjusted_bid, _ = _apply_annotation_to_price(bid_price, hist_price, ann)
                    if adjusted_bid is not None:
                        adjusted_savings = (hist_price - adjusted_bid) * qty
                        agg["normalized_savings"] += adjusted_savings
                        agg["n_normalized"] += 1
                        agg["n_resolved_by_annotation"] += 1
                        totals["normalized"] += adjusted_savings

    by_supplier = list(by_supplier_dict.values())
    by_supplier.sort(key=lambda d: -d["normalized_savings"])
    for s in by_supplier:
        for k in ("raw_savings", "clean_savings", "strict_savings", "normalized_savings"):
            s[k] = round(s[k], 2)
    return {
        "by_supplier": by_supplier,
        "totals": {k: round(v, 2) for k, v in totals.items()},
    }


# ---------------------------------------------------------------------------
# UOM resolution — manual annotations + auto-fill from supplier notes
# ---------------------------------------------------------------------------
#
# The problem: when a bid's UOM doesn't match the history's UOM (e.g., Fastenal
# quotes per-each $0.0123, McMaster anchor is per-package $1.23), comparing them
# directly produces nonsense savings. The STRICT savings tier in
# compute_clean_savings_summary excludes these from headline math.
#
# This module lets analysts ANNOTATE items with the conversion factor so the
# bid price can be normalized to the history's UOM, moving the item from
# STRICT-excluded to NORMALIZED-comparable.
#
# Two ways an annotation gets set:
#   1. AUTO — extracted from a parseable supplier note (e.g., "(50' Spool)"
#      tells us 1 spool = 50 ft, so per-foot bid × 50 = per-spool comparable)
#   2. MANUAL — analyst looks up McMaster catalog (or stockroom) offline and
#      types in the conversion factor via the UOM Resolution Queue UI
#
# Persists in _STATE["uom_annotations"] keyed by (item_key, supplier). Rides
# through serialize_state / restore_state so the work survives save/load.
# ---------------------------------------------------------------------------

import re as _uom_re

# Patterns we can auto-extract from supplier notes. Each tuple:
#   (regex, capture-group-name, factor-derivation-fn(match) -> int|float|None)
# Returning None means the pattern matched but we couldn't pin down a factor
# (e.g., "(Per Inch)" tells us the bid UOM but not how many inches per piece).
_AUTO_PATTERNS = [
    # "(50' Spool)" / "(300' Spool)" → factor = the number (length of one spool)
    (_uom_re.compile(r"\((\d+)\s*['′]?\s*(?:Spool|Roll|Reel)\)", _uom_re.IGNORECASE),
     "spool_size",
     lambda m: int(m.group(1))),
    # "Pack of 10", "Pack of 25" → factor = pack count
    (_uom_re.compile(r"Pack\s*of\s*(\d+)", _uom_re.IGNORECASE),
     "pack_of",
     lambda m: int(m.group(1))),
    # "(10 Pack)", "(25/Pack)" → same
    (_uom_re.compile(r"\(\s*(\d+)\s*[/\s]?Pack\s*\)", _uom_re.IGNORECASE),
     "pack_count",
     lambda m: int(m.group(1))),
    # "Bag of 100" / "Box of 50" / "Carton of 12"
    (_uom_re.compile(r"(?:Bag|Box|Carton|Bundle)\s*of\s*(\d+)", _uom_re.IGNORECASE),
     "container_of",
     lambda m: int(m.group(1))),
    # "(Per Inch)" / "(Per Foot)" — we know the supplier's per-X but not the
    # piece dimensions. Return None for factor; flag as needs-review.
    (_uom_re.compile(r"\(?\s*Per\s+(Inch|Foot|Each|Meter|Centimeter|Yard)\s*\)?", _uom_re.IGNORECASE),
     "per_x",
     lambda m: None),
]


def _extract_pack_size_from_notes(notes_text: str) -> dict:
    """Look at a supplier-bid note for parseable UOM-conversion patterns.

    Returns:
        {"factor": int|float|None,
         "pattern": str,                  # which pattern matched
         "raw_match": str,                # the matched substring
         "confidence": "high" | "low"}    # high = explicit count; low = unit-only
        OR None if no pattern matched.

    Examples:
        "UOM Discrepancy (50' Spool)"     -> {factor: 50, pattern: "spool_size", confidence: "high"}
        "Sold per Pack of 25"             -> {factor: 25, pattern: "pack_of", confidence: "high"}
        "(Per Inch)"                       -> {factor: None, pattern: "per_x", confidence: "low"}
        "REVIEWED AND CORRECT!"            -> None
    """
    if not notes_text:
        return None
    s = str(notes_text)
    for rx, name, factor_fn in _AUTO_PATTERNS:
        m = rx.search(s)
        if m:
            try:
                factor = factor_fn(m)
            except Exception:
                factor = None
            return {
                "factor": factor,
                "pattern": name,
                "raw_match": m.group(0),
                "confidence": "high" if factor is not None else "low",
            }
    return None


def _annotation_key(item_key: str, supplier: str) -> str:
    """Composite key — annotations are per-(item, supplier) since the same
    item can have different UOM mismatches with different suppliers."""
    return f"{item_key}|{supplier}"


def set_uom_annotation(item_key: str, supplier: str, factor, hist_uom: str = "",
                       bid_uom: str = "", note: str = "", status: str = "resolved",
                       set_by: str = "", direction: str = "auto_detect") -> dict:
    """Manually annotate the conversion factor for a (item, supplier) pair.

    Args:
        item_key:  the item's RFQ key (item.key from extract_rfq_list)
        supplier:  the supplier name (matches one of _STATE["bids"] keys)
        factor:    the integer/float relating the two UOMs. Direction matters:
                   - direction="multiply": adjusted_bid = bid_price × factor
                                            (use when supplier quoted in smaller units)
                   - direction="divide": adjusted_bid = bid_price / factor
                                          (use when supplier quoted in larger units)
                   - direction="auto_detect": at apply-time, pick whichever
                                              direction puts adjusted_bid closer
                                              to hist_price. Useful when the
                                              user is uncertain; risky for
                                              cases where both directions are
                                              far from history.
                   Pass None to mark needs-review without committing a factor.
        hist_uom:  the history's UOM string (snapshot, for sanity-check)
        bid_uom:   the bid's UOM string (snapshot)
        note:      optional analyst memo (e.g., "McMaster pkg/50 per catalog")
        status:    "resolved" | "skipped" | "needs_review"
        set_by:    optional user name for audit trail
        direction: see factor docstring above

    Returns the stored annotation dict.
    """
    if not item_key or not supplier:
        raise ValueError("item_key and supplier are required")
    if status not in ("resolved", "skipped", "needs_review"):
        raise ValueError(f"invalid status: {status!r}")
    if direction not in ("multiply", "divide", "auto_detect"):
        raise ValueError(f"invalid direction: {direction!r}")

    if "uom_annotations" not in _STATE:
        _STATE["uom_annotations"] = {}
    key = _annotation_key(item_key, supplier)
    annotation = {
        "item_key": item_key,
        "supplier": supplier,
        "factor": factor,
        "direction": direction,
        "hist_uom": hist_uom,
        "bid_uom": bid_uom,
        "note": note,
        "status": status,
        "set_by": set_by,
        "set_at": datetime.utcnow().isoformat() + "Z",
    }
    _STATE["uom_annotations"][key] = annotation
    log_event("uom_annotation_set",
              f"factor={factor} dir={direction} status={status}" + (f" note={note!r}" if note else ""),
              related=f"{supplier}:{item_key}")
    return annotation


def _apply_annotation_to_price(bid_price: float, hist_price: float, ann: dict):
    """Apply a UOM annotation to a bid price, respecting the direction.

    Returns (adjusted_bid, direction_used) or (None, None) if the annotation
    can't be applied (missing factor, invalid type, etc.).
    """
    factor = ann.get("factor")
    if not factor:
        return None, None
    try:
        f = float(factor)
        if f <= 0:
            return None, None
    except (TypeError, ValueError):
        return None, None
    direction = ann.get("direction", "auto_detect")
    if direction == "multiply":
        return bid_price * f, "multiply"
    if direction == "divide":
        return bid_price / f, "divide"
    # auto_detect — pick the direction that puts adjusted closer to hist_price.
    # Compute both candidates' relative distance from hist; prefer the smaller.
    cand_mul = bid_price * f
    cand_div = bid_price / f
    dist_mul = abs(cand_mul - hist_price) / hist_price if hist_price > 0 else float("inf")
    dist_div = abs(cand_div - hist_price) / hist_price if hist_price > 0 else float("inf")
    if dist_mul <= dist_div:
        return cand_mul, "multiply (auto)"
    return cand_div, "divide (auto)"


def get_uom_annotation(item_key: str, supplier: str) -> dict:
    """Return the stored annotation for (item, supplier), or None if not set."""
    return (_STATE.get("uom_annotations") or {}).get(_annotation_key(item_key, supplier))


def clear_uom_annotation(item_key: str, supplier: str) -> bool:
    """Remove an annotation. Returns True if one was present, False if not."""
    key = _annotation_key(item_key, supplier)
    annotations = _STATE.get("uom_annotations") or {}
    if key in annotations:
        del annotations[key]
        log_event("uom_annotation_cleared", "removed", related=f"{supplier}:{item_key}")
        return True
    return False


def list_items_needing_uom_resolution(spend_threshold: float = 0) -> list:
    """Build the queue for the UOM Resolution UI.

    Returns a list of {item, supplier, bid, history_qty, history_price,
    auto_suggestion, annotation, ranked_by_spend_24mo} for every (item,
    supplier) pair that:
        - Has a priced bid (effective_price > 0)
        - Has a usable history anchor (last_unit_price > 0, qty_24mo > 0)
        - Either:
              a) bid status is UOM_DISC, OR
              b) bid_uom != hist_uom after normalization
          AND has NOT already been resolved/skipped (status != "resolved" / "skipped")

    Items already annotated with status="resolved" or "skipped" are filtered
    OUT (they're already handled). Items annotated "needs_review" stay in the
    list because they're still pending analyst input.

    Auto-suggestion field includes any parseable pattern from supplier notes
    (e.g., "(50' Spool)" → {factor: 50}).

    spend_threshold: optional filter to skip low-value items.
    """
    matrix = compute_comparison_matrix()
    rows = matrix.get("rows", [])
    annotations = _STATE.get("uom_annotations") or {}
    out = []

    for r in rows:
        hist_uom = _norm_uom(r.get("uom"))
        hist_price = r.get("last_unit_price") or 0
        qty = r.get("qty_24mo") or 0
        if hist_price <= 0 or qty <= 0:
            continue
        spend_24 = hist_price * qty
        if spend_24 < spend_threshold:
            continue
        item_key = r.get("rfq_key")

        for sup, bid in (r.get("bids") or {}).items():
            if not isinstance(bid, dict): continue
            bid_price = bid.get("price") or 0
            if bid_price <= 0: continue

            bid_raw = bid.get("raw") or {}
            bid_uom = _norm_uom(bid_raw.get("uom"))
            status = bid.get("status", "")
            notes = bid.get("notes", "") or bid_raw.get("notes", "")

            # Two reasons an item needs resolution:
            #   1) Status was flagged UOM_DISC by the parser
            #   2) UOMs differ silently (status == PRICED but units don't match)
            needs_resolution = (status == BID_STATUS_UOM_DISC) or (
                bid_uom and hist_uom and bid_uom != hist_uom
            )
            if not needs_resolution:
                continue

            # Skip already-handled annotations
            existing = annotations.get(_annotation_key(item_key, sup))
            if existing and existing["status"] in ("resolved", "skipped"):
                continue

            # Auto-suggest from notes if possible
            auto = _extract_pack_size_from_notes(notes)

            out.append({
                "item_key": item_key,
                "item_num": r.get("item_num"),
                "description": r.get("description"),
                "supplier": sup,
                "hist_uom": r.get("uom"),
                "bid_uom": bid_raw.get("uom"),
                "hist_price": hist_price,
                "bid_price": bid_price,
                "qty_24mo": qty,
                "spend_24mo": spend_24,
                "notes": notes,
                "auto_suggestion": auto,
                "annotation": existing,   # may be None or a "needs_review" stub
            })

    # Sort by spend desc — highest-impact items first
    out.sort(key=lambda x: -x["spend_24mo"])
    return out


def get_uom_resolution_summary() -> dict:
    """Quick stats for the UOM resolution UI header.

    Returns:
        {"n_total": ..., "n_resolved": ..., "n_skipped": ...,
         "n_needs_review": ..., "n_remaining": ..., "n_auto_resolvable": ...}
    """
    queue = list_items_needing_uom_resolution()
    annotations = _STATE.get("uom_annotations") or {}
    n_resolved = sum(1 for a in annotations.values() if a["status"] == "resolved")
    n_skipped = sum(1 for a in annotations.values() if a["status"] == "skipped")
    n_needs_review = sum(1 for a in annotations.values() if a["status"] == "needs_review")
    n_auto = sum(1 for q in queue if q["auto_suggestion"] and q["auto_suggestion"].get("factor"))
    return {
        "n_total": len(queue) + n_resolved + n_skipped,
        "n_resolved": n_resolved,
        "n_skipped": n_skipped,
        "n_needs_review": n_needs_review,
        "n_remaining": len(queue),
        "n_auto_resolvable": n_auto,
    }


def remove_supplier_bid(supplier_name: str) -> bool:
    bids = _STATE.get("bids", {}) or {}
    if supplier_name in bids:
        del bids[supplier_name]
        _STATE["bids"] = bids
        log_event("remove_supplier_bid", "bid removed", related=supplier_name)
        return True
    return False


# ---------------------------------------------------------------------------
# Cross-supplier comparison + outlier detection
# ---------------------------------------------------------------------------

def _outlier_factor():
    return get_thresholds()["outlier_factor"]


def compute_comparison_matrix(included_keys=None) -> dict:
    """Build the cross-supplier comparison view for the loaded RFQ items
    against the loaded supplier bids.

    Returns:
        {
          "suppliers": [supplier_name, ...],   # ordered, columns of the matrix
          "rows": [
             {
               "rfq_key", "item_num", "part_number", "description", "mfg_name",
               "uom", "qty_24mo", "last_unit_price", "tier", "score",
               "bids": {supplier: {price, status, notes, alt_part, ...}, ...},
               "n_quoted": int,
               "coverage": "FULL" | "PARTIAL" | "SINGLE" | "NONE",
               "lowest_supplier": str | None,
               "lowest_price": float | None,
               "winner_savings_vs_history": float | None,
               "spread_pct": float | None,    # (max - min) / median
               "flags": [str, ...],
             }, ...
          ],
          "summary": {
             "n_items": int,
             "n_with_3plus_bids": int,
             "n_with_2_bids": int,
             "n_with_1_bid": int,
             "n_with_0_bids": int,
             "n_outliers_flagged": int,
             "total_lowest_value": float,
             "total_historical_value": float,
             "estimated_savings": float,
          }
        }
    """
    items = _STATE.get("items", [])
    bids_by_supplier = _STATE.get("bids", {}) or {}
    suppliers = list(bids_by_supplier.keys())

    if included_keys is not None:
        keep = set(included_keys)
        items = [it for it in items if it["item_num"] in keep]

    # Build a lookup: (supplier, rfq_key) -> bid record
    bid_lookup = {}
    for sup, parsed in bids_by_supplier.items():
        for b in parsed.get("bids", []):
            bid_lookup[(sup, b["rfq_key"])] = b

    rows = []
    n_full = n_two = n_one = n_zero = n_outliers = 0
    total_lowest = 0.0
    total_historical = 0.0
    est_savings = 0.0
    n_sup = len(suppliers)
    outlier_factor = _outlier_factor()

    for it in items:
        rfq_key = it["key"]
        per_sup = {}
        priced_values = []  # only PRICED-status bids contribute to outlier math
        for sup in suppliers:
            b = bid_lookup.get((sup, rfq_key))
            if b is None:
                per_sup[sup] = {"status": "MISSING", "price": None, "notes": "", "raw": None}
                continue
            per_sup[sup] = {
                "status": b["status"],
                "price": b["effective_price"],
                "notes": b["notes"],
                "raw": b,
            }
            if b["status"] in (BID_STATUS_PRICED, BID_STATUS_UOM_DISC, BID_STATUS_SUBSTITUTE) \
               and b["effective_price"] is not None and b["effective_price"] > 0:
                priced_values.append((sup, b["effective_price"]))

        n_quoted = len(priced_values)
        coverage = (
            "FULL" if n_quoted >= n_sup and n_sup >= 3
            else "PARTIAL" if n_quoted >= 2
            else "SINGLE" if n_quoted == 1
            else "NONE"
        )
        if n_quoted == 0: n_zero += 1
        elif n_quoted == 1: n_one += 1
        elif n_quoted == 2: n_two += 1
        else: n_full += 1

        lowest_supplier = None
        lowest_price = None
        spread_pct = None
        flags = []

        if priced_values:
            priced_values.sort(key=lambda t: t[1])
            lowest_supplier, lowest_price = priced_values[0]
            highest_price = priced_values[-1][1]
            if len(priced_values) >= 2:
                med = _median([p for _, p in priced_values])
                if med and med > 0:
                    spread_pct = (highest_price - lowest_price) / med * 100.0
                    # Outlier flag — any bid > Nx median or < median/N
                    for sup, p in priced_values:
                        if p >= med * outlier_factor:
                            flags.append(f"OUTLIER_HIGH:{sup}=${p:.2f}vs_median${med:.2f}")
                            n_outliers += 1
                        elif p <= med / outlier_factor:
                            flags.append(f"OUTLIER_LOW:{sup}=${p:.2f}vs_median${med:.2f}")
                            n_outliers += 1
            # vs historical
            hist = it.get("last_unit_price")
            if hist and hist > 0:
                if lowest_price >= hist * outlier_factor:
                    flags.append(f"ALL_BIDS_HIGH:lowest${lowest_price:.2f}_vs_paid${hist:.2f}")
                elif lowest_price <= hist / outlier_factor:
                    flags.append(f"BIG_SAVINGS:lowest${lowest_price:.2f}_vs_paid${hist:.2f}")

        # No-bid coverage flag (high-spend items with no quotes need follow-up)
        if n_quoted == 0:
            spend_24 = it.get("spend_24mo_actual") or 0
            if spend_24 >= 1000:
                flags.append(f"NO_BID_HIGH_SPEND:${spend_24:,.0f}_24mo_at_risk")

        # Award math — best-case scenario at lowest non-flagged bid
        qty_24 = it.get("qty_24mo") or 0
        hist_price = it.get("last_unit_price") or 0
        if lowest_price and qty_24:
            total_lowest += lowest_price * qty_24
        if hist_price and qty_24:
            total_historical += hist_price * qty_24

        row_record = {
            "rfq_key": rfq_key,
            "item_num": it["item_num"],
            "part_number": it.get("part_number") or "",
            "description": it["description"],
            "mfg_name": it["mfg_name"],
            "uom": it["uom"],
            "qty_24mo": qty_24,
            "last_unit_price": hist_price,
            "tier": it.get("tier"),
            "score": it.get("score"),
            "bids": per_sup,
            "n_quoted": n_quoted,
            "coverage": coverage,
            "lowest_supplier": lowest_supplier,
            "lowest_price": lowest_price,
            "spread_pct": spread_pct,
            "flags": flags,
        }
        rec = recommend_for_item(row_record)
        row_record["recommendation"] = rec["recommendation"]
        row_record["recommendation_target"] = rec["target_supplier"]
        row_record["recommendation_reason"] = rec["reason"]
        row_record["secondary_actions"] = rec["secondary_actions"]
        rows.append(row_record)

    est_savings = total_historical - total_lowest

    # Rec-distribution summary
    rec_counts = {}
    for r in rows:
        rec_counts[r["recommendation"]] = rec_counts.get(r["recommendation"], 0) + 1

    return {
        "suppliers": suppliers,
        "rows": rows,
        "summary": {
            "n_items": len(rows),
            "n_with_3plus_bids": n_full,
            "n_with_2_bids": n_two,
            "n_with_1_bid": n_one,
            "n_with_0_bids": n_zero,
            "n_outliers_flagged": n_outliers,
            "total_lowest_value": total_lowest,
            "total_historical_value": total_historical,
            "estimated_savings": est_savings,
            "recommendation_counts": rec_counts,
        },
    }


# ---------------------------------------------------------------------------
# Difficulty snapshot history — appended each extract for period-end reporting.
# Stored as a list on _STATE so the save-state JSON captures the full series.
# ---------------------------------------------------------------------------

def record_difficulty_snapshot(difficulty: dict) -> None:
    if not isinstance(difficulty, dict) or not difficulty.get("score"):
        return
    history = _STATE.get("difficulty_history", [])
    snap = {
        "snapshot_at": difficulty.get("snapshot_at") or datetime.now().isoformat(),
        "score": difficulty.get("score"),
        "level": difficulty.get("level"),
        "summary": difficulty.get("summary"),
        "signals": dict(difficulty.get("signals") or {}),
    }
    # Don't duplicate — if last snapshot has the same signals, skip
    if history:
        last = history[-1]
        if last.get("score") == snap["score"] and last.get("signals") == snap["signals"]:
            return
    history.append(snap)
    # Cap at 50 snapshots so the save state doesn't grow unbounded
    if len(history) > 50:
        history = history[-50:]
    _STATE["difficulty_history"] = history


def list_difficulty_history() -> list:
    return list(_STATE.get("difficulty_history", []))


def clear_difficulty_history() -> None:
    _STATE["difficulty_history"] = []


# ---------------------------------------------------------------------------
# Recommendation engine — 5-tier per-item recommendations with mandatory reasons.
#
# Categories (deterministic, no AI):
#   ACCEPT             — clear winner, no exceptions, savings exceed threshold
#   PUSH_BACK          — quote is high vs baseline OR vs lowest competitor
#   ASK_CLARIFICATION  — UOM mismatch, missing data, alternate offered, etc.
#   EXCLUDE            — no bid, invalid price, item identity changed
#   MANUAL_REVIEW      — data conflict, low usage, outlier, ties
#
# Every recommendation carries a `reason` string with concrete numbers.
# ---------------------------------------------------------------------------
RECOMMENDATION_ACCEPT      = "ACCEPT"
RECOMMENDATION_PUSH_BACK   = "PUSH_BACK"
RECOMMENDATION_CLARIFY     = "ASK_CLARIFICATION"
RECOMMENDATION_EXCLUDE     = "EXCLUDE"
RECOMMENDATION_MANUAL      = "MANUAL_REVIEW"


def _fmt_pct(x):
    return f"{x*100:.0f}%" if x is not None else "—"


def _fmt_money(x):
    return f"${x:,.2f}" if x is not None else "—"


def recommend_for_item(matrix_row: dict) -> dict:
    """Given one row of the comparison matrix, return:
        {
          "recommendation": ACCEPT | PUSH_BACK | ASK_CLARIFICATION | EXCLUDE | MANUAL_REVIEW,
          "target_supplier": str | None,
          "reason": str,
          "secondary_actions": [str, ...],   # additional follow-ups (e.g. push back AND ask MOQ)
        }

    Decision rules (in priority order — first match wins):
      1. n_quoted == 0 → EXCLUDE (no bid from anyone)
      2. ALL bids have status NEED_INFO → ASK_CLARIFICATION
      3. Any UOM_DISC on the lowest bid → ASK_CLARIFICATION (verify UOM first)
      4. Any SUBSTITUTE on the lowest bid → ASK_CLARIFICATION (validate alt part)
      5. Lowest bid >> baseline (>pushback_threshold AND no cheaper alt) → PUSH_BACK
      6. Lowest bid is an outlier among bids OR vs history → MANUAL_REVIEW
      7. Lowest bid available, savings >= min_savings_pct_to_switch → ACCEPT
      8. Otherwise → MANUAL_REVIEW (low savings; not worth switching cost)
    """
    th = get_thresholds()
    pushback_pct = th["pushback_threshold_pct"]
    min_switch_pct = th["min_savings_pct_to_switch"]

    n_quoted = matrix_row.get("n_quoted", 0)
    bids = matrix_row.get("bids") or {}
    flags = matrix_row.get("flags") or []
    lowest_supplier = matrix_row.get("lowest_supplier")
    lowest_price = matrix_row.get("lowest_price")
    historical = matrix_row.get("last_unit_price") or 0
    qty = matrix_row.get("qty_24mo") or 0

    secondary = []

    # Rule 1: nobody bid
    if n_quoted == 0:
        return {
            "recommendation": RECOMMENDATION_EXCLUDE,
            "target_supplier": None,
            "reason": "No supplier quoted this item. Either follow up with all suppliers or drop the item from the RFQ.",
            "secondary_actions": [],
        }

    # Rule 2: all bids need info
    statuses = [b.get("status") for b in bids.values() if b.get("status") not in (None, "MISSING")]
    if statuses and all(s == BID_STATUS_NEED_INFO for s in statuses):
        return {
            "recommendation": RECOMMENDATION_CLARIFY,
            "target_supplier": None,
            "reason": "All responding suppliers asked for more information. Provide additional spec or sample before re-quoting.",
            "secondary_actions": [],
        }

    if not lowest_supplier or lowest_price is None:
        return {
            "recommendation": RECOMMENDATION_MANUAL,
            "target_supplier": None,
            "reason": "No usable bid (all responses are no-bid / need-info / invalid). Manual review.",
            "secondary_actions": [],
        }

    lowest_bid = bids.get(lowest_supplier) or {}
    lowest_status = lowest_bid.get("status")
    lowest_raw = lowest_bid.get("raw") or {}

    # Rule 3: UOM discrepancy on the lowest bid
    if lowest_status == BID_STATUS_UOM_DISC:
        return {
            "recommendation": RECOMMENDATION_CLARIFY,
            "target_supplier": lowest_supplier,
            "reason": f"Lowest bid is from {lowest_supplier} at {_fmt_money(lowest_price)} but they flagged a UOM discrepancy. Confirm UOM before treating as real savings.",
            "secondary_actions": ["UOM_VERIFY"],
        }

    # Rule 4: substitute offered on the lowest bid
    if lowest_status == BID_STATUS_SUBSTITUTE:
        sub_part = lowest_raw.get("sub_part") or "(unspecified)"
        return {
            "recommendation": RECOMMENDATION_CLARIFY,
            "target_supplier": lowest_supplier,
            "reason": f"Lowest bid from {lowest_supplier} is for an alternate part ({sub_part}). Validate that the substitute meets spec before awarding.",
            "secondary_actions": ["SPEC_VERIFY"],
        }

    # Rule 5: all bids high vs history (push back)
    if historical > 0:
        gap_vs_hist = (lowest_price - historical) / historical
        if gap_vs_hist >= pushback_pct:
            return {
                "recommendation": RECOMMENDATION_PUSH_BACK,
                "target_supplier": lowest_supplier,
                "reason": f"Lowest bid ({_fmt_money(lowest_price)} from {lowest_supplier}) is {_fmt_pct(gap_vs_hist)} above the historical paid price ({_fmt_money(historical)}). Push back before accepting.",
                "secondary_actions": [],
            }

    # Rule 6: outlier flags
    if flags:
        # Detect specific flag types
        outlier_flag = any("OUTLIER" in f for f in flags)
        all_high_flag = any("ALL_BIDS_HIGH" in f for f in flags)
        big_savings_flag = any("BIG_SAVINGS" in f for f in flags)
        if all_high_flag:
            return {
                "recommendation": RECOMMENDATION_PUSH_BACK,
                "target_supplier": lowest_supplier,
                "reason": f"All bids are well above historical paid price. Lowest is {_fmt_money(lowest_price)} from {lowest_supplier} — push back across the board or drop the item.",
                "secondary_actions": [],
            }
        if big_savings_flag:
            return {
                "recommendation": RECOMMENDATION_MANUAL,
                "target_supplier": lowest_supplier,
                "reason": f"Lowest bid ({_fmt_money(lowest_price)} from {lowest_supplier}) is dramatically below historical paid ({_fmt_money(historical)}). Verify it's the same item / spec / UOM before accepting.",
                "secondary_actions": ["SPEC_VERIFY", "UOM_VERIFY"],
            }
        if outlier_flag:
            return {
                "recommendation": RECOMMENDATION_MANUAL,
                "target_supplier": lowest_supplier,
                "reason": f"Lowest bid ({_fmt_money(lowest_price)} from {lowest_supplier}) is statistically an outlier vs the other bids. Verify before awarding.",
                "secondary_actions": [],
            }

    # Rule 7: clear acceptance
    if historical > 0:
        savings_pct = (historical - lowest_price) / historical
        savings_total = (historical - lowest_price) * qty
        if savings_pct >= min_switch_pct:
            return {
                "recommendation": RECOMMENDATION_ACCEPT,
                "target_supplier": lowest_supplier,
                "reason": f"Award to {lowest_supplier} at {_fmt_money(lowest_price)}. Saves {_fmt_pct(savings_pct)} vs historical ({_fmt_money(historical)}) — about {_fmt_money(savings_total)} on 24-mo qty.",
                "secondary_actions": [],
            }
        else:
            # Savings below the switch threshold — recommend manual review
            return {
                "recommendation": RECOMMENDATION_MANUAL,
                "target_supplier": lowest_supplier,
                "reason": f"Lowest bid {_fmt_money(lowest_price)} from {lowest_supplier} only saves {_fmt_pct(savings_pct)} vs historical. Below the {_fmt_pct(min_switch_pct)} switching threshold — may not be worth a supplier change.",
                "secondary_actions": [],
            }

    # No historical baseline — accept the lowest bid but note it
    return {
        "recommendation": RECOMMENDATION_ACCEPT,
        "target_supplier": lowest_supplier,
        "reason": f"Award to {lowest_supplier} at {_fmt_money(lowest_price)} (no historical baseline to compare against).",
        "secondary_actions": [],
    }


# ---------------------------------------------------------------------------
# Consolidation analysis — Ryan's actual award strategy.
#
# Default: consolidate to ONE supplier (operational simplicity beats per-line
# lowest pricing when the savings are small). Then carve out a limited set
# of "extreme exceptions" — items where another supplier saves enough on a
# single item to justify splitting the award.
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Configurable project thresholds
#
# All tunable knobs live in one dict so the UI can read/write them and they
# get persisted in the save state. Defaults are sensible for MRO RFQs but
# every project / supplier mix should be allowed to override.
# ---------------------------------------------------------------------------
DEFAULT_THRESHOLDS = {
    # Carve-out logic is OR between % and $/yr — a carve fires when EITHER the
    # percent savings is structural (typical industry default 15-25%) OR the
    # absolute annual dollar savings is large enough to justify a separate PO
    # (typical $1K-5K/yr). This dual rule under-carves on long-tail items
    # (where 50% of $200/yr is admin overhead) and catches high-volume items
    # where 5% of $50K/yr is real money.
    "carve_out_min_savings_pct": 0.20,                # 20% — fires on long-tail anomalies
    "carve_out_min_savings_annual_dollar": 3000.0,    # $3K/yr — fires on high-volume rate plays
    "outlier_factor": 3.0,                   # bid is outlier if >= Nx median or <= 1/N of median
    "spike_factor": 1.5,                     # latest price is "spike" if >= 1.5x 90-day median
    "uom_suspect_ratio": 20.0,               # carve-out price ratio >= this → mark UOM-verify
    "min_savings_pct_to_switch": 0.05,       # below 5% savings, prefer incumbent (avoid switching cost)
    "pushback_threshold_pct": 0.10,          # quote >10% above baseline → push back
    "max_acceptable_lead_time_days": 60,     # bids with lead time > this are flagged
    "max_acceptable_moq_vs_annual": 1.0,     # bids with MOQ > annual qty are flagged
    "min_quote_validity_days": 30,           # quotes valid <30 days are flagged
    "min_spend_for_review": 100.0,           # below this 24mo spend, item not worth manual review
    "high_spend_no_bid_threshold": 1000.0,   # no-bid items with >$X 24mo spend get follow-up flag
}


def get_thresholds() -> dict:
    """Return current thresholds (defaults merged with any saved overrides)."""
    overrides = _STATE.get("thresholds", {})
    out = dict(DEFAULT_THRESHOLDS)
    out.update(overrides)
    return out


def set_thresholds(updates: dict) -> dict:
    """Update one or more thresholds. Returns the new merged state."""
    current = _STATE.get("thresholds", {}) or {}
    keys_changed = []
    for k, v in (updates or {}).items():
        if k in DEFAULT_THRESHOLDS and current.get(k) != v:
            current[k] = v
            keys_changed.append(k)
    _STATE["thresholds"] = current
    if keys_changed:
        log_event("set_thresholds", f"changed: {', '.join(keys_changed)}")
    return get_thresholds()


def reset_thresholds() -> dict:
    _STATE["thresholds"] = {}
    return get_thresholds()


# ---------------------------------------------------------------------------
# User profile — operator name + contact info used in outbound xlsx files.
# Persisted with the session save state so different operators on the same
# machine can each have their own.
# ---------------------------------------------------------------------------

DEFAULT_USER_PROFILE = {
    "name": "",
    "email": "",
    "title": "Procurement Analyst",
    "company": "Andersen",
}


def get_user_profile() -> dict:
    p = _STATE.get("user_profile", {}) or {}
    out = dict(DEFAULT_USER_PROFILE)
    out.update(p)
    return out


def set_user_profile(updates: dict) -> dict:
    cur = _STATE.get("user_profile", {}) or {}
    for k in ("name", "email", "title", "company"):
        if k in (updates or {}):
            cur[k] = (updates[k] or "").strip()
    _STATE["user_profile"] = cur
    log_event("set_user_profile", "user profile updated")
    return get_user_profile()


# Legacy alias (kept for back-compat with existing call sites; use thresholds)
DEFAULT_CARVE_OUT_THRESHOLD = 0.30


def compute_consolidation_analysis(included_keys=None, carve_threshold: float = None,
                                   uom_suspect_ratio: float = None,
                                   carve_threshold_dollar: float = None) -> dict:
    """Rank suppliers as candidate consolidation winners and quantify the
    impact of carving out exceptions where someone else is meaningfully
    cheaper on individual items.

    Returns:
      {
        "candidates": [
          {
            "supplier": str,
            "n_items_quoted": int,
            "pct_items_quoted": float,
            "n_items_lowest": int,
            "pct_items_lowest": float,
            "consolidation_value": float,    # award everything they quoted, at their prices
            "items_not_quoted": int,         # items in the RFQ this supplier didn't quote
            "consolidation_coverage_pct": float,   # share of total RFQ qty they cover
            "missing_value_at_history": float,     # value of items they didn't quote, at hist price
          }, ... sorted by consolidation_value asc (cheapest first)
        ],
        "winner": {
          "supplier": str,
          "consolidation_value": float,
          "carve_outs": [
            {
              "rfq_key", "item_num", "description",
              "winner_price", "winner_supplier",
              "carve_supplier", "carve_price",
              "qty_24mo", "savings_per_unit", "savings_total",
              "savings_pct"
            }, ... sorted by savings_total desc
          ],
          "carve_out_savings_total": float,
          "items_winner_didnt_quote": [
            {"rfq_key", "item_num", "description", "qty_24mo",
             "best_alt_supplier", "best_alt_price",
             "value_at_best_alt", "value_at_history"}, ...
          ],
          "final_award_value": float,    # winner-base − carveout-savings + missing-items-at-best-alt
          "final_award_value_vs_history": float,  # vs historical at last_unit_price
        } | None,
        "summary": {
          "n_suppliers": int,
          "n_items": int,
          "carve_out_threshold_pct": float,
        }
      }
    """
    th = get_thresholds()
    if carve_threshold is None:
        carve_threshold = th["carve_out_min_savings_pct"]
    if carve_threshold_dollar is None:
        carve_threshold_dollar = th["carve_out_min_savings_annual_dollar"]
    if uom_suspect_ratio is None:
        uom_suspect_ratio = th["uom_suspect_ratio"]

    items = _STATE.get("items", [])
    bids_by_supplier = _STATE.get("bids", {}) or {}
    suppliers = list(bids_by_supplier.keys())

    if included_keys is not None:
        keep = set(included_keys)
        items = [it for it in items if it["item_num"] in keep]

    # Build (supplier, rfq_key) → bid lookup
    bid_lookup = {}
    for sup, parsed in bids_by_supplier.items():
        for b in parsed.get("bids", []):
            if b["status"] in (BID_STATUS_PRICED, BID_STATUS_UOM_DISC, BID_STATUS_SUBSTITUTE) \
               and b["effective_price"] is not None and b["effective_price"] > 0:
                bid_lookup[(sup, b["rfq_key"])] = b

    n_items = len(items)
    candidates = []
    for sup in suppliers:
        n_quoted = 0
        n_lowest = 0
        consol_value = 0.0
        missing_at_history = 0.0
        items_not_quoted = 0
        for it in items:
            qty = it.get("qty_24mo") or 0
            sup_bid = bid_lookup.get((sup, it["key"]))
            if sup_bid:
                n_quoted += 1
                consol_value += sup_bid["effective_price"] * qty
                # Are they the lowest among all suppliers for this item?
                others = [bid_lookup.get((s, it["key"])) for s in suppliers]
                priced = [b["effective_price"] for b in others if b]
                if priced and sup_bid["effective_price"] == min(priced):
                    n_lowest += 1
            else:
                items_not_quoted += 1
                hist = it.get("last_unit_price") or 0
                if hist > 0:
                    missing_at_history += hist * qty

        candidates.append({
            "supplier": sup,
            "n_items_quoted": n_quoted,
            "pct_items_quoted": (100.0 * n_quoted / n_items) if n_items else 0.0,
            "n_items_lowest": n_lowest,
            "pct_items_lowest": (100.0 * n_lowest / n_quoted) if n_quoted else 0.0,
            "consolidation_value": consol_value,
            "items_not_quoted": items_not_quoted,
            "consolidation_coverage_pct": (100.0 * n_quoted / n_items) if n_items else 0.0,
            "missing_value_at_history": missing_at_history,
        })

    # Two-tier ranking: COVERAGE FIRST, then aggregate value.
    # Without the coverage tier, a supplier who only quotes 1 cheap item beats
    # a supplier who quotes everything (their aggregate is tiny). The realistic
    # consolidation question is "who can we award the most to?" — that's
    # bounded by what they actually quoted. Top tier = suppliers with at least
    # `min_coverage_pct` of items quoted; among them, lowest aggregate wins.
    # Falls through to the legacy ranking if nobody clears the bar.
    min_coverage_pct = 50.0   # quote at least half the RFQ to be a serious consolidation candidate
    high_coverage = [c for c in candidates if c["pct_items_quoted"] >= min_coverage_pct]
    if high_coverage:
        # Within the top tier, lowest value wins; tiebreak: more items.
        high_coverage.sort(key=lambda c: (c["consolidation_value"], -c["n_items_quoted"]))
        low_coverage = [c for c in candidates if c["pct_items_quoted"] < min_coverage_pct]
        low_coverage.sort(key=lambda c: (c["consolidation_value"], -c["n_items_quoted"]))
        candidates = high_coverage + low_coverage
    else:
        candidates.sort(key=lambda c: (c["consolidation_value"], -c["n_items_quoted"]))

    winner_block = None
    if candidates:
        winner = candidates[0]
        winner_sup = winner["supplier"]
        carve_outs = []
        items_not_quoted_details = []
        carve_savings_total = 0.0
        items_at_best_alt = 0.0

        for it in items:
            qty = it.get("qty_24mo") or 0
            winner_bid = bid_lookup.get((winner_sup, it["key"]))
            other_bids = [
                (s, bid_lookup[(s, it["key"])]["effective_price"])
                for s in suppliers
                if s != winner_sup and (s, it["key"]) in bid_lookup
            ]
            if winner_bid:
                # Consolidation winner quoted — check if a carve-out is justified
                w_price = winner_bid["effective_price"]
                if other_bids:
                    other_bids.sort(key=lambda t: t[1])
                    cheapest_other_sup, cheapest_other_price = other_bids[0]
                    savings_per_unit = w_price - cheapest_other_price
                    savings_total = savings_per_unit * qty                  # 24-mo total savings at the qty windowed in the RFQ
                    annual_savings = savings_total / 2.0 if qty else 0.0    # 24-mo qty halved → ~annual run-rate
                    pct_savings = (savings_per_unit / w_price) if w_price else 0.0
                    fires_pct = pct_savings >= carve_threshold
                    fires_dollar = annual_savings >= carve_threshold_dollar
                    if fires_pct or fires_dollar:
                        # Record which rule fired (or both) — drives the matrix
                        # tooltip / decision-log explanation. Both is common
                        # on big high-volume rate plays.
                        if fires_pct and fires_dollar:
                            carve_rule_fired = "BOTH"
                        elif fires_pct:
                            carve_rule_fired = "PCT"
                        else:
                            carve_rule_fired = "DOLLAR"
                        # UOM-discrepancy guard — if EITHER side has a UOM
                        # warning OR the price ratio is extreme (>20×), flag
                        # the carve-out as needing UOM verification before
                        # being trusted as real savings. Fastenal's "per each
                        # vs McMaster per package" notes are the canonical case.
                        carve_bid_record = bid_lookup.get((cheapest_other_sup, it["key"]))
                        winner_bid_record = bid_lookup.get((winner_sup, it["key"]))
                        carve_status = (carve_bid_record or {}).get("status")
                        winner_status = (winner_bid_record or {}).get("status")
                        verify_uom = (
                            carve_status == BID_STATUS_UOM_DISC or
                            winner_status == BID_STATUS_UOM_DISC or
                            (cheapest_other_price > 0 and w_price / cheapest_other_price >= uom_suspect_ratio)
                        )
                        carve_outs.append({
                            "rfq_key": it["key"],
                            "item_num": it["item_num"],
                            "description": it["description"],
                            "qty_24mo": qty,
                            "winner_supplier": winner_sup,
                            "winner_price": w_price,
                            "carve_supplier": cheapest_other_sup,
                            "carve_price": cheapest_other_price,
                            "savings_per_unit": savings_per_unit,
                            "savings_total": savings_total,
                            "savings_annual": annual_savings,
                            "savings_pct": (savings_per_unit / w_price * 100.0) if w_price else 0.0,
                            "carve_rule_fired": carve_rule_fired,   # "PCT", "DOLLAR", or "BOTH"
                            "verify_uom": verify_uom,
                            "carve_notes": (carve_bid_record or {}).get("notes", ""),
                            "winner_notes": (winner_bid_record or {}).get("notes", ""),
                        })
                        # Only count savings from carve-outs that DON'T need UOM verify
                        if not verify_uom:
                            carve_savings_total += savings_total
            else:
                # Winner didn't quote — pick best alternative
                if other_bids:
                    other_bids.sort(key=lambda t: t[1])
                    best_alt_sup, best_alt_price = other_bids[0]
                    items_at_best_alt += best_alt_price * qty
                    items_not_quoted_details.append({
                        "rfq_key": it["key"],
                        "item_num": it["item_num"],
                        "description": it["description"],
                        "qty_24mo": qty,
                        "best_alt_supplier": best_alt_sup,
                        "best_alt_price": best_alt_price,
                        "value_at_best_alt": best_alt_price * qty,
                        "value_at_history": (it.get("last_unit_price") or 0) * qty,
                    })
                else:
                    # Nobody quoted this item — flag for follow-up
                    items_not_quoted_details.append({
                        "rfq_key": it["key"],
                        "item_num": it["item_num"],
                        "description": it["description"],
                        "qty_24mo": qty,
                        "best_alt_supplier": None,
                        "best_alt_price": None,
                        "value_at_best_alt": 0.0,
                        "value_at_history": (it.get("last_unit_price") or 0) * qty,
                    })

        carve_outs.sort(key=lambda c: c["savings_total"], reverse=True)

        # Final award value = winner's full quote − carve savings + items at best alt
        final_award_value = winner["consolidation_value"] - carve_savings_total + items_at_best_alt

        # vs historical at last_unit_price (for items either supplier covers)
        historical_value = 0.0
        for it in items:
            qty = it.get("qty_24mo") or 0
            hist = it.get("last_unit_price") or 0
            historical_value += hist * qty

        winner_block = {
            "supplier": winner_sup,
            "consolidation_value": winner["consolidation_value"],
            "carve_outs": carve_outs,
            "carve_out_savings_total": carve_savings_total,
            "items_winner_didnt_quote": items_not_quoted_details,
            "n_items_winner_didnt_quote": len(items_not_quoted_details),
            "items_at_best_alt_value": items_at_best_alt,
            "final_award_value": final_award_value,
            "historical_value": historical_value,
            "savings_vs_history": historical_value - final_award_value,
        }

    return {
        "candidates": candidates,
        "winner": winner_block,
        "summary": {
            "n_suppliers": len(suppliers),
            "n_items": n_items,
            "carve_out_threshold_pct": carve_threshold * 100.0,
            "carve_out_threshold_dollar": carve_threshold_dollar,
        },
    }


# ---------------------------------------------------------------------------
# Supplier follow-up xlsx generator — multi-tab pushback / clarification packet
#
# Per the brief's recommended structure:
#   1. Summary
#   2. Items Needing Price Review  (PUSH_BACK recommendations)
#   3. Missing Information         (NEED_INFO bids)
#   4. UOM or MOQ Exceptions       (UOM_DISC, MOQ flags)
#   5. Alternate Parts             (SUBSTITUTE bids)
#   6. No-Bids                     (NO_BID lines, ask why + pricing window)
#   7. Full Quote Detail           (every line this supplier responded to)
#
# Template-based prose only — no AI. Per the cross-supplier isolation rule,
# this xlsx contains ONLY the named supplier's own data — never another
# supplier's bid prices or our internal target/cost/margin.
# ---------------------------------------------------------------------------

PUSHBACK_TEMPLATES = {
    "price_too_high_vs_history": (
        "Your quoted price for {item_desc} is {gap_pct} above our historical "
        "paid price of {hist_price} per {uom}. Annualized volume is "
        "approximately {qty} units. Please review whether you have room to "
        "improve pricing on this item."
    ),
    "price_too_high_vs_competition": (
        "Your quoted price for {item_desc} is {gap_pct} above the lowest "
        "qualified quote we received. We would like to consolidate with you "
        "if possible — please confirm whether you can revisit pricing."
    ),
    "missing_info_request": (
        "We were unable to evaluate your quote for {item_desc} because "
        "additional information was needed. Please provide unit price, UOM, "
        "and confirmation that the part offered matches our specification."
    ),
    "uom_clarification": (
        "Your quote for {item_desc} indicates a UOM mismatch with our "
        "purchasing UOM ({our_uom}). Please confirm your quoted UOM and "
        "whether the price needs to be normalized to our UOM."
    ),
    "moq_concern": (
        "The minimum order quantity on your quote for {item_desc} ({moq}) "
        "exceeds our typical annual consumption ({qty}). Please confirm "
        "whether a smaller MOQ is available."
    ),
    "alternate_part_review": (
        "You quoted an alternate part ({alt_part}) for our requested item "
        "{item_desc}. Please confirm form/fit/function equivalence and "
        "whether your alternate is in stock and current."
    ),
    "no_bid_clarification": (
        "You did not bid on {item_desc} (note: \"{notes}\"). If this item "
        "is something you can supply with additional information or a longer "
        "lead time, please let us know."
    ),
}


def gen_supplier_followup_xlsx(supplier_name: str, included_keys=None) -> bytes:
    """Build the follow-up packet xlsx for one supplier.

    Strict isolation: only this supplier's own bid data appears.
    No other supplier's prices, no internal target/cost/margin.
    """
    if not supplier_name:
        raise ValueError("supplier_name required")
    bids_by_supplier = _STATE.get("bids", {}) or {}
    parsed = bids_by_supplier.get(supplier_name)
    if not parsed:
        raise ValueError(f"No loaded bids for supplier {supplier_name!r}")

    # Match each bid back to RFQ items so we can include item context
    items = _STATE.get("items", [])
    if included_keys is not None:
        keep = set(included_keys)
        items = [it for it in items if it["item_num"] in keep]
    item_by_key = {it["key"]: it for it in items}

    matrix = compute_comparison_matrix(included_keys=included_keys)
    rows_by_key = {r["rfq_key"]: r for r in matrix["rows"]}

    th = get_thresholds()
    pushback_pct = th["pushback_threshold_pct"]

    # Categorize this supplier's bids by recommendation tab
    bids_for_review = []     # PUSH_BACK
    bids_missing_info = []   # NEED_INFO
    bids_uom_or_moq = []     # UOM_DISC or MOQ flagged
    bids_alternates = []     # SUBSTITUTE
    bids_no_bid = []         # NO_BID
    bids_all = []            # everything (Full Quote Detail tab)

    for b in parsed.get("bids", []):
        rfq_key = b["rfq_key"]
        item = item_by_key.get(rfq_key)
        row = rows_by_key.get(rfq_key)
        rec_target_is_them = (row and row.get("recommendation_target") == supplier_name)
        rec = row.get("recommendation") if row else None

        ctx = {
            "rfq_key": rfq_key,
            "item_num": (item["item_num"] if item else b.get("item_num", "")),
            "part_number": (item.get("part_number") if item else b.get("part_number", "")),
            "description": (item["description"] if item else b.get("description", "")),
            "mfg_name": (item["mfg_name"] if item else b.get("mfg_name", "")),
            "our_uom": (item["uom"] if item else b.get("uom", "")),
            "qty_24mo": (item.get("qty_24mo") if item else (b.get("qty") or 0)),
            "historical_price": (item.get("last_unit_price") if item else None),
            "your_price": b.get("effective_price"),
            "your_quoted_uom": b.get("uom", ""),
            "your_notes": b.get("notes", ""),
            "your_part": b.get("exact_part") or b.get("part_number", ""),
            "your_alt_part": b.get("sub_part") or "",
            "your_alt_desc": b.get("sub_desc") or "",
            "status": b["status"],
        }
        bids_all.append(ctx)

        # NO_BID
        if b["status"] == BID_STATUS_NO_BID:
            bids_no_bid.append(ctx)
            continue
        # NEED_INFO
        if b["status"] == BID_STATUS_NEED_INFO:
            bids_missing_info.append(ctx)
            continue
        # UOM discrepancy
        if b["status"] == BID_STATUS_UOM_DISC:
            bids_uom_or_moq.append(ctx)
            continue
        # SUBSTITUTE
        if b["status"] == BID_STATUS_SUBSTITUTE:
            bids_alternates.append(ctx)
            continue
        # PRICED — check if pushback applies (vs history OR vs competition)
        their_price = b["effective_price"] or 0
        hist = ctx["historical_price"] or 0
        # Vs history
        push_reason = None
        if hist > 0 and their_price > 0:
            gap = (their_price - hist) / hist
            if gap >= pushback_pct:
                push_reason = ("vs_history", gap)
        # Vs competition (do not name the cheaper supplier — isolation rule)
        if row and not push_reason:
            lowest_price = row.get("lowest_price")
            lowest_supplier = row.get("lowest_supplier")
            if (lowest_price and lowest_supplier and lowest_supplier != supplier_name
                    and lowest_price > 0 and their_price > lowest_price * (1 + pushback_pct)):
                comp_gap = (their_price - lowest_price) / lowest_price
                push_reason = ("vs_competition", comp_gap)
        if push_reason:
            ctx["pushback_kind"] = push_reason[0]
            ctx["pushback_gap_pct"] = push_reason[1]
            bids_for_review.append(ctx)

    # ---- Build the workbook ----
    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="0a0e1a")
    BAND_FILL = PatternFill("solid", fgColor="2a3658")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    LABEL_FONT = Font(bold=True, color="b4c0d4", size=10)
    BANNER_FONT = Font(bold=True, color="ffb733", size=14)

    # ---- TAB 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"RFQ Follow-Up — {supplier_name}"])
    ws["A1"].font = BANNER_FONT
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["This packet contains items from your prior bid response that we'd like you to revisit."])
    ws.append(["Each tab covers a different category. Please review and respond at your convenience."])
    ws.append([])
    ws.append(["Section", "Item count", "Notes"])
    for c in ws[7]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.append(["Items Needing Price Review", len(bids_for_review), "Quoted but priced high vs our baseline or vs the qualified market"])
    ws.append(["Missing Information", len(bids_missing_info), "Items where you indicated 'need more information'"])
    ws.append(["UOM / MOQ Exceptions", len(bids_uom_or_moq), "UOM mismatch or MOQ exceeds typical usage"])
    ws.append(["Alternate Parts", len(bids_alternates), "Where you offered a substitute part"])
    ws.append(["No-Bids", len(bids_no_bid), "Items you declined to quote — please confirm category & reason"])
    ws.append(["Full Quote Detail", len(bids_all), "Every line you responded to (reference)"])
    autosize(ws)

    # ---- TAB 2: Items Needing Price Review ----
    ws2 = wb.create_sheet("Items Needing Price Review")
    headers = ["Item #", "Description", "Manufacturer", "Our UOM", "Annual Qty",
               "Your Price", "Your UOM", "Reason for review", "Pushback message",
               "Your revised price", "Your notes"]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    yellow = PatternFill("solid", fgColor="FFF59D")
    for ctx in bids_for_review:
        kind = ctx.get("pushback_kind", "")
        gap = ctx.get("pushback_gap_pct", 0)
        if kind == "vs_history":
            reason_short = f"{gap*100:.0f}% above our prior paid price"
            tmpl = PUSHBACK_TEMPLATES["price_too_high_vs_history"].format(
                item_desc=(ctx["description"] or ctx["item_num"])[:80],
                gap_pct=f"{gap*100:.0f}%",
                hist_price=_fmt_money(ctx["historical_price"]),
                uom=ctx["our_uom"] or "unit",
                qty=f"{(ctx['qty_24mo'] or 0):,.0f}",
            )
        else:
            reason_short = f"{gap*100:.0f}% above the lowest qualified quote"
            tmpl = PUSHBACK_TEMPLATES["price_too_high_vs_competition"].format(
                item_desc=(ctx["description"] or ctx["item_num"])[:80],
                gap_pct=f"{gap*100:.0f}%",
            )
        ws2.append([
            ctx["item_num"], ctx["description"], ctx["mfg_name"], ctx["our_uom"],
            ctx["qty_24mo"], ctx["your_price"], ctx["your_quoted_uom"],
            reason_short, tmpl, "", "",
        ])
    # Yellow-fill the response columns (J=10, K=11) so they stand out
    for r in ws2.iter_rows(min_row=2, min_col=10, max_col=11):
        for c in r:
            c.fill = yellow
    for r in ws2.iter_rows(min_row=2, min_col=6, max_col=6):
        for c in r:
            c.number_format = "$#,##0.00"
    for r in ws2.iter_rows(min_row=2, min_col=10, max_col=10):
        for c in r:
            c.number_format = "$#,##0.00"
    autosize(ws2)
    ws2.freeze_panes = "A2"

    # ---- TAB 3: Missing Information ----
    ws3 = wb.create_sheet("Missing Information")
    ws3.append(["Item #", "Description", "Manufacturer", "Our UOM", "Annual Qty",
                "Your note", "What we need", "Your follow-up price", "Your follow-up UOM", "Your notes"])
    for c in ws3[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for ctx in bids_missing_info:
        ws3.append([
            ctx["item_num"], ctx["description"], ctx["mfg_name"], ctx["our_uom"], ctx["qty_24mo"],
            ctx["your_notes"],
            PUSHBACK_TEMPLATES["missing_info_request"].format(item_desc=(ctx["description"] or ctx["item_num"])[:80]),
            "", "", "",
        ])
    for r in ws3.iter_rows(min_row=2, min_col=8, max_col=10):
        for c in r:
            c.fill = yellow
    for r in ws3.iter_rows(min_row=2, min_col=8, max_col=8):
        for c in r:
            c.number_format = "$#,##0.00"
    autosize(ws3)
    ws3.freeze_panes = "A2"

    # ---- TAB 4: UOM / MOQ Exceptions ----
    ws4 = wb.create_sheet("UOM and MOQ Exceptions")
    ws4.append(["Item #", "Description", "Our UOM", "Your quoted UOM", "Your price",
                "Your notes", "What we're asking", "Your confirmed UOM", "Your normalized price", "Your notes"])
    for c in ws4[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for ctx in bids_uom_or_moq:
        ws4.append([
            ctx["item_num"], ctx["description"], ctx["our_uom"], ctx["your_quoted_uom"],
            ctx["your_price"], ctx["your_notes"],
            PUSHBACK_TEMPLATES["uom_clarification"].format(
                item_desc=(ctx["description"] or ctx["item_num"])[:80],
                our_uom=ctx["our_uom"] or "(unknown)",
            ),
            "", "", "",
        ])
    for r in ws4.iter_rows(min_row=2, min_col=8, max_col=10):
        for c in r:
            c.fill = yellow
    for r in ws4.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in r:
            c.number_format = "$#,##0.00"
    for r in ws4.iter_rows(min_row=2, min_col=9, max_col=9):
        for c in r:
            c.number_format = "$#,##0.00"
    autosize(ws4)
    ws4.freeze_panes = "A2"

    # ---- TAB 5: Alternate Parts ----
    ws5 = wb.create_sheet("Alternate Parts")
    ws5.append(["Item #", "Description (ours)", "Manufacturer (ours)", "Annual Qty",
                "Your alternate part", "Your alternate description", "Your price",
                "Form/fit/function confirmed?", "Your notes"])
    for c in ws5[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for ctx in bids_alternates:
        ws5.append([
            ctx["item_num"], ctx["description"], ctx["mfg_name"], ctx["qty_24mo"],
            ctx["your_alt_part"], ctx["your_alt_desc"], ctx["your_price"],
            "", "",
        ])
    for r in ws5.iter_rows(min_row=2, min_col=8, max_col=9):
        for c in r:
            c.fill = yellow
    for r in ws5.iter_rows(min_row=2, min_col=7, max_col=7):
        for c in r:
            c.number_format = "$#,##0.00"
    autosize(ws5)
    ws5.freeze_panes = "A2"

    # ---- TAB 6: No-Bids ----
    ws6 = wb.create_sheet("No-Bids")
    ws6.append(["Item #", "Description", "Manufacturer", "Our UOM", "Annual Qty",
                "Your no-bid note", "Reason category", "Could you bid with more info?", "Your notes"])
    for c in ws6[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for ctx in bids_no_bid:
        ws6.append([
            ctx["item_num"], ctx["description"], ctx["mfg_name"], ctx["our_uom"], ctx["qty_24mo"],
            ctx["your_notes"],
            "", "", "",
        ])
    for r in ws6.iter_rows(min_row=2, min_col=7, max_col=9):
        for c in r:
            c.fill = yellow
    autosize(ws6)
    ws6.freeze_panes = "A2"

    # ---- TAB 7: Full Quote Detail ----
    ws7 = wb.create_sheet("Full Quote Detail")
    ws7.append(["Item #", "Description", "Manufacturer", "Our UOM", "Annual Qty",
                "Status", "Your price", "Your UOM", "Your part #", "Your alt part",
                "Your notes"])
    for c in ws7[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for ctx in bids_all:
        ws7.append([
            ctx["item_num"], ctx["description"], ctx["mfg_name"], ctx["our_uom"], ctx["qty_24mo"],
            ctx["status"], ctx["your_price"], ctx["your_quoted_uom"],
            ctx["your_part"], ctx["your_alt_part"], ctx["your_notes"],
        ])
    for r in ws7.iter_rows(min_row=2, min_col=7, max_col=7):
        for c in r:
            c.number_format = "$#,##0.00"
    autosize(ws7)
    ws7.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Outbound RFQ generator — multi-tab Andersen-branded supplier xlsx.
#
# Per the brief's recommended structure:
#   1. Instructions
#   2. RFQ Lines  (with hidden item_key + rfq_line_id columns for round-trip)
#   3. Terms and Assumptions
#   4. Data Dictionary
#   5. Supplier Response Template
#
# Strict isolation: no historical paid price, no Andersen target / cost /
# margin. Each supplier file contains ONLY this supplier's identification
# and the items being asked. McMaster-style anchor: Part Number column is
# the unique join key for round-trip ingest.
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Visual style palette for supplier-bound xlsx files. Inspired by polished
# supplier-side templates (e.g. Grainger's "Data Spreadsheet" with bright-
# green response columns + bold compact headers).
# ---------------------------------------------------------------------------
COLOR_BRAND_DARK   = "0a0e1a"   # deep navy ground (header bands)
COLOR_BRAND_AMBER  = "ffb733"   # accent
COLOR_BRAND_AMBER_DEEP = "e89c1a"
COLOR_BAND_DARK    = "1c2540"   # secondary band
COLOR_RESPONSE_YELLOW = "FFF59D"  # supplier-response cells (your move)
COLOR_RESPONSE_GREEN  = "C8E6C9"  # supplier-response cells (alt: "fill this in")
COLOR_REFERENCE_GRAY  = "ECEFF1"  # read-only reference cells
COLOR_BORDER       = "999999"
COLOR_INFO_BLUE    = "E3F2FD"
COLOR_WARN_RED     = "FFCDD2"
COLOR_TEXT_LIGHT   = "FFFFFF"
COLOR_TEXT_INK     = "1c2540"
COLOR_INK_MUTED    = "5C6B80"


def _fill(rgb): return PatternFill("solid", fgColor=rgb)
def _border(thin_color=COLOR_BORDER, weight="thin"):
    s = Side(border_style=weight, color=thin_color)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_banner_row(ws, row_idx: int, text: str, span_cols: int,
                      fill_color: str = COLOR_BRAND_DARK,
                      text_color: str = COLOR_BRAND_AMBER,
                      font_size: int = 18,
                      height: int = 38):
    """Write a merged title bar across `span_cols` columns at row `row_idx`."""
    ws.cell(row=row_idx, column=1, value=text)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span_cols)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = Font(bold=True, color=text_color, size=font_size)
    cell.fill = _fill(fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = height


def _write_subbanner(ws, row_idx: int, text: str, span_cols: int,
                     fill_color: str = COLOR_BAND_DARK,
                     text_color: str = COLOR_TEXT_LIGHT,
                     font_size: int = 12, height: int = 26):
    ws.cell(row=row_idx, column=1, value=text)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=span_cols)
    cell = ws.cell(row=row_idx, column=1)
    cell.font = Font(bold=True, color=text_color, size=font_size)
    cell.fill = _fill(fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row_idx].height = height


def _style_table_header(ws, row_idx: int, n_cols: int,
                        fill_color: str = COLOR_BRAND_DARK,
                        text_color: str = COLOR_TEXT_LIGHT,
                        font_size: int = 10):
    border = _border(weight="thin")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = Font(bold=True, color=text_color, size=font_size)
        cell.fill = _fill(fill_color)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row_idx].height = 32


def _alternate_row_stripes(ws, start_row: int, end_row: int, n_cols: int,
                           stripe_color: str = "F5F7FA"):
    fill = _fill(stripe_color)
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if not (cell.fill and cell.fill.patternType and cell.fill.fgColor.rgb not in (None, "00000000")):
                    cell.fill = fill


def gen_outbound_rfq_xlsx(supplier_name: str, rfq_id: str = "",
                          response_due_date: str = "",
                          contact_name: str = "",
                          contact_email: str = "",
                          included_keys=None) -> bytes:
    # Pull from user profile if not explicitly passed
    profile = get_user_profile()
    if not contact_name:
        contact_name = profile.get("name") or "(operator name not set)"
    if not contact_email:
        contact_email = profile.get("email") or "(operator email not set)"
    contact_company = profile.get("company") or "Andersen"
    """Generate one supplier's outbound RFQ workbook.

    Strict isolation rules baked in:
      - No historical paid price column
      - No internal target / cost / margin
      - Hidden item_key and rfq_line_id columns for round-trip matching
      - Locked sheet, only response cells editable
      - Dropdown validation on Yes/No / UOM / No Bid Reason
    """
    if not supplier_name:
        raise ValueError("supplier_name required")

    items = _STATE.get("items", [])
    if included_keys is not None:
        keep = set(str(k) for k in included_keys)
        items = [it for it in items if it["item_num"] in keep]
    else:
        items = [it for it in items if it.get("included")]

    if not rfq_id:
        rfq_id = f"RFQ-{datetime.now().strftime('%Y-%m')}-001"
    if not response_due_date:
        # Default: 14 days from now
        from datetime import timedelta as _td
        response_due_date = (datetime.now() + _td(days=14)).strftime("%Y-%m-%d")

    wb = Workbook()

    # ---- TAB 1: Instructions ----
    ws = wb.active
    ws.title = "Instructions"
    _write_banner_row(ws, 1, f"REQUEST FOR QUOTATION  ·  {rfq_id}", span_cols=4, font_size=20, height=46)
    _write_subbanner(ws, 2, f"Issued to: {supplier_name}", span_cols=4, font_size=14, height=30)
    # Meta row
    meta = [
        ("Issue date",        datetime.now().strftime('%Y-%m-%d')),
        ("Response due",      response_due_date),
        (f"{contact_company} contact",  f"{contact_name} — {contact_email}"),
        ("RFQ identifier",    rfq_id),
    ]
    ws.append([])
    for label, val in meta:
        ws.append([label, val])
        cell_label = ws.cell(row=ws.max_row, column=1)
        cell_label.font = Font(bold=True, color=COLOR_TEXT_INK, size=11)
        cell_label.fill = _fill(COLOR_REFERENCE_GRAY)
        cell_label.alignment = Alignment(vertical="center", indent=1)
        cell_val = ws.cell(row=ws.max_row, column=2)
        cell_val.font = Font(size=11)
        cell_val.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[ws.max_row].height = 22

    ws.append([])
    _write_subbanner(ws, ws.max_row + 1, "INSTRUCTIONS  ·  HOW TO COMPLETE THIS RFQ", span_cols=4)
    ws.append([])
    instructions = [
        "1. Open the 'Supplier Response Template' tab. Yellow cells are for your responses; gray reference cells are read-only.",
        "2. Do not change the order of rows or columns. Hidden columns (item_key, rfq_line_id) are used for round-trip matching when you return the file.",
        "3. Quote each item in the UOM specified in the gray reference column. If you must quote in a different UOM, set 'Quote UOM' accordingly and add a note in 'Supplier Notes'.",
        "4. If you cannot bid an item, mark 'No Bid' = Yes and select a reason from the dropdown. If you have a comment, add it in 'Supplier Notes'.",
        "5. If you offer an alternate part, set 'Alternate Part Offered' = Yes and complete the alternate part columns.",
        "6. Lead time should be in calendar days from PO receipt.",
        "7. Quote validity: please indicate how long your prices are firm in the 'Valid Through Date' column.",
        "8. All prices should be NET (excluding freight unless 'Freight Included' = Yes).",
        "9. Please return the completed workbook to the Andersen contact above by the response-due date.",
        "10. Any item without a response will be treated as a no-bid.",
    ]
    for line in instructions:
        ws.append([line])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
        ws.row_dimensions[ws.max_row].height = 32

    ws.append([])
    _write_subbanner(ws, ws.max_row + 1, "WORKBOOK CONTENTS", span_cols=4)
    ws.append([])
    for tab_line in [
        "  •  Instructions  —  this tab",
        "  •  RFQ Lines  —  read-only summary of items being requested",
        "  •  Terms and Assumptions  —  ground rules for this RFQ",
        "  •  Data Dictionary  —  definition of each column in the response template",
        "  •  Supplier Response Template  —  YOUR INPUT GOES HERE",
    ]:
        ws.append([tab_line])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(size=11)
        cell.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 1
    ws.column_dimensions["D"].width = 1

    # ---- TAB 2: RFQ Lines (read-only summary) ----
    ws2 = wb.create_sheet("RFQ Lines")
    headers = ["Line #", "Andersen Item #", "EAM Part #", "Manufacturer Part #",
               "Manufacturer", "Description", "Commodity", "Annual Qty", "UOM"]
    _write_banner_row(ws2, 1, f"RFQ LINES  ·  {len(items):,} items requested", span_cols=len(headers))
    _write_subbanner(ws2, 2, "Read-only summary — for your reference. Quote in the 'Supplier Response Template' tab.", span_cols=len(headers), font_size=10, height=22)
    ws2.append([])
    ws2.append(headers)
    _style_table_header(ws2, ws2.max_row, len(headers))
    sorted_items = sorted(items, key=lambda x: (x.get("qty_24mo") or 0) * (x.get("last_unit_price") or 0), reverse=True)
    data_start = ws2.max_row + 1
    for i, it in enumerate(sorted_items, start=1):
        ws2.append([
            i,
            it["item_num"],
            it.get("eam_pn") or "",
            it.get("mfg_pn") or "",
            it.get("mfg_name") or "",
            it.get("description") or "",
            it.get("commodity") or "",
            it.get("qty_24mo") or 0,
            it.get("uom") or "",
        ])
    _alternate_row_stripes(ws2, data_start, ws2.max_row, len(headers))
    # Borders + alignment on data rows
    border = _border()
    for r in range(data_start, ws2.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    autosize(ws2)
    ws2.freeze_panes = f"A{data_start}"

    # ---- TAB 3: Terms and Assumptions ----
    ws3 = wb.create_sheet("Terms and Assumptions")
    _write_banner_row(ws3, 1, "TERMS AND ASSUMPTIONS", span_cols=2)
    ws3.append([])
    terms = [
        ("Quote currency", "USD unless otherwise stated."),
        ("Quote basis", "Net unit prices, excluding freight unless 'Freight Included' = Yes."),
        ("Quantity basis", "Annual quantities are estimates based on 24-month order history. Actual orders may vary."),
        ("Award terms", "This RFQ is not a commitment to purchase. Awards will be issued by separate purchase order."),
        ("Substitutes", "Alternate parts will be evaluated for form/fit/function equivalence. Quote the OEM part as primary; alternates are secondary."),
        ("Pricing audit", "Andersen reserves the right to request manufacturer cost documentation for items priced significantly above market."),
        ("Confidentiality", "Pricing in this RFQ and your responses are confidential between Andersen and your company."),
        ("Validity", "Quotes should remain firm for at least 90 days unless otherwise noted in the 'Valid Through' field."),
        ("Tariffs", "If tariffs apply, indicate whether they're included in your quoted price ('Tariff Included' field)."),
        ("Round-trip data", "Hidden columns in the response template (item_key, rfq_line_id) are used for matching your response back to our items. Do not delete or modify these columns."),
    ]
    ws3.append(["Topic", "Detail"])
    _style_table_header(ws3, ws3.max_row, 2)
    border = _border()
    data_start_3 = ws3.max_row + 1
    for topic, detail in terms:
        ws3.append([topic, detail])
        ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=11, color=COLOR_TEXT_INK)
        ws3.cell(row=ws3.max_row, column=1).fill = _fill(COLOR_REFERENCE_GRAY)
    _alternate_row_stripes(ws3, data_start_3, ws3.max_row, 2)
    for r in range(data_start_3, ws3.max_row + 1):
        for c in range(1, 3):
            cell = ws3.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        ws3.row_dimensions[r].height = 32
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 110

    # ---- TAB 4: Data Dictionary ----
    ws4 = wb.create_sheet("Data Dictionary")
    _write_banner_row(ws4, 1, "DATA DICTIONARY  ·  every column defined", span_cols=3)
    ws4.append([])
    ws4.append(["Column", "Required?", "Definition"])
    _style_table_header(ws4, ws4.max_row, 3)
    data_start_4 = ws4.max_row + 1
    dictionary = [
        ("Andersen Item #", "Reference", "Our internal item identifier."),
        ("EAM Part #", "Reference", "Our EAM/maintenance system part number (primary key for most suppliers)."),
        ("Manufacturer Part #", "Reference", "OEM manufacturer part number, when known."),
        ("Annual Qty", "Reference", "Estimated annual consumption based on 24-month history."),
        ("UOM", "Reference", "Our purchasing unit of measure."),
        ("Quote Price", "Required", "Your net unit price in the Quote UOM. To-the-penny precision."),
        ("Quote UOM", "Required", "The unit of measure your price is quoted in (EA, BX, PK, etc.). Use the dropdown."),
        ("Your Part #", "Required", "Your catalog SKU for this item."),
        ("Minimum Order Quantity", "Required if applicable", "Minimum order qty in your Quote UOM. Leave blank if no minimum."),
        ("Lead Time Days", "Required", "Calendar days from PO receipt to ship."),
        ("Country of Origin", "Recommended", "Two-letter ISO code (US, CN, MX, etc.)."),
        ("Manufacturer", "Recommended", "OEM manufacturer name."),
        ("Manufacturer Part Number", "Recommended", "OEM part number you're quoting."),
        ("Alternate Part Offered", "Optional", "Yes/No. Use the dropdown."),
        ("Alternate Item Number", "If alternate", "Your alternate part number, if you're proposing a substitute."),
        ("Alternate Description", "If alternate", "Description of the alternate part."),
        ("Tariff Included", "Required", "Yes/No. Whether your price includes any applicable tariffs."),
        ("Freight Included", "Required", "Yes/No. Whether your price includes freight to our facility."),
        ("No Bid", "Optional", "Yes/No. Mark Yes if you cannot bid on this item; select a reason."),
        ("No Bid Reason", "If no-bid", "Reason from the dropdown (Discontinued, Not Available, Need More Info, etc.)."),
        ("Supplier Notes", "Optional", "Any free-text comments — alternate proposals, UOM concerns, etc."),
        ("Valid Through Date", "Required", "Date your quote remains firm. YYYY-MM-DD format preferred."),
        ("item_key", "DO NOT EDIT", "Hidden internal matching key. Do not modify or delete."),
        ("rfq_line_id", "DO NOT EDIT", "Hidden internal line identifier. Do not modify or delete."),
    ]
    for col, req, defn in dictionary:
        ws4.append([col, req, defn])
    border = _border()
    _alternate_row_stripes(ws4, data_start_4, ws4.max_row, 3)
    for r in range(data_start_4, ws4.max_row + 1):
        for c in range(1, 4):
            cell = ws4.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        # Bold the column-name col
        ws4.cell(row=r, column=1).font = Font(bold=True, color=COLOR_TEXT_INK, size=11)
        # Color-code the Required? column
        req_val = (ws4.cell(row=r, column=2).value or "").lower()
        if "required" in req_val:
            ws4.cell(row=r, column=2).font = Font(bold=True, color="C62828", size=10)
        elif "do not edit" in req_val:
            ws4.cell(row=r, column=2).font = Font(bold=True, color="6A1B9A", size=10)
        elif "recommended" in req_val:
            ws4.cell(row=r, column=2).font = Font(bold=True, color="EF6C00", size=10)
        else:
            ws4.cell(row=r, column=2).font = Font(color=COLOR_INK_MUTED, size=10)
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 22
    ws4.column_dimensions["C"].width = 90

    # ---- TAB 5: Supplier Response Template ----
    ws5 = wb.create_sheet("Supplier Response Template")
    n_template_cols = 26
    _write_banner_row(ws5, 1, f"SUPPLIER RESPONSE TEMPLATE  ·  {supplier_name}  ·  {rfq_id}", span_cols=n_template_cols, font_size=18, height=42)
    _write_subbanner(ws5, 2,
                     "  YELLOW cells = your input  ·  GRAY cells = read-only reference (do not modify)  ·  Hidden cols Y/Z = round-trip keys (do not delete)",
                     span_cols=n_template_cols, font_size=10, height=22,
                     fill_color=COLOR_BRAND_AMBER, text_color=COLOR_TEXT_INK)
    ws5.append([])
    template_headers = [
        # Reference (read-only) cols
        "Andersen Item #",   # A
        "EAM Part #",        # B
        "Manufacturer Part #",  # C
        "Manufacturer",      # D
        "Description",       # E
        "Annual Qty",        # F
        "UOM",               # G
        # Response cols (yellow)
        "Quote Price",       # H *
        "Quote UOM",         # I *
        "Your Part #",       # J *
        "Minimum Order Quantity",  # K
        "Lead Time Days",    # L *
        "Country of Origin", # M
        "Manufacturer (your)",  # N
        "Manufacturer Part Number (your)",  # O
        "Alternate Part Offered",  # P
        "Alternate Item Number",   # Q
        "Alternate Description",   # R
        "Tariff Included",   # S *
        "Freight Included",  # T *
        "No Bid",            # U
        "No Bid Reason",     # V
        "Supplier Notes",    # W
        "Valid Through Date",  # X *
        # Hidden round-trip columns
        "item_key",          # Y (hidden)
        "rfq_line_id",       # Z (hidden)
    ]
    ws5.append(template_headers)
    # Style the header row — different fill for reference (gray) vs response (amber)
    HDR_FILL_REF = _fill(COLOR_BRAND_DARK)
    HDR_FILL_RESP = _fill(COLOR_BRAND_AMBER)
    HDR_FILL_HIDDEN = _fill("CCCCCC")
    border = _border()
    for c_idx in range(1, n_template_cols + 1):
        cell = ws5.cell(row=4, column=c_idx)
        cell.font = Font(bold=True, color=COLOR_TEXT_LIGHT if c_idx <= 7 else COLOR_TEXT_INK, size=10)
        if c_idx <= 7:           # Reference cols (A-G)
            cell.fill = HDR_FILL_REF
        elif c_idx <= 24:        # Response cols (H-X)
            cell.fill = HDR_FILL_RESP
        else:                    # Hidden round-trip cols (Y-Z)
            cell.fill = HDR_FILL_HIDDEN
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        cell.border = border
    ws5.row_dimensions[4].height = 38

    # Row data
    for i, it in enumerate(sorted_items, start=1):
        rfq_line_id = f"{rfq_id}-{i:05d}"
        ws5.append([
            it["item_num"],            # A
            it.get("eam_pn") or "",    # B
            it.get("mfg_pn") or "",    # C
            it.get("mfg_name") or "",  # D
            it.get("description") or "",   # E
            it.get("qty_24mo") or 0,   # F
            it.get("uom") or "",       # G
            None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,  # H..X (yellow)
            it["key"],                 # Y item_key
            rfq_line_id,               # Z rfq_line_id
        ])

    # Color-code response cells (yellow) vs reference cells (gray) for all data rows
    n_data_rows = len(sorted_items)
    YELLOW_FILL = _fill(COLOR_RESPONSE_YELLOW)
    REF_FILL = _fill(COLOR_REFERENCE_GRAY)
    for r_idx in range(5, 5 + n_data_rows):
        for col_idx in range(8, 25):  # H through X inclusive — supplier responses
            c = ws5.cell(row=r_idx, column=col_idx)
            c.fill = YELLOW_FILL
            c.border = border
        for col_idx in range(1, 8):  # A..G — read-only reference
            c = ws5.cell(row=r_idx, column=col_idx)
            c.fill = REF_FILL
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)

    # Number formats
    for r in ws5.iter_rows(min_row=5, min_col=6, max_col=6):  # F = Annual Qty
        for c in r:
            c.number_format = "#,##0"
    for r in ws5.iter_rows(min_row=5, min_col=8, max_col=8):  # H = Quote Price
        for c in r:
            c.number_format = "$#,##0.00"

    # Hide the item_key + rfq_line_id columns (Y, Z = 25, 26)
    ws5.column_dimensions[get_column_letter(25)].hidden = True
    ws5.column_dimensions[get_column_letter(26)].hidden = True

    # Dropdown validation on key fields
    try:
        from openpyxl.worksheet.datavalidation import DataValidation
        # Yes/No on cols P, S, T, U
        yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws5.add_data_validation(yn)
        for col_letter in ("P", "S", "T", "U"):
            yn.add(f"{col_letter}5:{col_letter}{4 + n_data_rows}")
        # No Bid Reason dropdown on V
        nobid = DataValidation(
            type="list",
            formula1='"Discontinued,Not Available,Not Stocked,Need More Information,Outside Capability,Pricing Not Competitive,Obsolete,Other"',
            allow_blank=True,
        )
        ws5.add_data_validation(nobid)
        nobid.add(f"V5:V{4 + n_data_rows}")
        # UOM dropdown on I
        uom = DataValidation(
            type="list",
            formula1='"EA,BX,PK,FT,IN,LB,KG,GAL,QT,PT,OZ,RL,CS,DZ,M,CM,MM,SET,KIT,Other"',
            allow_blank=True,
        )
        ws5.add_data_validation(uom)
        uom.add(f"I5:I{4 + n_data_rows}")
    except Exception as e:
        # Validation is nice-to-have; if openpyxl barfs, the file still works
        pass

    # Sheet protection — lock everything except yellow response cells
    try:
        # Mark response cells as unlocked first
        from openpyxl.styles import Protection
        for r_idx in range(5, 5 + n_data_rows):
            for col_idx in range(8, 25):  # H..X = response cols
                c = ws5.cell(row=r_idx, column=col_idx)
                c.protection = Protection(locked=False)
        ws5.protection.sheet = True
        ws5.protection.password = "rfq"  # weak by design — deters accidents, not malice
        ws5.protection.formatCells = False
        ws5.protection.formatColumns = False
        ws5.protection.formatRows = False
    except Exception:
        pass

    autosize(ws5, max_w=40)
    ws5.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def list_loaded_supplier_names() -> list:
    """Convenience for the JS to know which suppliers have loaded bids."""
    return list((_STATE.get("bids", {}) or {}).keys())


# ---------------------------------------------------------------------------
# State serialization for save/restore — pulls everything Python-side that
# needs to survive a session save/reload.
#
# NOT serialized:
#   - items / po_lines_by_key / annual_spend / kpis / difficulty
#     (these are derived from extract_rfq_list and re-computed on restore
#     when the user re-loads the source xlsx)
#
# Serialized:
#   - bids (parsed supplier responses — re-loading them is otherwise tedious)
#   - scenarios
#   - thresholds (overrides)
#   - difficulty_history (the period-end snapshot series)
#   - supplier_name (incumbent)
#   - data_anchor_date
# ---------------------------------------------------------------------------

def serialize_state() -> dict:
    """Return a JSON-safe dict capturing everything in _STATE that should
    persist across a save/reload cycle."""
    return {
        "bids":                 _STATE.get("bids", {}),
        "scenarios":            _STATE.get("scenarios", {}),
        "thresholds":           _STATE.get("thresholds", {}),
        "difficulty_history":   _STATE.get("difficulty_history", []),
        "supplier_name":        _STATE.get("supplier_name", ""),
        "data_anchor_date":     _STATE.get("data_anchor_date"),
        "audit_log":            _STATE.get("audit_log", []),
        "user_profile":         _STATE.get("user_profile", {}),
        # Manual UOM resolution annotations — keyed by "<item_key>|<supplier>".
        # Survives save/load so the analyst's catalog-lookup work isn't lost,
        # AND rides through to a colleague's app when the JSON save is shared.
        "uom_annotations":      _STATE.get("uom_annotations", {}),
        # Per-item outlier-line exclusions from the per-item history modal.
        # Maps item_num → [excluded_indices into ascending-sorted po_lines].
        # Survives save/load so the analyst's outlier curation persists across
        # reopens of the RFQ event.
        "item_exclusions":      _STATE.get("item_exclusions", {}),
        # Per-item supplier locks — analyst-confirmed pins of an item's
        # award to a specific supplier. Apply across every scenario.
        "item_locks":           _STATE.get("item_locks", {}),
        # Excluded-line review log — every per-item modal exclusion +
        # un-exclusion event with the line snapshot + before-baseline.
        # Survives save/load AND a JSON colleague-share so the master
        # audit record stays intact across sessions.
        "exclusion_log":        _STATE.get("exclusion_log", []),
        # Round 2/Rn focused-RFQ selection — the analyst's picks from the
        # comparison matrix for the next negotiation round.
        "round2_selection":     _STATE.get("round2_selection", []),
        "current_round":        _STATE.get("current_round", 1),
    }


# ---------------------------------------------------------------------------
# Audit log — discrete action-level event trail.
#
# Per the brief: every important action gets a row. Useful for:
#   - "What did I do last Tuesday on the McMaster RFQ?" (period-end recap)
#   - Director conversations: "show me the trail for this award decision"
#   - Debugging: "why is the difficulty score 64 — when did it change?"
# ---------------------------------------------------------------------------

AUDIT_MAX_ENTRIES = 500


def log_event(action_type: str, action_detail: str = "", related: str = "") -> None:
    """Append an audit-log entry. action_type is short (verb + noun).
    related is a free-form id (rfq_id, supplier_name, scenario_name, etc.)."""
    if not action_type:
        return
    log = _STATE.setdefault("audit_log", [])
    log.append({
        "timestamp": datetime.now().isoformat(),
        "action_type": action_type,
        "action_detail": action_detail,
        "related": related,
    })
    if len(log) > AUDIT_MAX_ENTRIES:
        del log[:len(log) - AUDIT_MAX_ENTRIES]


def detect_item_conflicts(items: list) -> dict:
    """Surface data-hygiene issues across the full item list.

    Returns:
      {
        "summary": {n_conflicts_total, n_items_affected, ...},
        "by_type": {
          "MFG_PN_MULTI_ITEM": [...],   # same MFG PN under multiple item_num
          "ITEM_NUM_MULTI_DESC": [...], # same item_num with different descriptions
          "DESC_MULTI_ITEM": [...],     # same description under multiple item_nums
          "MFG_PN_MULTI_MFR": [...],    # same MFG PN under multiple manufacturers
        }
      }

    Useful for trimming the candidate RFQ list of items that need cleanup
    BEFORE going out to suppliers. Per the brief: don't silently auto-resolve.
    """
    by_mfg_pn = defaultdict(list)        # mfg_pn → [item dicts]
    by_item_num = defaultdict(list)      # item_num → [item dicts]
    by_desc = defaultdict(list)          # desc_norm → [item dicts]
    by_mfg_pn_for_mfr_check = defaultdict(set)  # mfg_pn → set of mfg_names

    for it in items:
        mfg_pn = (it.get("mfg_pn") or "").strip()
        item_num = (it.get("item_num") or "").strip()
        desc = (it.get("description") or "").strip()
        mfg = (it.get("mfg_name") or "").strip()

        if mfg_pn:
            mp_key = norm_pn(mfg_pn)
            by_mfg_pn[mp_key].append(it)
            if mfg:
                by_mfg_pn_for_mfr_check[mp_key].add(mfg)
        if item_num:
            by_item_num[item_num].append(it)
        if desc and len(desc) >= 10:
            d_norm = re.sub(r"\s+", " ", desc.lower())[:120]
            by_desc[d_norm].append(it)

    conflicts = {
        "MFG_PN_MULTI_ITEM": [],
        "ITEM_NUM_MULTI_DESC": [],
        "DESC_MULTI_ITEM": [],
        "MFG_PN_MULTI_MFR": [],
    }

    # Same MFG PN with different item_num
    for mp_key, group in by_mfg_pn.items():
        if not mp_key:
            continue
        distinct_items = set(it["item_num"] for it in group if it.get("item_num"))
        if len(distinct_items) >= 2:
            sample_mfg_pn = group[0].get("mfg_pn") or mp_key
            conflicts["MFG_PN_MULTI_ITEM"].append({
                "mfg_pn": sample_mfg_pn,
                "item_nums": sorted(distinct_items),
                "n_items": len(distinct_items),
                "n_lines": len(group),
            })

    # Same item_num with different descriptions
    for item_num, group in by_item_num.items():
        # Already deduped by extract_rfq_list, so each item_num only has ONE row.
        # We get conflicts here when items_by_item_num[k] has length > 1, which
        # only happens if dedup-key fell through to part_number (i.e. one
        # item_num appears in multiple part_number-keyed records). Skip
        # this block for now — extract_rfq_list de-duplicates at item_key
        # level, not at item_num level.
        pass

    # Same description under different item_nums
    for d_norm, group in by_desc.items():
        distinct_items = set(it["item_num"] for it in group if it.get("item_num"))
        if len(distinct_items) >= 2:
            sample_desc = group[0].get("description") or d_norm
            conflicts["DESC_MULTI_ITEM"].append({
                "description": sample_desc[:80],
                "item_nums": sorted(distinct_items),
                "n_items": len(distinct_items),
            })

    # Same MFG PN under multiple manufacturers
    for mp_key, mfrs in by_mfg_pn_for_mfr_check.items():
        if len(mfrs) >= 2:
            sample_group = by_mfg_pn[mp_key]
            sample_mfg_pn = sample_group[0].get("mfg_pn") or mp_key
            conflicts["MFG_PN_MULTI_MFR"].append({
                "mfg_pn": sample_mfg_pn,
                "manufacturers": sorted(mfrs),
                "n_items": len(sample_group),
            })

    # Mark conflict-affected items so the UI can flag them
    affected_keys = set()
    for c in conflicts["MFG_PN_MULTI_ITEM"]:
        affected_keys.update(c["item_nums"])
    for c in conflicts["DESC_MULTI_ITEM"]:
        affected_keys.update(c["item_nums"])
    for c in conflicts["MFG_PN_MULTI_MFR"]:
        # Find item_nums for that mfg_pn
        for it in items:
            if it.get("mfg_pn") and norm_pn(it["mfg_pn"]) == norm_pn(c["mfg_pn"]):
                affected_keys.add(it["item_num"])

    summary = {
        "n_conflicts_total": sum(len(v) for v in conflicts.values()),
        "n_items_affected": len(affected_keys),
        "n_mfg_pn_multi_item": len(conflicts["MFG_PN_MULTI_ITEM"]),
        "n_desc_multi_item": len(conflicts["DESC_MULTI_ITEM"]),
        "n_mfg_pn_multi_mfr": len(conflicts["MFG_PN_MULTI_MFR"]),
    }
    return {
        "summary": summary,
        "by_type": conflicts,
        "affected_item_nums": sorted(affected_keys),
    }


def list_audit_log(limit: int = None) -> list:
    log = list(_STATE.get("audit_log", []))
    log.reverse()  # most recent first
    return log[:limit] if limit else log


def clear_audit_log() -> None:
    _STATE["audit_log"] = []


def restore_state(payload: dict) -> None:
    """Apply a previously-serialized state payload back into _STATE.
    Items + po_lines + difficulty are NOT restored from the save — they
    require re-running extract_rfq_list against the source xlsx."""
    if not isinstance(payload, dict):
        return
    if "bids" in payload and isinstance(payload["bids"], dict):
        _STATE["bids"] = payload["bids"]
    if "scenarios" in payload and isinstance(payload["scenarios"], dict):
        _STATE["scenarios"] = payload["scenarios"]
    if "thresholds" in payload and isinstance(payload["thresholds"], dict):
        _STATE["thresholds"] = payload["thresholds"]
    if "difficulty_history" in payload and isinstance(payload["difficulty_history"], list):
        _STATE["difficulty_history"] = payload["difficulty_history"]
    if "supplier_name" in payload:
        _STATE["supplier_name"] = payload["supplier_name"]
    if "data_anchor_date" in payload:
        _STATE["data_anchor_date"] = payload["data_anchor_date"]
    if "audit_log" in payload and isinstance(payload["audit_log"], list):
        _STATE["audit_log"] = payload["audit_log"]
    if "user_profile" in payload and isinstance(payload["user_profile"], dict):
        _STATE["user_profile"] = payload["user_profile"]
    if "uom_annotations" in payload and isinstance(payload["uom_annotations"], dict):
        _STATE["uom_annotations"] = payload["uom_annotations"]
    if "item_exclusions" in payload and isinstance(payload["item_exclusions"], dict):
        # Coerce values to int lists — JSON round-trip preserves them but
        # belt-and-suspenders against hand-edited save files.
        cleaned = {}
        for item_num, idxs in payload["item_exclusions"].items():
            if not isinstance(idxs, (list, tuple)):
                continue
            ints = []
            for v in idxs:
                try:
                    ints.append(int(v))
                except (TypeError, ValueError):
                    continue
            if ints:
                cleaned[str(item_num)] = sorted(set(ints))
        _STATE["item_exclusions"] = cleaned
    if "item_locks" in payload and isinstance(payload["item_locks"], dict):
        cleaned_locks = {}
        for item_num, rec in payload["item_locks"].items():
            if not isinstance(rec, dict):
                continue
            sup = rec.get("supplier")
            if not sup:
                continue
            cleaned_locks[str(item_num)] = {
                "supplier": str(sup),
                "reason": rec.get("reason") or "",
                "locked_at": rec.get("locked_at") or "",
            }
        _STATE["item_locks"] = cleaned_locks
    if "exclusion_log" in payload and isinstance(payload["exclusion_log"], list):
        # Pass-through with light validation — entries are authored by this
        # module so we trust shape, but cap to a defensive max so a hand-
        # edited save can't blow memory.
        _STATE["exclusion_log"] = list(payload["exclusion_log"])[:10000]
    if "round2_selection" in payload and isinstance(payload["round2_selection"], list):
        _STATE["round2_selection"] = sorted({str(n) for n in payload["round2_selection"] if n})
    if "current_round" in payload:
        try:
            _STATE["current_round"] = max(1, int(payload["current_round"]))
        except (TypeError, ValueError):
            pass


# ---------------------------------------------------------------------------
# Award scenarios — named, saveable, side-by-side comparison.
#
# A "scenario" is a strategy for awarding the RFQ:
#   - lowest_price       : per item, pick the lowest non-flagged bid
#   - lowest_qualified   : like lowest_price but excludes UOM_DISC + SUBSTITUTE
#   - incumbent_preferred: stay with the historical supplier when their bid
#                          is within `min_savings_pct_to_switch` of lowest
#   - consolidate_to     : award everything to one named supplier (with carves
#                          where another supplier's price beats them by
#                          carve_out_min_savings_pct)
#   - manual             : user-defined per-item award
#
# Scenarios are stored on _STATE so they persist with the session save state.
# Each scenario carries its strategy + parameters + manual overrides + a
# computed totals snapshot (so re-loading is fast even before re-evaluation).
# ---------------------------------------------------------------------------

SCENARIO_STRATEGIES = (
    "lowest_price",
    "lowest_qualified",
    "incumbent_preferred",
    "consolidate_to",
    "manual",
)


def _get_scenarios() -> dict:
    return _STATE.setdefault("scenarios", {})


def list_award_scenarios() -> list:
    """Return all scenarios as a list (most-recent-saved first)."""
    sc = _get_scenarios()
    out = []
    for name, s in sc.items():
        out.append({
            "name": name,
            "strategy": s.get("strategy"),
            "parameters": s.get("parameters", {}),
            "saved_at": s.get("saved_at"),
            "totals": s.get("totals", {}),
            "n_overrides": len(s.get("overrides") or {}),
        })
    out.sort(key=lambda x: x.get("saved_at") or "", reverse=True)
    return out


def delete_award_scenario(name: str) -> bool:
    sc = _get_scenarios()
    if name in sc:
        del sc[name]
        log_event("delete_award_scenario", "scenario removed", related=name)
        return True
    return False


def _evaluate_scenario(strategy: str, parameters: dict, overrides: dict, included_keys=None) -> dict:
    """Run an award strategy against current items + bids, returning the per-item
    award decisions + roll-up totals.

    overrides: {rfq_key: {"supplier": supplier_name | None, "reason": str}}
               Per-scenario manual awards. Highest priority — beats both
               item locks and strategy logic.

    Cross-scenario item locks (``_STATE["item_locks"]``) are applied AFTER
    overrides but BEFORE strategy. A lock on item_num=X with supplier=S
    forces award to S for every scenario, as long as S has a priced bid
    for the item. If S did not bid (or bid NO_BID/NEED_INFO), the lock is
    impotent for THIS evaluation (recorded in decision_basis as a warning)
    and strategy logic falls through.
    """
    th = get_thresholds()
    items = _STATE.get("items", [])
    if included_keys is not None:
        keep = set(included_keys)
        items = [it for it in items if it["item_num"] in keep]

    bids_by_supplier = _STATE.get("bids", {}) or {}
    suppliers = list(bids_by_supplier.keys())
    bid_lookup = {}
    for sup, parsed in bids_by_supplier.items():
        for b in parsed.get("bids", []):
            if b["status"] in (BID_STATUS_PRICED, BID_STATUS_UOM_DISC, BID_STATUS_SUBSTITUTE) \
               and b["effective_price"] is not None and b["effective_price"] > 0:
                bid_lookup[(sup, b["rfq_key"])] = b

    item_locks = _STATE.get("item_locks", {}) or {}

    overrides = overrides or {}
    parameters = parameters or {}
    consolidate_supplier = parameters.get("supplier") if strategy == "consolidate_to" else None
    incumbent_threshold = parameters.get("incumbent_keep_threshold_pct", th["min_savings_pct_to_switch"])
    carve_threshold = parameters.get("carve_threshold", th["carve_out_min_savings_pct"])
    carve_threshold_dollar = parameters.get("carve_threshold_dollar", th["carve_out_min_savings_annual_dollar"])
    exclude_uom = parameters.get("exclude_uom_disc", strategy == "lowest_qualified")
    exclude_subs = parameters.get("exclude_substitutes", strategy == "lowest_qualified")

    awards = []
    n_no_award = 0
    n_manual = 0
    n_carved = 0
    n_locked = 0          # locks that successfully forced an award this scenario
    n_locks_unhonored = 0 # locks where the locked supplier didn't bid this item
    award_total = 0.0
    historical_total = 0.0
    items_switched = 0
    incumbent_retained = 0
    by_supplier = {}

    for it in items:
        rfq_key = it["key"]
        qty = it.get("qty_24mo") or 0
        hist_price = it.get("last_unit_price") or 0
        # Note: incumbent supplier from historical data; may be the same
        # supplier name as one of the bidders, OR a different name (e.g.
        # "MCMASTER-CARR") that doesn't match any bidder
        incumbent = (it.get("supplier") if "supplier" in it else None) or _STATE.get("supplier_name", "")

        # Manual override beats everything
        if rfq_key in overrides:
            ov = overrides[rfq_key]
            sup = ov.get("supplier")
            if sup:
                bid = bid_lookup.get((sup, rfq_key))
                price = bid["effective_price"] if bid else None
                awards.append({
                    "rfq_key": rfq_key,
                    "item_num": it["item_num"],
                    "description": it["description"],
                    "qty_24mo": qty,
                    "awarded_supplier": sup,
                    "awarded_price": price,
                    "awarded_value": (price or 0) * qty,
                    "historical_price": hist_price,
                    "historical_value": hist_price * qty,
                    "savings_value": (hist_price - (price or 0)) * qty if hist_price and price else 0,
                    "decision_basis": f"MANUAL: {ov.get('reason') or 'user override'}",
                })
                n_manual += 1
                if price is not None and price > 0:
                    award_total += price * qty
                historical_total += hist_price * qty
                by_supplier[sup] = by_supplier.get(sup, 0) + (price or 0) * qty
                continue
            else:
                # Explicit "no award" override
                awards.append({
                    "rfq_key": rfq_key, "item_num": it["item_num"], "description": it["description"],
                    "qty_24mo": qty, "awarded_supplier": None, "awarded_price": None,
                    "awarded_value": 0, "historical_price": hist_price,
                    "historical_value": hist_price * qty, "savings_value": 0,
                    "decision_basis": f"MANUAL: {ov.get('reason') or 'no award'}",
                })
                n_no_award += 1
                historical_total += hist_price * qty
                continue

        # Item lock — analyst-confirmed pin to a specific supplier. Applies
        # across every scenario. Beats strategy logic; loses to per-scenario
        # explicit overrides above. If the locked supplier didn't bid this
        # item, the lock is recorded as an unhonored warning and strategy
        # logic continues normally below.
        lock_record = item_locks.get(it["item_num"])
        lock_warning = None
        if lock_record:
            locked_sup = lock_record.get("supplier")
            locked_bid = bid_lookup.get((locked_sup, rfq_key)) if locked_sup else None
            if locked_bid is not None:
                price = locked_bid["effective_price"]
                reason_txt = lock_record.get("reason") or "audited bid"
                if incumbent and locked_sup.lower() == (incumbent or "").lower():
                    incumbent_retained += 1
                else:
                    items_switched += 1
                award_total += price * qty
                historical_total += hist_price * qty
                by_supplier[locked_sup] = by_supplier.get(locked_sup, 0) + price * qty
                awards.append({
                    "rfq_key": rfq_key, "item_num": it["item_num"], "description": it["description"],
                    "qty_24mo": qty, "awarded_supplier": locked_sup, "awarded_price": price,
                    "awarded_value": price * qty, "historical_price": hist_price,
                    "historical_value": hist_price * qty,
                    "savings_value": (hist_price - price) * qty if hist_price else 0,
                    "decision_basis": f"LOCKED → {locked_sup} ({reason_txt})",
                })
                n_locked += 1
                continue
            else:
                # Lock present but locked supplier didn't bid — surface as
                # warning in the eventual decision_basis below.
                n_locks_unhonored += 1
                lock_warning = (
                    f"NOTE: lock to {locked_sup} ignored — no priced bid"
                )

        # Gather eligible bids for this item per the strategy filters
        candidates = []
        for sup in suppliers:
            b = bid_lookup.get((sup, rfq_key))
            if not b:
                continue
            if exclude_uom and b["status"] == BID_STATUS_UOM_DISC:
                continue
            if exclude_subs and b["status"] == BID_STATUS_SUBSTITUTE:
                continue
            candidates.append((sup, b["effective_price"], b["status"]))

        if not candidates:
            awards.append({
                "rfq_key": rfq_key, "item_num": it["item_num"], "description": it["description"],
                "qty_24mo": qty, "awarded_supplier": None, "awarded_price": None,
                "awarded_value": 0, "historical_price": hist_price,
                "historical_value": hist_price * qty, "savings_value": 0,
                "decision_basis": "NO_BID — no eligible supplier",
            })
            n_no_award += 1
            historical_total += hist_price * qty
            continue

        # Pick the awarded supplier per strategy
        candidates.sort(key=lambda c: c[1])  # cheapest first
        chosen_sup, chosen_price, chosen_status = candidates[0]
        decision = "lowest"

        if strategy == "incumbent_preferred":
            # If the incumbent has a bid, prefer them unless competition saves
            # at least incumbent_threshold pct
            inc_bid = next((c for c in candidates if c[0].lower() == (incumbent or "").lower()), None)
            if inc_bid:
                inc_price = inc_bid[1]
                if chosen_price <= 0 or (inc_price - chosen_price) / inc_price < incumbent_threshold:
                    chosen_sup, chosen_price, chosen_status = inc_bid
                    decision = f"INCUMBENT_KEPT (savings to switch < {incumbent_threshold*100:.0f}%)"

        elif strategy == "consolidate_to" and consolidate_supplier:
            # Default to consolidate winner unless another supplier saves enough
            # to justify a carve. OR-logic — fires when EITHER pct savings is
            # structural (>=carve_threshold) OR annual $ savings clears the
            # absolute floor (>=carve_threshold_dollar). Mirrors the rule in
            # compute_consolidation_analysis so the headline analysis and the
            # saved-scenario evaluation never disagree.
            target_bid = next((c for c in candidates if c[0] == consolidate_supplier), None)
            if target_bid:
                target_price = target_bid[1]
                pct_savings = (target_price - chosen_price) / target_price if target_price else 0.0
                annual_savings = (target_price - chosen_price) * (qty / 2.0) if qty else 0.0
                fires_pct = pct_savings >= carve_threshold
                fires_dollar = annual_savings >= carve_threshold_dollar
                if fires_pct or fires_dollar:
                    rule = "BOTH" if (fires_pct and fires_dollar) else ("PCT" if fires_pct else "DOLLAR")
                    decision = f"CARVE[{rule}]: {chosen_sup} saves {pct_savings*100:.0f}% / ${annual_savings:,.0f}/yr vs {consolidate_supplier}"
                    n_carved += 1
                else:
                    chosen_sup, chosen_price, chosen_status = target_bid
                    decision = f"CONSOLIDATE_TO {consolidate_supplier}"
            else:
                # Consolidation winner didn't bid this item — fall through to lowest
                decision = f"WINNER_NOBID — fell through to {chosen_sup} (lowest)"

        # If chosen supplier matches incumbent, count retained
        if incumbent and chosen_sup.lower() == (incumbent or "").lower():
            incumbent_retained += 1
        else:
            items_switched += 1

        award_value = chosen_price * qty
        award_total += award_value
        historical_total += hist_price * qty
        by_supplier[chosen_sup] = by_supplier.get(chosen_sup, 0) + award_value

        decision_with_warning = (
            f"{decision} — {lock_warning}" if lock_warning else decision
        )
        awards.append({
            "rfq_key": rfq_key, "item_num": it["item_num"], "description": it["description"],
            "qty_24mo": qty, "awarded_supplier": chosen_sup, "awarded_price": chosen_price,
            "awarded_value": award_value, "historical_price": hist_price,
            "historical_value": hist_price * qty, "savings_value": (hist_price - chosen_price) * qty if hist_price else 0,
            "decision_basis": decision_with_warning,
        })

    return {
        "strategy": strategy,
        "parameters": parameters,
        "n_items": len(awards),
        "n_awarded": len([a for a in awards if a["awarded_supplier"]]),
        "n_no_award": n_no_award,
        "n_manual_overrides": n_manual,
        "n_carved": n_carved,
        "n_locked": n_locked,
        "n_locks_unhonored": n_locks_unhonored,
        "items_switched": items_switched,
        "incumbent_retained": incumbent_retained,
        "award_total": award_total,
        "historical_total": historical_total,
        "savings_total": historical_total - award_total,
        "savings_pct": (historical_total - award_total) / historical_total * 100.0 if historical_total else 0.0,
        "award_by_supplier": by_supplier,
        "awards": awards,
    }


def _summarize_eval_for_headline(evaluated: dict, consolidate_supplier: str = None) -> dict:
    """Compress a full _evaluate_scenario result into the slim shape the
    headline card needs: which supplier wins the most $, totals, savings,
    carve-out count. Used only for the chip-strip pre-compute — never mutates
    state. Picks `supplier_primary` as the supplier with the largest awarded
    $ (concentration) so the headline's "AWARD TO X" reads true even when
    a strategy splits awards across suppliers.
    """
    by_sup = evaluated.get("award_by_supplier") or {}
    award_total = evaluated.get("award_total") or 0.0
    primary_sup = None
    primary_value = 0.0
    if by_sup:
        primary_sup, primary_value = max(by_sup.items(), key=lambda kv: kv[1] or 0.0)
    primary_pct = (primary_value / award_total * 100.0) if award_total else 0.0
    return {
        "supplier_primary": primary_sup,
        "supplier_primary_value": primary_value,
        "supplier_primary_pct": primary_pct,
        "award_total": award_total,
        "historical_total": evaluated.get("historical_total") or 0.0,
        "savings_total": evaluated.get("savings_total") or 0.0,
        "savings_pct": evaluated.get("savings_pct") or 0.0,
        "n_items": evaluated.get("n_items") or 0,
        "n_awarded": evaluated.get("n_awarded") or 0,
        "n_carved": evaluated.get("n_carved") or 0,
        "n_locked": evaluated.get("n_locked") or 0,
        "n_locks_unhonored": evaluated.get("n_locks_unhonored") or 0,
        "n_no_award": evaluated.get("n_no_award") or 0,
        "award_by_supplier": dict(by_sup),
        "consolidate_supplier": consolidate_supplier,
    }


def compute_headline_strategies(consolidate_supplier: str = None) -> dict:
    """Pre-compute headline numbers for all 5 award strategies so the chip
    switcher in step 4's headline card can flip between them with no Python
    round-trip. Each strategy's totals / supplier mix / savings / carve-outs
    are computed from the current bids + item_locks + item_exclusions state.

    Argument:
      consolidate_supplier — which supplier to consolidate to. If None, defaults
        to the top consolidation candidate from compute_consolidation_analysis.

    Returns:
      {
        "strategies": {
          "lowest_price":        {summary},
          "lowest_qualified":    {summary},
          "incumbent_preferred": {summary},
          "consolidate_to":      {summary, "consolidate_supplier": str},
          "manual":              {summary} | None,   # only present if scenarios named "manual" exist
        },
        "default_chip":                  "consolidate_to" if a supplier is available else "lowest_qualified",
        "default_consolidate_supplier":  str | None,
        "available_consolidate_suppliers": [str, ...],     # all suppliers with priced bids
        "manual_overrides": {
          "n_locks":           int,
          "n_exclusions":      int,    # number of items with at least one excluded line
          "n_uom_resolutions": int,    # UOM annotations the analyst has applied
        },
        "thresholds": {
          "carve_out_min_savings_pct":            float,
          "carve_out_min_savings_annual_dollar":  float,
        },
      }

    All five strategies use the SAME item_locks / overrides plumbing in
    _evaluate_scenario, so a locked item shows up identically in every chip.
    The carve-out OR-rule (% or $/yr) applies to consolidate_to only.
    """
    bids_by_supplier = _STATE.get("bids", {}) or {}
    suppliers_with_priced = []
    for sup, parsed in bids_by_supplier.items():
        for b in parsed.get("bids", []):
            if b.get("status") == BID_STATUS_PRICED and b.get("effective_price"):
                suppliers_with_priced.append(sup)
                break

    consol = compute_consolidation_analysis()
    candidates = consol.get("candidates") or []
    default_consolidate = candidates[0]["supplier"] if candidates else None
    if consolidate_supplier is None:
        consolidate_supplier = default_consolidate

    strategies = {}
    for strat in ("lowest_price", "lowest_qualified", "incumbent_preferred"):
        evaluated = _evaluate_scenario(strat, {}, {})
        strategies[strat] = _summarize_eval_for_headline(evaluated)

    if consolidate_supplier:
        evaluated = _evaluate_scenario("consolidate_to", {"supplier": consolidate_supplier}, {})
        strategies["consolidate_to"] = _summarize_eval_for_headline(evaluated, consolidate_supplier)
    else:
        strategies["consolidate_to"] = None

    # Manual is only meaningful if the analyst saved one — not pre-computed
    # here. The chip is enabled regardless; clicking it surfaces saved manual
    # scenarios from the drawer.
    strategies["manual"] = None

    item_locks = _STATE.get("item_locks", {}) or {}
    item_exclusions = _STATE.get("item_exclusions", {}) or {}
    uom_annotations = _STATE.get("uom_annotations", {}) or {}
    manual_overrides = {
        "n_locks": len(item_locks),
        "n_exclusions": sum(1 for v in item_exclusions.values() if v),
        "n_uom_resolutions": len([k for k, v in uom_annotations.items() if v and v.get("factor") is not None]),
    }

    th = get_thresholds()
    return {
        "strategies": strategies,
        "default_chip": "consolidate_to" if default_consolidate else "lowest_qualified",
        "default_consolidate_supplier": default_consolidate,
        "available_consolidate_suppliers": suppliers_with_priced,
        "manual_overrides": manual_overrides,
        "thresholds": {
            "carve_out_min_savings_pct": th["carve_out_min_savings_pct"],
            "carve_out_min_savings_annual_dollar": th["carve_out_min_savings_annual_dollar"],
        },
    }


def reset_to_auto() -> dict:
    """Clear analyst-applied manual overrides — locks, item exclusions, and
    UOM annotations — so the system returns to its purely-auto recommendation.
    Audit-logged. Returns the post-reset manual_overrides counts (should all
    be zero) for the UI to confirm.
    """
    n_locks = len(_STATE.get("item_locks") or {})
    n_excl = sum(1 for v in (_STATE.get("item_exclusions") or {}).values() if v)
    n_uom = len(_STATE.get("uom_annotations") or {})

    # Recompute aggregates for items we're un-excluding so last_unit_price /
    # qty_*/spend_* go back to the raw historicals.
    affected_items = [k for k, v in (_STATE.get("item_exclusions") or {}).items() if v]
    _STATE["item_exclusions"] = {}
    _STATE["item_locks"] = {}
    _STATE["uom_annotations"] = {}
    for it_num in affected_items:
        _recompute_item_aggregates_for(it_num)
    _rebuild_kpis_from_items()

    log_event(
        "reset_to_auto",
        f"cleared {n_locks} locks · {n_excl} item exclusions · {n_uom} UOM annotations",
    )
    return {
        "cleared": {
            "n_locks": n_locks,
            "n_exclusions": n_excl,
            "n_uom_resolutions": n_uom,
        },
        "manual_overrides": {
            "n_locks": 0,
            "n_exclusions": 0,
            "n_uom_resolutions": 0,
        },
    }


def save_award_scenario(name: str, strategy: str, parameters: dict = None,
                        overrides: dict = None, included_keys=None,
                        notes: str = "") -> dict:
    """Evaluate the strategy + parameters and save the result as a named scenario."""
    if not name:
        raise ValueError("name required")
    if strategy not in SCENARIO_STRATEGIES:
        raise ValueError(f"strategy must be one of {SCENARIO_STRATEGIES}")
    parameters = parameters or {}
    overrides = overrides or {}

    evaluated = _evaluate_scenario(strategy, parameters, overrides, included_keys)

    # Persist (drop the per-item awards array from the saved scenario — it's
    # large; re-derived on demand. Keep totals + parameters + overrides.)
    sc = _get_scenarios()
    sc[name] = {
        "name": name,
        "strategy": strategy,
        "parameters": parameters,
        "overrides": overrides,
        "included_keys": list(included_keys) if included_keys is not None else None,
        "notes": notes,
        "saved_at": datetime.now().isoformat(),
        "totals": {k: evaluated[k] for k in
                   ("n_items", "n_awarded", "n_no_award", "n_manual_overrides",
                    "n_carved", "items_switched", "incumbent_retained",
                    "award_total", "historical_total", "savings_total",
                    "savings_pct", "award_by_supplier")},
    }
    log_event(
        "save_award_scenario",
        f"strategy={strategy} / award_total=${evaluated['award_total']:,.0f} / savings=${evaluated['savings_total']:,.0f}",
        related=name,
    )
    return sc[name]


def evaluate_award_scenario(name: str) -> dict:
    """Re-run a saved scenario against current items + bids and return the
    full per-item awards array + roll-up totals."""
    sc = _get_scenarios()
    s = sc.get(name)
    if not s:
        raise ValueError(f"No scenario named {name!r}")
    return _evaluate_scenario(
        s["strategy"], s.get("parameters", {}),
        s.get("overrides", {}), s.get("included_keys"),
    )


def compare_award_scenarios(name_a: str, name_b: str) -> dict:
    """Side-by-side comparison of two scenarios — totals delta + per-item
    differences (items where the awarded supplier differs)."""
    a = evaluate_award_scenario(name_a)
    b = evaluate_award_scenario(name_b)
    awards_a = {x["rfq_key"]: x for x in a["awards"]}
    awards_b = {x["rfq_key"]: x for x in b["awards"]}
    all_keys = set(awards_a) | set(awards_b)
    diffs = []
    for k in all_keys:
        ax = awards_a.get(k) or {}
        bx = awards_b.get(k) or {}
        sup_a = ax.get("awarded_supplier")
        sup_b = bx.get("awarded_supplier")
        if sup_a != sup_b or ax.get("awarded_price") != bx.get("awarded_price"):
            diffs.append({
                "rfq_key": k,
                "item_num": ax.get("item_num") or bx.get("item_num"),
                "description": ax.get("description") or bx.get("description"),
                "qty_24mo": ax.get("qty_24mo") or bx.get("qty_24mo") or 0,
                "supplier_a": sup_a, "price_a": ax.get("awarded_price"), "value_a": ax.get("awarded_value"),
                "supplier_b": sup_b, "price_b": bx.get("awarded_price"), "value_b": bx.get("awarded_value"),
                "value_delta": (bx.get("awarded_value") or 0) - (ax.get("awarded_value") or 0),
            })
    diffs.sort(key=lambda d: abs(d["value_delta"]), reverse=True)
    # Drop the per-item awards arrays from a/b before returning to keep payload manageable
    summary_a = {k: v for k, v in a.items() if k != "awards"}
    summary_b = {k: v for k, v in b.items() if k != "awards"}
    return {
        "scenario_a": {"name": name_a, **summary_a},
        "scenario_b": {"name": name_b, **summary_b},
        "summary_delta": {
            "award_total":     b["award_total"]     - a["award_total"],
            "historical_total":b["historical_total"]- a["historical_total"],
            "savings_total":   b["savings_total"]   - a["savings_total"],
            "items_switched":  b["items_switched"]  - a["items_switched"],
            "incumbent_retained": b["incumbent_retained"] - a["incumbent_retained"],
            "n_awarded":       b["n_awarded"]       - a["n_awarded"],
            "n_no_award":      b["n_no_award"]      - a["n_no_award"],
        },
        "n_items_differ": len(diffs),
        "diffs": diffs[:500],   # cap
    }


# ---------------------------------------------------------------------------
# Award letter xlsx generator — per-supplier, with strict isolation guard.
#
# Generates one xlsx per awarded supplier from a saved scenario. Hard rule:
# every row in the output filters to a single awarded_supplier; defensive
# double-check at write refuses the export if any row carries another
# supplier's id. No other supplier's bids, no internal target / cost /
# margin in the file.
# ---------------------------------------------------------------------------

class IsolationViolation(Exception):
    """Raised if an award letter would contain rows for a supplier other than
    the named one. Blocks the export rather than silently leaking data."""


def gen_award_letter_xlsx(scenario_name: str, supplier_name: str,
                          rfq_id: str = "",
                          contact_name: str = "",
                          contact_email: str = "") -> bytes:
    profile = get_user_profile()
    if not contact_name:
        contact_name = profile.get("name") or "(operator name not set)"
    if not contact_email:
        contact_email = profile.get("email") or "(operator email not set)"
    contact_company = profile.get("company") or "Andersen"
    """Build the award-letter xlsx for one awarded supplier from a saved
    scenario. Strict isolation enforced both during row collection and
    via a defensive re-scan before write."""
    if not scenario_name:
        raise ValueError("scenario_name required")
    if not supplier_name:
        raise ValueError("supplier_name required")

    evaluated = evaluate_award_scenario(scenario_name)
    all_awards = evaluated.get("awards", [])
    # Strict filter: only this supplier's awards
    rows = [a for a in all_awards if a.get("awarded_supplier") == supplier_name]
    if not rows:
        raise ValueError(
            f"No items awarded to {supplier_name!r} in scenario {scenario_name!r}. "
            f"Nothing to send."
        )

    # Defensive: scan again to be absolutely sure
    for r in rows:
        actual = r.get("awarded_supplier")
        if actual != supplier_name:
            raise IsolationViolation(
                f"Row {r.get('item_num')!r} has awarded_supplier={actual!r}, "
                f"expected {supplier_name!r}. Refusing to export."
            )

    if not rfq_id:
        rfq_id = f"RFQ-{datetime.now().strftime('%Y-%m')}-001"

    award_total = sum((r.get("awarded_value") or 0) for r in rows)
    n_items = len(rows)

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="0a0e1a")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BANNER_FONT = Font(bold=True, color="ffb733", size=16)
    LABEL_FONT = Font(bold=True, color="000000", size=11)
    THIN = Side(border_style="thin", color="999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # ---- TAB 1: Cover ----
    ws = wb.active
    ws.title = "Award Letter"
    ws.append([f"AWARD NOTIFICATION — {rfq_id}"])
    ws["A1"].font = BANNER_FONT
    ws.append([])
    ws.append([f"Awarded to: {supplier_name}"])
    ws["A3"].font = LABEL_FONT
    ws.append([f"Award date: {datetime.now().strftime('%Y-%m-%d')}"])
    ws.append([f"Items awarded: {n_items:,}"])
    ws.append([f"Estimated annual value (at quoted prices): ${award_total:,.2f}"])
    ws.append([f"Contact: {contact_name} — {contact_email}"])
    ws.append([])
    ws.append(["Notification of award"])
    ws[f"A{ws.max_row}"].font = LABEL_FONT
    body = [
        f"This letter confirms that {supplier_name} has been awarded the items listed in the 'Awarded Items' tab,",
        "based on your response to our recent Request for Quotation. Awards are issued at the unit prices you",
        "quoted; estimated annual quantities are based on our prior 24 months of consumption and may vary.",
        "",
        "Purchase orders will be issued on a per-need basis. Pricing is firm through the validity period you",
        "indicated in your quote, or 90 days from this notification, whichever is later, unless otherwise agreed.",
        "",
        "If you cannot fulfill any awarded item at the quoted price, please notify us within 5 business days.",
        "",
        "Please confirm receipt of this award by replying to the contact above.",
    ]
    for line in body:
        ws.append([line])
    ws.column_dimensions["A"].width = 110
    for r_idx in range(1, ws.max_row + 1):
        ws.cell(row=r_idx, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # ---- TAB 2: Awarded Items ----
    ws2 = wb.create_sheet("Awarded Items")
    headers = [
        "Line #", "Andersen Item #", "Description",
        "Estimated Annual Qty", "Awarded Unit Price", "Estimated Annual Value",
        "Decision basis",
    ]
    ws2.append(headers)
    for c in ws2[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
    # Sort by value desc so the biggest awards are at the top
    sorted_rows = sorted(rows, key=lambda r: r.get("awarded_value") or 0, reverse=True)
    for i, r in enumerate(sorted_rows, start=1):
        ws2.append([
            i,
            r.get("item_num") or "",
            r.get("description") or "",
            r.get("qty_24mo") or 0,
            r.get("awarded_price"),
            r.get("awarded_value") or 0,
            r.get("decision_basis") or "",
        ])
    # Number formats
    for r in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for c in r: c.number_format = "#,##0"
    for r in ws2.iter_rows(min_row=2, min_col=5, max_col=6):
        for c in r: c.number_format = "$#,##0.00"
    # Total row
    ws2.append([])
    ws2.append(["", "", "", "TOTAL ESTIMATED ANNUAL VALUE", "", award_total, ""])
    ws2[ws2.max_row][3].font = LABEL_FONT
    ws2[ws2.max_row][5].font = Font(bold=True, size=12)
    ws2[ws2.max_row][5].number_format = "$#,##0.00"
    autosize(ws2)
    ws2.freeze_panes = "A2"

    # ---- TAB 3: Terms and Conditions ----
    ws3 = wb.create_sheet("Terms and Conditions")
    ws3.append(["Award Terms and Conditions"])
    ws3["A1"].font = BANNER_FONT
    ws3.append([])
    terms = [
        ("Award basis", "Net unit prices as quoted. Estimated annual quantities are forecasts only, not guaranteed minimums."),
        ("Pricing validity", "Firm through the validity period you stated, or 90 days from this notification, whichever is later."),
        ("Order issuance", "Individual purchase orders will be issued as needs arise. This award does not constitute a purchase order."),
        ("Lead time", "Quoted lead times apply. Notify us in advance if lead times will change."),
        ("Tariff and freight", "As noted on your quote. Notify us before any change."),
        ("Substitutes", "If you offered an alternate part on a quoted line, the alternate is awarded. Form/fit/function equivalence applies."),
        ("Quality", "All items must meet the specifications described in our RFQ documentation."),
        ("Payment terms", "Per the Andersen master supplier agreement currently in effect with your company."),
        ("Confidentiality", "Pricing in this award is confidential between Andersen and your company."),
    ]
    ws3.append(["Topic", "Detail"])
    for c in ws3[3]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for topic, detail in terms:
        ws3.append([topic, detail])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 100
    for row in ws3.iter_rows(min_row=4):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    # FINAL DEFENSIVE PASS — if anything in the awarded-items sheet has a
    # supplier name in a cell, refuse the export. (We don't write supplier
    # names in cells but this is belt-and-suspenders.)
    bids_by_supplier = _STATE.get("bids", {}) or {}
    other_supplier_names = [s for s in bids_by_supplier.keys() if s != supplier_name]
    for r in ws2.iter_rows(values_only=True):
        for v in r:
            if v is None: continue
            sv = str(v)
            for other in other_supplier_names:
                if other and other in sv:
                    raise IsolationViolation(
                        f"Award letter for {supplier_name!r} would contain "
                        f"the string {other!r}. Refusing to export."
                    )

    buf = io.BytesIO()
    wb.save(buf)
    log_event(
        "gen_award_letter",
        f"scenario={scenario_name} / {n_items:,} items / ${award_total:,.0f}",
        related=supplier_name,
    )
    return buf.getvalue()


def gen_award_letters_for_scenario(scenario_name: str, rfq_id: str = "") -> dict:
    """Generate one award letter per awarded supplier in a scenario.
    Returns: {supplier_name: xlsx_bytes, ...}. Skips suppliers with 0 awards."""
    evaluated = evaluate_award_scenario(scenario_name)
    awards = evaluated.get("awards", [])
    awarded_suppliers = sorted(set(
        a.get("awarded_supplier") for a in awards if a.get("awarded_supplier")
    ))
    out = {}
    for sup in awarded_suppliers:
        try:
            out[sup] = gen_award_letter_xlsx(scenario_name, sup, rfq_id=rfq_id)
        except (ValueError, IsolationViolation) as e:
            # Skip suppliers with no awards or isolation issues — surface in result
            out[sup] = None
    return out


# ---------------------------------------------------------------------------
# Award decision documentation — legal-hold record explaining WHY each award
# was made. Template-based prose (no AI). Designed for retention for several
# years; should be IMMUTABLE after creation (data embedded verbatim, not
# re-derived from current state). Generated alongside the award letters.
#
# Each award gets a multi-paragraph explanation:
#   1. PRIMARY REASON — why this supplier was chosen
#   2. ALTERNATIVES CONSIDERED — why other bidders weren't chosen (per supplier)
#   3. EXPECTED OUTCOME — qty / annual value / savings vs historical
#   4. RISK FACTORS — lead time, MOQ, single-source, UOM verification needed
#   5. METHODOLOGY NOTE — what scenario / strategy / thresholds were active
# ---------------------------------------------------------------------------

# ----------------- PRIMARY REASON templates -----------------
# Keyed by the decision_basis prefix string emitted by _evaluate_scenario.
# Each template is paragraph prose with named variables. Template language
# is intentionally consistent (not varied) — legal docs benefit from
# predictable phrasing.

PRIMARY_REASON_TEMPLATES = {
    "lowest": (
        "{supplier} was awarded this item as the lowest qualified quote received, "
        "at {price} per {uom}. "
        "{n_competing_text}"
        "Total annualized award value at this price is {annual_value}, "
        "representing an estimated {savings_text} versus our historical paid price of {hist_price} per {uom}."
    ),
    "lowest_no_history": (
        "{supplier} was awarded this item as the lowest qualified quote received, "
        "at {price} per {uom}. {n_competing_text}"
        "No historical paid price was available for direct savings comparison; "
        "this award establishes the baseline going forward."
    ),
    "INCUMBENT_KEPT": (
        "Award retained with the incumbent supplier {supplier} at {price} per {uom}. "
        "While {comp_supplier} did submit a lower quote ({comp_price} per {uom}, "
        "or {comp_savings_pct} below the incumbent), the savings from switching "
        "fall below our project-level switching threshold of {switch_threshold_pct}. "
        "Continuity with the incumbent supplier preserves established quality, "
        "lead-time, and operational relationships, and avoids the administrative "
        "cost and risk of supplier transition for a marginal-savings opportunity."
    ),
    "CONSOLIDATE_TO": (
        "{supplier} was awarded this item as part of a deliberate consolidation "
        "strategy under this RFQ. The procurement objective is to award the "
        "majority of the requested items to a single supplier where pricing is "
        "reasonably competitive, in order to reduce split-supplier complexity, "
        "simplify ordering and receiving, consolidate freight, and strengthen "
        "the negotiating position with the awarded supplier. {supplier}'s "
        "quoted price of {price} per {uom} was competitive with the alternatives "
        "received and within the carve-out tolerance of {carve_threshold_pct}, "
        "so this item was awarded to the consolidation supplier rather than "
        "carved out separately. {savings_text_vs_history}"
    ),
    "CARVE": (
        "{supplier} was awarded this item as a deliberate exception (carve-out) "
        "to the broader consolidation strategy. While the consolidation supplier "
        "({consol_supplier}) also bid this item ({consol_price} per {uom}), "
        "{supplier}'s quote of {price} per {uom} represents a {carve_savings_pct} "
        "savings versus the consolidation supplier on this specific item — exceeding "
        "our project-level carve-out threshold of {carve_threshold_pct}. "
        "On the estimated annual quantity of {qty} units, this represents "
        "{carve_savings_total} of additional annual savings versus consolidating "
        "this item. The materiality of the per-item savings justifies the "
        "operational complexity of a split award for this line. {savings_text_vs_history}"
    ),
    "WINNER_NOBID": (
        "{supplier} was awarded this item because the consolidation winner "
        "({consol_supplier}) did not provide a quote for it. Of the {n_responding} "
        "suppliers who did respond, {supplier} offered the lowest qualified price "
        "at {price} per {uom}. {savings_text_vs_history}"
    ),
    "MANUAL": (
        "{supplier} was awarded this item by manual decision of the procurement "
        "operator, overriding the engine's automated recommendation. The operator's "
        "stated rationale: \"{manual_reason}\". This award reflects the operator's "
        "judgment based on factors that may not be fully captured by the automated "
        "scoring (e.g., supplier relationship, prior quality experience, strategic "
        "alignment, or item criticality). All bid data has been preserved for review."
    ),
    "MANUAL_NO_AWARD": (
        "No supplier was awarded this item, by manual decision of the procurement "
        "operator. Operator rationale: \"{manual_reason}\". The bid data and "
        "engine recommendation are preserved for reference."
    ),
    "NO_BID": (
        "No supplier was awarded this item because no qualified bid was received. "
        "{n_no_bid_text}"
        "{recommendation_text}"
    ),
    "lowest_qualified": (
        "{supplier} was awarded this item at {price} per {uom} as the lowest "
        "qualified quote. {excluded_text}"
        "Total annualized award value at this price is {annual_value}, "
        "representing an estimated {savings_text} versus our historical paid price of {hist_price} per {uom}."
    ),
}

# ----------------- ALTERNATIVES-CONSIDERED templates (per non-winning bidder) -----------------

ALT_REASON_TEMPLATES = {
    "PRICE_HIGHER": (
        "{supplier}: quoted {price} per {uom}, which is {pct_higher} higher than the awarded price. "
        "Lower-cost option was awarded."
    ),
    "PRICE_HIGHER_BUT_QUALIFIED": (
        "{supplier}: quoted {price} per {uom} (qualified bid). At {pct_higher} above the awarded price, "
        "this bid was not selected on a price-only comparison."
    ),
    "UOM_DISCREPANCY": (
        "{supplier}: quoted {price} per {uom}, but flagged a unit-of-measure discrepancy "
        "in their notes (\"{notes_excerpt}\"). Without UOM normalization or supplier "
        "confirmation, this quote could not be directly compared to the others. "
        "Will follow up via the supplier follow-up packet."
    ),
    "SUBSTITUTE_OFFERED": (
        "{supplier}: quoted {price} per {uom} for an alternate part ({alt_part_text}). "
        "Substitute parts require form/fit/function validation before award; "
        "the bid was preserved but not selected pending substitute approval."
    ),
    "NEED_INFO": (
        "{supplier}: did not provide a price; their response indicated that additional "
        "information was needed (\"{notes_excerpt}\"). Will follow up with clarification."
    ),
    "NO_BID": (
        "{supplier}: declined to quote this item{notes_excerpt_text}."
    ),
    "OUTLIER_HIGH": (
        "{supplier}: quoted {price} per {uom}, which is {factor:.1f}× the median of the "
        "received bids. This is statistically anomalous and likely indicates a "
        "wrong-part match, pricing error, or UOM mismatch on the supplier side. "
        "Excluded from award consideration pending supplier clarification."
    ),
    "OUTLIER_LOW": (
        "{supplier}: quoted {price} per {uom}, which is only {factor:.2f}× the median bid. "
        "This is suspiciously low and was not awarded on price alone — the operator "
        "should verify part identity and UOM with the supplier before treating "
        "as a real opportunity."
    ),
    "DID_NOT_RESPOND": (
        "{supplier}: did not respond on this item (no row submitted in their bid response file)."
    ),
}

# ----------------- EXPECTED OUTCOME templates -----------------

OUTCOME_TEMPLATES = {
    "WITH_SAVINGS": (
        "Estimated annual quantity is {qty} units, derived from the trailing 24 months "
        "of consumption. At the awarded unit price of {price} per {uom}, total "
        "annualized award value is {annual_value}. Versus the historical paid price "
        "of {hist_price} per {uom}, this award is projected to deliver {savings_amount} "
        "in annual savings, or approximately {savings_pct} of historical spend on this item."
    ),
    "WITH_INCREASE": (
        "Estimated annual quantity is {qty} units, derived from the trailing 24 months "
        "of consumption. At the awarded unit price of {price} per {uom}, total "
        "annualized award value is {annual_value}. The awarded price is {increase_pct} "
        "higher than the historical paid price of {hist_price} per {uom}, representing "
        "an annualized cost increase of {increase_amount}. Acceptance reflects the "
        "absence of a more competitive qualified quote in this RFQ."
    ),
    "NO_HISTORY": (
        "Estimated annual quantity is {qty} units, derived from the trailing 24 months "
        "of consumption. At the awarded unit price of {price} per {uom}, total "
        "annualized award value is {annual_value}. No historical paid price is on file "
        "for this item, so this award establishes the new baseline."
    ),
    "ZERO_QTY": (
        "Annual quantity estimate is zero based on the trailing 24-month order history "
        "(this item has not been ordered in that window). The awarded unit price of {price} "
        "per {uom} will apply if and when the item is ordered, but no annualized value "
        "is being committed."
    ),
}

# ----------------- RISK FACTOR templates (appended when conditions are met) -----------------

RISK_TEMPLATES = {
    "SINGLE_SOURCE_AWARD": (
        "Risk: This item received only one qualified bid, from {supplier}. "
        "Until the next RFQ cycle, Andersen has no competitive backup for this item. "
        "Operator should monitor lead times and pricing closely."
    ),
    "UOM_VERIFICATION_REQUIRED": (
        "Risk: The awarded supplier flagged a UOM discrepancy on their quote. "
        "Procurement should confirm the UOM in writing before issuing the first PO; "
        "the apparent savings may not be real once UOM is normalized."
    ),
    "SUBSTITUTE_PENDING_VALIDATION": (
        "Risk: The awarded supplier offered an alternate part ({alt_part_text}). "
        "Form/fit/function equivalence has not yet been validated by engineering. "
        "First-article approval should precede full release to production."
    ),
    "DEMAND_VOLATILITY": (
        "Risk: This item's recent demand pattern shows {demand_change_text}, "
        "so the annualized quantity estimate carries above-average uncertainty. "
        "Actual order volumes may differ materially from the estimate; "
        "supplier capacity planning should not assume the estimate is firm."
    ),
    "DORMANT_ITEM": (
        "Risk: This item has not been ordered in the last 12 months. "
        "Award price is a forward-looking commitment that may never be invoked. "
        "Consider whether this item should remain in the active item master."
    ),
    "ONE_TIME_BUY_HISTORY": (
        "Risk: This item has only been ordered once historically. "
        "Annual quantity estimates are unreliable for one-time-buy patterns; "
        "actual demand may be much higher or zero."
    ),
    "PRICE_SPIKE_CONTEXT": (
        "Note: The historical paid price reflects a recent price change. "
        "The latest line in our purchase history was {latest_price} per {uom}, which is "
        "{spike_pct} above the 90-day median ({median_price} per {uom}). "
        "Operator should be aware that the savings calculation is sensitive to which "
        "historical baseline is used."
    ),
    "GENERIC_DESCRIPTION": (
        "Risk: This item's description in our system is generic or short, which makes "
        "supplier cross-referencing difficult. Confirm that the awarded supplier is "
        "quoting the same physical part by reviewing the manufacturer part number."
    ),
}


def _format_money(v):
    if v is None:
        return "—"
    return f"${v:,.2f}"


def _format_pct(v):
    if v is None:
        return "—"
    return f"{v*100:.1f}%"


def _format_factor(v):
    if v is None:
        return "—"
    return f"{v:.2f}×"


def _safe_text(s, fallback=""):
    if s is None:
        return fallback
    s = str(s).strip()
    return s if s else fallback


def _excerpt_notes(notes, max_len=120):
    if not notes:
        return ""
    n = str(notes).strip()
    if len(n) > max_len:
        n = n[:max_len].rsplit(" ", 1)[0] + "…"
    return n


def generate_award_rationale(award_record: dict, comparison_row: dict, scenario: dict) -> dict:
    """Build the full multi-paragraph rationale for one award. Returns:
        {
          "primary_reason": str,            — paragraph 1: why this supplier
          "alternatives_considered": [str], — paragraph 2: per-non-winner reasoning
          "expected_outcome": str,          — paragraph 3: annual qty / value / savings
          "risk_factors": [str],            — paragraph 4: warnings (may be empty)
          "methodology_note": str,          — paragraph 5: scenario / thresholds
          "full_text": str,                 — concatenation suitable for a single cell
        }

    All paragraphs are template-based (no AI). Numbers are formatted to-the-penny
    or to-the-percent for legal-grade precision.
    """
    th = get_thresholds()
    awarded_supplier = award_record.get("awarded_supplier")
    awarded_price = award_record.get("awarded_price")
    qty = award_record.get("qty_24mo") or 0
    annual_value = award_record.get("awarded_value") or 0
    hist_price = award_record.get("historical_price") or 0
    historical_value = award_record.get("historical_value") or 0
    savings_value = award_record.get("savings_value") or 0
    decision_basis = award_record.get("decision_basis") or ""
    item_uom = (comparison_row.get("uom") if comparison_row else None) or "unit"
    bids = (comparison_row.get("bids") if comparison_row else None) or {}

    def _bid_for(sup):
        return bids.get(sup) or {}

    # ---------- PRIMARY REASON ----------
    primary = ""
    qty_basis = "the trailing 24 months of order history"

    if not awarded_supplier:
        # No-award branch
        n_no_bid = sum(1 for b in bids.values() if b.get("status") == "NO_BID")
        n_need_info = sum(1 for b in bids.values() if b.get("status") == "NEED_INFO")
        rec_text = ""
        if hist_price > 0 and qty > 0:
            rec_text = (
                f"Recommend either re-issuing the request to a broader supplier list, "
                f"sole-sourcing with the incumbent, or removing this item from the active RFQ. "
                f"The historical annual run-rate at this item is approximately "
                f"{_format_money(hist_price * qty)}, which provides context for the level of effort warranted."
            )
        else:
            rec_text = (
                "Given the limited or absent historical run-rate, recommend dropping this item from "
                "the active RFQ unless a specific need is identified."
            )
        if decision_basis.startswith("MANUAL"):
            manual_reason = decision_basis.replace("MANUAL: ", "").replace("MANUAL", "").strip(": ")
            primary = PRIMARY_REASON_TEMPLATES["MANUAL_NO_AWARD"].format(
                manual_reason=manual_reason or "no rationale provided"
            )
        else:
            no_bid_text = ""
            if n_no_bid:
                no_bid_text = f"{n_no_bid} supplier{'s' if n_no_bid != 1 else ''} declined to quote. "
            if n_need_info:
                no_bid_text += f"{n_need_info} supplier{'s' if n_need_info != 1 else ''} indicated more information was needed. "
            if not no_bid_text:
                no_bid_text = "No qualified responses were received. "
            primary = PRIMARY_REASON_TEMPLATES["NO_BID"].format(
                n_no_bid_text=no_bid_text,
                recommendation_text=rec_text,
            )
    else:
        # Award branch — pick template by decision_basis
        # Calculate competitive context
        priced_alternatives = [
            (s, b.get("price")) for s, b in bids.items()
            if s != awarded_supplier and b.get("price") is not None and b.get("status") in ("PRICED", "UOM_DISC", "SUBSTITUTE")
        ]
        n_competing = len(priced_alternatives)
        if n_competing == 0:
            n_competing_text = "This was the only qualified quote received for this item. "
        elif n_competing == 1:
            other_sup, other_price = priced_alternatives[0]
            pct_higher = ((other_price - awarded_price) / awarded_price) if awarded_price else 0
            n_competing_text = (
                f"One other qualified bid was received "
                f"({other_sup} at {_format_money(other_price)} per {item_uom}, "
                f"{_format_pct(pct_higher)} higher). "
            )
        else:
            other_min = min(p for _, p in priced_alternatives)
            other_max = max(p for _, p in priced_alternatives)
            n_competing_text = (
                f"{n_competing} other qualified bids were received, ranging from "
                f"{_format_money(other_min)} to {_format_money(other_max)} per {item_uom}, "
                f"all higher than the awarded price. "
            )

        # Savings text
        if hist_price and hist_price > 0:
            sv_pct = (hist_price - awarded_price) / hist_price
            if sv_pct > 0:
                savings_text = (
                    f"savings of {_format_pct(sv_pct)} on this item "
                    f"(approximately {_format_money(savings_value)} per year on the estimated annual qty of {qty:,.0f})"
                )
            elif sv_pct < 0:
                savings_text = (
                    f"price increase of {_format_pct(abs(sv_pct))} versus the prior price "
                    f"(approximately {_format_money(abs(savings_value))} per year on the estimated annual qty of {qty:,.0f})"
                )
            else:
                savings_text = "essentially flat versus the prior price"
            savings_text_vs_history = "Versus the historical paid price of " + _format_money(hist_price) + " per " + item_uom + ", this represents a " + savings_text + "."
        else:
            savings_text = "no historical baseline available"
            savings_text_vs_history = "No prior paid price was on file for this item, so this award establishes the baseline."

        # Pick the right template
        if decision_basis.startswith("MANUAL"):
            manual_reason = decision_basis.split(":", 1)[1].strip() if ":" in decision_basis else "no rationale captured"
            primary = PRIMARY_REASON_TEMPLATES["MANUAL"].format(
                supplier=awarded_supplier,
                manual_reason=manual_reason,
            )
        elif decision_basis.startswith("INCUMBENT_KEPT"):
            comp = next(((s, p) for s, p in priced_alternatives), (None, None))
            comp_supplier = comp[0] or "competitor"
            comp_price = comp[1] or 0
            comp_savings_pct = ((awarded_price - comp_price) / awarded_price) if (awarded_price and comp_price) else 0
            primary = PRIMARY_REASON_TEMPLATES["INCUMBENT_KEPT"].format(
                supplier=awarded_supplier,
                price=_format_money(awarded_price),
                uom=item_uom,
                comp_supplier=comp_supplier,
                comp_price=_format_money(comp_price),
                comp_savings_pct=_format_pct(comp_savings_pct),
                switch_threshold_pct=_format_pct(th["min_savings_pct_to_switch"]),
            )
        elif decision_basis.startswith("CONSOLIDATE_TO"):
            primary = PRIMARY_REASON_TEMPLATES["CONSOLIDATE_TO"].format(
                supplier=awarded_supplier,
                price=_format_money(awarded_price),
                uom=item_uom,
                carve_threshold_pct=_format_pct(th["carve_out_min_savings_pct"]),
                savings_text_vs_history=savings_text_vs_history,
            )
        elif decision_basis.startswith("CARVE"):
            # Decision basis like "CARVE: MSC saves 30% vs Grainger"
            consol_supplier = (scenario.get("parameters") or {}).get("supplier") if scenario else "the consolidation winner"
            consol_bid = _bid_for(consol_supplier)
            consol_price = consol_bid.get("price") or 0
            carve_savings_pct = ((consol_price - awarded_price) / consol_price) if (consol_price and consol_price > 0) else 0
            carve_savings_total = (consol_price - awarded_price) * qty if (consol_price and qty) else 0
            primary = PRIMARY_REASON_TEMPLATES["CARVE"].format(
                supplier=awarded_supplier,
                consol_supplier=consol_supplier or "the consolidation winner",
                price=_format_money(awarded_price),
                consol_price=_format_money(consol_price),
                uom=item_uom,
                qty=f"{qty:,.0f}",
                carve_savings_pct=_format_pct(carve_savings_pct),
                carve_savings_total=_format_money(carve_savings_total),
                carve_threshold_pct=_format_pct(th["carve_out_min_savings_pct"]),
                savings_text_vs_history=savings_text_vs_history,
            )
        elif decision_basis.startswith("WINNER_NOBID"):
            consol_supplier = (scenario.get("parameters") or {}).get("supplier") if scenario else "the consolidation winner"
            primary = PRIMARY_REASON_TEMPLATES["WINNER_NOBID"].format(
                supplier=awarded_supplier,
                consol_supplier=consol_supplier or "the consolidation winner",
                price=_format_money(awarded_price),
                uom=item_uom,
                n_responding=n_competing + 1,
                savings_text_vs_history=savings_text_vs_history,
            )
        else:
            # Default: lowest-priced award
            tmpl_key = "lowest" if hist_price > 0 else "lowest_no_history"
            primary = PRIMARY_REASON_TEMPLATES[tmpl_key].format(
                supplier=awarded_supplier,
                price=_format_money(awarded_price),
                uom=item_uom,
                n_competing_text=n_competing_text,
                annual_value=_format_money(annual_value),
                savings_text=savings_text,
                hist_price=_format_money(hist_price),
            )

    # ---------- ALTERNATIVES CONSIDERED ----------
    alt_lines = []
    for sup, b in bids.items():
        if sup == awarded_supplier:
            continue
        status = b.get("status")
        price = b.get("price")
        notes = b.get("notes") or ""
        notes_excerpt = _excerpt_notes(notes, 120)
        raw = b.get("raw") or {}
        if status == "MISSING":
            alt_lines.append(ALT_REASON_TEMPLATES["DID_NOT_RESPOND"].format(supplier=sup))
        elif status == "NO_BID":
            note_text = f' (\"{notes_excerpt}\")' if notes_excerpt else ""
            alt_lines.append(ALT_REASON_TEMPLATES["NO_BID"].format(
                supplier=sup,
                notes_excerpt_text=note_text,
            ))
        elif status == "NEED_INFO":
            alt_lines.append(ALT_REASON_TEMPLATES["NEED_INFO"].format(
                supplier=sup,
                notes_excerpt=notes_excerpt or "no detail provided",
            ))
        elif status == "UOM_DISC":
            alt_lines.append(ALT_REASON_TEMPLATES["UOM_DISCREPANCY"].format(
                supplier=sup,
                price=_format_money(price),
                uom=item_uom,
                notes_excerpt=notes_excerpt or "UOM mismatch noted",
            ))
        elif status == "SUBSTITUTE":
            alt_part = raw.get("sub_part") or "alternate part"
            alt_desc = raw.get("sub_desc") or ""
            alt_part_text = f"{alt_part}" + (f" — {alt_desc}" if alt_desc else "")
            alt_lines.append(ALT_REASON_TEMPLATES["SUBSTITUTE_OFFERED"].format(
                supplier=sup,
                price=_format_money(price),
                uom=item_uom,
                alt_part_text=alt_part_text,
            ))
        elif price is not None and awarded_price is not None and price > 0 and awarded_price > 0:
            # Outlier check
            other_priced = [b2.get("price") for b2 in bids.values()
                            if b2.get("price") is not None and b2.get("price") > 0]
            if len(other_priced) >= 2:
                med = _median(other_priced)
                if med and price >= med * th["outlier_factor"]:
                    alt_lines.append(ALT_REASON_TEMPLATES["OUTLIER_HIGH"].format(
                        supplier=sup,
                        price=_format_money(price),
                        uom=item_uom,
                        factor=price / med,
                    ))
                    continue
                elif med and price <= med / th["outlier_factor"]:
                    alt_lines.append(ALT_REASON_TEMPLATES["OUTLIER_LOW"].format(
                        supplier=sup,
                        price=_format_money(price),
                        uom=item_uom,
                        factor=price / med,
                    ))
                    continue
            pct_higher = ((price - awarded_price) / awarded_price) if awarded_price else 0
            alt_lines.append(ALT_REASON_TEMPLATES["PRICE_HIGHER"].format(
                supplier=sup,
                price=_format_money(price),
                uom=item_uom,
                pct_higher=_format_pct(pct_higher),
            ))
        else:
            alt_lines.append(ALT_REASON_TEMPLATES["DID_NOT_RESPOND"].format(supplier=sup))

    # ---------- EXPECTED OUTCOME ----------
    if not awarded_supplier:
        outcome = ""
    elif qty == 0:
        outcome = OUTCOME_TEMPLATES["ZERO_QTY"].format(
            price=_format_money(awarded_price), uom=item_uom,
        )
    elif hist_price and hist_price > 0:
        if awarded_price <= hist_price:
            outcome = OUTCOME_TEMPLATES["WITH_SAVINGS"].format(
                qty=f"{qty:,.0f}",
                price=_format_money(awarded_price),
                uom=item_uom,
                annual_value=_format_money(annual_value),
                hist_price=_format_money(hist_price),
                savings_amount=_format_money(savings_value),
                savings_pct=_format_pct(savings_value / historical_value if historical_value else 0),
            )
        else:
            increase_pct = (awarded_price - hist_price) / hist_price if hist_price else 0
            increase_amount = (awarded_price - hist_price) * qty
            outcome = OUTCOME_TEMPLATES["WITH_INCREASE"].format(
                qty=f"{qty:,.0f}",
                price=_format_money(awarded_price),
                uom=item_uom,
                annual_value=_format_money(annual_value),
                hist_price=_format_money(hist_price),
                increase_pct=_format_pct(increase_pct),
                increase_amount=_format_money(increase_amount),
            )
    else:
        outcome = OUTCOME_TEMPLATES["NO_HISTORY"].format(
            qty=f"{qty:,.0f}",
            price=_format_money(awarded_price),
            uom=item_uom,
            annual_value=_format_money(annual_value),
        )

    # ---------- RISK FACTORS ----------
    risks = []
    if awarded_supplier:
        # Single-source check
        n_priced = sum(1 for b in bids.values() if b.get("price") is not None and b.get("price") > 0)
        if n_priced == 1:
            risks.append(RISK_TEMPLATES["SINGLE_SOURCE_AWARD"].format(supplier=awarded_supplier))
        # UOM verification
        awarded_bid = _bid_for(awarded_supplier)
        if awarded_bid.get("status") == "UOM_DISC":
            risks.append(RISK_TEMPLATES["UOM_VERIFICATION_REQUIRED"])
        # Substitute pending
        if awarded_bid.get("status") == "SUBSTITUTE":
            raw = awarded_bid.get("raw") or {}
            alt_part = raw.get("sub_part") or "an alternate part"
            alt_desc = raw.get("sub_desc") or ""
            alt_part_text = alt_part + (f" — {alt_desc}" if alt_desc else "")
            risks.append(RISK_TEMPLATES["SUBSTITUTE_PENDING_VALIDATION"].format(alt_part_text=alt_part_text))
    # Item-level risk flags (from comparison_row)
    if comparison_row:
        # Demand flags
        # We don't have demand flags in comparison_row directly — they're on the
        # underlying item record. Fetch from _STATE.items for completeness.
        item_record = next(
            (it for it in _STATE.get("items", []) if it["item_num"] == comparison_row.get("item_num")),
            None,
        )
        if item_record:
            dflags = item_record.get("demand_flags") or []
            if "DORMANT_12MO" in dflags:
                risks.append(RISK_TEMPLATES["DORMANT_ITEM"])
            elif "SINGLE_ORDER" in dflags:
                risks.append(RISK_TEMPLATES["ONE_TIME_BUY_HISTORY"])
            elif "DEMAND_DROP_50" in dflags:
                risks.append(RISK_TEMPLATES["DEMAND_VOLATILITY"].format(
                    demand_change_text="a drop of more than 50% versus the prior 12 months"
                ))
            elif "DEMAND_SURGE_50" in dflags:
                risks.append(RISK_TEMPLATES["DEMAND_VOLATILITY"].format(
                    demand_change_text="a surge of more than 50% versus the prior 12 months"
                ))
            # Generic description
            desc_flags = item_record.get("desc_flags") or []
            if "generic" in desc_flags:
                risks.append(RISK_TEMPLATES["GENERIC_DESCRIPTION"])

    # ---------- METHODOLOGY NOTE ----------
    strategy = (scenario or {}).get("strategy", "?")
    parameters = (scenario or {}).get("parameters", {})
    th = get_thresholds()
    methodology = (
        f"This award was generated under the \"{(scenario or {}).get('name', '?')}\" scenario "
        f"using the {strategy} strategy"
    )
    if strategy == "consolidate_to" and parameters.get("supplier"):
        methodology += f" with consolidation supplier set to {parameters['supplier']}"
    methodology += (
        f". Carve-out threshold at the time of decision: {_format_pct(th['carve_out_min_savings_pct'])}; "
        f"minimum savings to switch from incumbent: {_format_pct(th['min_savings_pct_to_switch'])}; "
        f"price-pushback threshold: {_format_pct(th['pushback_threshold_pct'])}; "
        f"outlier factor: {th['outlier_factor']}×; "
        f"UOM-suspect ratio: {th['uom_suspect_ratio']}×. "
        f"All historical pricing reflects the trailing 24 months of order data; "
        f"\"last paid price\" is the unit price of the most recent priced order line."
    )

    # ---------- COMPOSE FULL TEXT ----------
    sections = []
    if primary:
        sections.append("PRIMARY REASON:\n" + primary)
    if alt_lines:
        sections.append("ALTERNATIVES CONSIDERED:\n" + "\n".join(f"• {l}" for l in alt_lines))
    if outcome:
        sections.append("EXPECTED OUTCOME:\n" + outcome)
    if risks:
        sections.append("RISK FACTORS / FOLLOW-UP ACTIONS:\n" + "\n".join(f"• {r}" for r in risks))
    if methodology:
        sections.append("METHODOLOGY:\n" + methodology)
    full_text = "\n\n".join(sections)

    return {
        "primary_reason": primary,
        "alternatives_considered": alt_lines,
        "expected_outcome": outcome,
        "risk_factors": risks,
        "methodology_note": methodology,
        "full_text": full_text,
    }


def gen_decision_log_xlsx(scenario_name: str, rfq_id: str = "",
                          decided_by: str = "",
                          retention_years: int = 7) -> bytes:
    if not decided_by:
        profile = get_user_profile()
        decided_by = profile.get("name") or "(operator name not set)"
    """Build the immutable, legal-hold decision log for one award scenario.

    Captures every awarded item's full rationale + the methodology snapshot +
    the audit log of all engine actions. Designed for retention for several
    years; embeds data verbatim rather than re-deriving on reopen.
    """
    if not scenario_name:
        raise ValueError("scenario_name required")
    sc = _get_scenarios().get(scenario_name)
    if not sc:
        raise ValueError(f"No scenario named {scenario_name!r}")

    evaluated = evaluate_award_scenario(scenario_name)
    matrix = compute_comparison_matrix(included_keys=sc.get("included_keys"))
    rows_by_key = {r["rfq_key"]: r for r in matrix["rows"]}

    if not rfq_id:
        rfq_id = f"RFQ-{datetime.now().strftime('%Y-%m')}-001"

    th_snapshot = get_thresholds()
    generated_at = datetime.now()
    retention_until = datetime(generated_at.year + retention_years, generated_at.month, generated_at.day)

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="0a0e1a")
    BANNER_FILL = PatternFill("solid", fgColor="1c2540")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BANNER_FONT = Font(bold=True, color="ffb733", size=18)
    LABEL_FONT = Font(bold=True, color="000000", size=11)
    SECTION_FONT = Font(bold=True, color="000000", size=12)
    NOTE_FONT = Font(italic=True, color="666666", size=10)

    # ---- TAB 1: Cover + Provenance ----
    ws = wb.active
    ws.title = "Cover"
    ws.append([f"DECISION LOG — {rfq_id}"])
    ws["A1"].font = BANNER_FONT
    ws.append([])
    ws.append(["This is the official record of award decisions for the above RFQ."])
    ws.append(["It is intended for retention as a defensible audit record of how each supplier was selected."])
    ws.append([])
    ws.append(["IMMUTABILITY NOTICE"])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    ws.append(["This document embeds the bid data, recommendation reasoning, threshold values, and audit trail"])
    ws.append(["that were in effect at the time the awards were made. It is not regenerated from current state"])
    ws.append(["and should not be modified after creation. If the underlying data, thresholds, or scenario"])
    ws.append(["change, generate a new decision log dated to that change rather than editing this one."])
    ws.append([])
    ws.append(["RFQ Provenance"])
    ws[f"A{ws.max_row}"].font = SECTION_FONT
    rows_meta = [
        ("RFQ ID", rfq_id),
        ("Scenario name", scenario_name),
        ("Award strategy", sc.get("strategy")),
        ("Strategy parameters", json.dumps(sc.get("parameters") or {}, default=str) or "{}"),
        ("Items in scenario", evaluated.get("n_items")),
        ("Items awarded", evaluated.get("n_awarded")),
        ("Items not awarded", evaluated.get("n_no_award")),
        ("Manual overrides", evaluated.get("n_manual_overrides")),
        ("Carve-outs (consolidation)", evaluated.get("n_carved", 0)),
        ("Total awarded value (annualized)", evaluated.get("award_total")),
        ("Historical baseline (annualized)", evaluated.get("historical_total")),
        ("Estimated annual savings", evaluated.get("savings_total")),
        ("Estimated savings %", f"{evaluated.get('savings_pct', 0):.1f}%"),
        ("", ""),
        ("Decided by", decided_by),
        ("Decision date", generated_at.strftime("%Y-%m-%d %H:%M")),
        ("Retain until at least", retention_until.strftime("%Y-%m-%d") + f"  ({retention_years}-year retention)"),
        ("Tool version", "auto-rfq-banana"),
    ]
    for label, value in rows_meta:
        ws.append([label, value])
    # Format $ rows
    for r_idx, (label, _) in enumerate(rows_meta, start=ws.max_row - len(rows_meta) + 1):
        if "value" in label.lower() or "baseline" in label.lower() or "savings" in label.lower():
            try:
                cell = ws.cell(row=r_idx, column=2)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "$#,##0.00"
            except Exception:
                pass
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    # ---- TAB 2: Threshold Snapshot ----
    ws_th = wb.create_sheet("Thresholds At Decision Time")
    ws_th.append([f"Threshold Snapshot — {generated_at.strftime('%Y-%m-%d %H:%M')}"])
    ws_th["A1"].font = SECTION_FONT
    ws_th.append([])
    ws_th.append(["These are the engine threshold values that were in effect when the awards were made."])
    ws_th.append(["If thresholds are changed later, this record reflects the values used at decision time, not current values."])
    ws_th.append([])
    ws_th.append(["Threshold key", "Value at decision time"])
    for c in ws_th[ws_th.max_row]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for k, v in th_snapshot.items():
        ws_th.append([k, v])
    ws_th.column_dimensions["A"].width = 40
    ws_th.column_dimensions["B"].width = 25

    # ---- TAB 3: Award-by-Item Rationale ----
    ws3 = wb.create_sheet("Award Rationale")
    ws3.append([
        "Item #", "Description", "MFG", "MFG PN", "Annual Qty", "UOM",
        "Awarded Supplier", "Awarded Price", "Annual Award Value",
        "Historical Price", "Annual Savings",
        "Decision Basis (engine)",
        "Primary Reason",
        "Alternatives Considered",
        "Expected Outcome",
        "Risk Factors",
        "Methodology Note",
    ])
    for c in ws3[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for award in sorted(evaluated.get("awards", []),
                        key=lambda a: a.get("awarded_value") or 0, reverse=True):
        comp_row = rows_by_key.get(award.get("rfq_key"), {})
        rationale = generate_award_rationale(award, comp_row, sc)
        item_record = next((it for it in _STATE.get("items", []) if it["item_num"] == award.get("item_num")), {})
        ws3.append([
            award.get("item_num"),
            award.get("description"),
            item_record.get("mfg_name", ""),
            item_record.get("mfg_pn", ""),
            award.get("qty_24mo"),
            item_record.get("uom", ""),
            award.get("awarded_supplier") or "(no award)",
            award.get("awarded_price"),
            award.get("awarded_value") or 0,
            award.get("historical_price") or 0,
            award.get("savings_value") or 0,
            award.get("decision_basis") or "",
            rationale["primary_reason"],
            "\n".join(f"• {l}" for l in rationale["alternatives_considered"]),
            rationale["expected_outcome"],
            "\n".join(f"• {r}" for r in rationale["risk_factors"]),
            rationale["methodology_note"],
        ])
    # Number formats
    for r in ws3.iter_rows(min_row=2, min_col=5, max_col=5):
        for c in r: c.number_format = "#,##0"
    for r in ws3.iter_rows(min_row=2, min_col=8, max_col=11):
        for c in r: c.number_format = "$#,##0.00"
    # Wrap rationale columns
    for r in ws3.iter_rows(min_row=2, min_col=13, max_col=17):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    # Column widths
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 40
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 11
    ws3.column_dimensions["F"].width = 8
    ws3.column_dimensions["G"].width = 18
    ws3.column_dimensions["H"].width = 14
    ws3.column_dimensions["I"].width = 16
    ws3.column_dimensions["J"].width = 14
    ws3.column_dimensions["K"].width = 14
    ws3.column_dimensions["L"].width = 32
    ws3.column_dimensions["M"].width = 90
    ws3.column_dimensions["N"].width = 70
    ws3.column_dimensions["O"].width = 70
    ws3.column_dimensions["P"].width = 60
    ws3.column_dimensions["Q"].width = 80
    ws3.freeze_panes = "A2"

    # ---- TAB 4: Bid Data Preserved (immutable record of bids received) ----
    ws4 = wb.create_sheet("Bid Data")
    ws4.append([
        "Item #", "Supplier", "Status", "Quoted Price", "Verified Price",
        "Effective Price", "Quoted UOM", "Notes",
        "Their Part #", "Alternate Part Offered", "Alternate Description",
    ])
    for c in ws4[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bids_by_supplier = _STATE.get("bids", {}) or {}
    for sup, parsed in bids_by_supplier.items():
        for b in parsed.get("bids", []):
            # Only include bids for items in this scenario
            if sc.get("included_keys") is not None:
                # Match by item_num to scenario's included_keys
                pass
            ws4.append([
                b.get("item_num", "") or b.get("part_number", ""),
                sup,
                b.get("status", ""),
                b.get("quoted_price"),
                b.get("verified_price"),
                b.get("effective_price"),
                b.get("uom", ""),
                _excerpt_notes(b.get("notes", ""), 200),
                b.get("exact_part") or b.get("part_number", ""),
                b.get("sub_part", ""),
                b.get("sub_desc", ""),
            ])
    for r in ws4.iter_rows(min_row=2, min_col=4, max_col=6):
        for c in r: c.number_format = "$#,##0.00"
    autosize(ws4, max_w=50)
    ws4.freeze_panes = "A2"

    # ---- TAB 5: Audit Trail ----
    ws5 = wb.create_sheet("Audit Trail")
    ws5.append(["Timestamp", "Action", "Detail", "Related"])
    for c in ws5[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for entry in list_audit_log():
        ws5.append([
            entry.get("timestamp", ""),
            entry.get("action_type", ""),
            entry.get("action_detail", ""),
            entry.get("related", ""),
        ])
    autosize(ws5, max_w=80)
    ws5.freeze_panes = "A2"

    # ---- TAB 6: Methodology Notes ----
    ws6 = wb.create_sheet("Methodology")
    ws6.append(["How the engine made these decisions"])
    ws6["A1"].font = SECTION_FONT
    ws6.append([])
    methodology_paragraphs = [
        "Item identity: Items are identified by an internal key derived from the Andersen Item Number, falling back to EAM Part Number, then to the supplier's catalog Part Number. For most Andersen suppliers EAM is the primary identifier; cXML / PunchOut suppliers (e.g. McMaster) frequently leave Item # and EAM blank, in which case the supplier's own catalog SKU is used as the join key.",
        "Window aggregations: Quantities and spend are aggregated across rolling 12-month, 24-month, and 36-month windows anchored to the most recent order date in the source data. The annualized quantity used for award value calculations is the trailing 24-month quantity divided by 2.",
        "Last paid price: The unit price of the most recent priced order line. No medians, smoothing, or rolling averages are used in displayed RFQ pricing — to-the-penny precision is preserved throughout. Where this may obscure a price spike, a separate analytical reference (90-day median) is shown on the per-item drill-down chart but is not used in award math.",
        "Recommendation engine: Each item with bids receives a deterministic recommendation in one of five categories: ACCEPT, PUSH_BACK, ASK_CLARIFICATION, EXCLUDE, or MANUAL_REVIEW. The decision rules are evaluated in priority order: no-bid → EXCLUDE; all-need-info → ASK_CLARIFICATION; UOM discrepancy on lowest bid → ASK_CLARIFICATION; substitute on lowest bid → ASK_CLARIFICATION; lowest bid above the historical baseline by more than the pushback threshold → PUSH_BACK; statistical outlier → MANUAL_REVIEW; savings exceed the switch threshold → ACCEPT; otherwise MANUAL_REVIEW.",
        "Award scenarios: The actual awards in this decision log were generated under a named scenario with one of five strategies — lowest_price, lowest_qualified, consolidate_to (with a named consolidation supplier), incumbent_preferred, or manual. Manual overrides at the per-item level take precedence over scenario logic.",
        "Consolidation with carve-outs: The default award strategy at Andersen for this category is to consolidate the bulk of items to a single supplier and to carve out individual items where another supplier saves significantly more. The carve-out threshold (the per-item savings vs the consolidation winner that triggers a carve-out) is recorded in the Threshold Snapshot tab.",
        "UOM-discrepancy guard: Carve-outs where the alternative supplier flagged a unit-of-measure mismatch (or where the price ratio versus the consolidation winner is extreme) are flagged separately and excluded from counted savings until the supplier confirms the UOM. This prevents false-positive savings from per-each vs per-package pricing mismatches.",
        "Cross-supplier isolation: Every supplier-bound output (RFQ files, award letters, follow-up packets) is filtered at write time to a single supplier's data. A defensive cell-level scan refuses the export if any other supplier's name appears in the workbook. The verify_isolation.py sibling script provides a third-party cross-check on a folder of outputs.",
    ]
    for p in methodology_paragraphs:
        ws6.append([p])
        cell = ws6.cell(row=ws6.max_row, column=1)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws6.row_dimensions[ws6.max_row].height = 75
    ws6.column_dimensions["A"].width = 130

    log_event("gen_decision_log", f"scenario={scenario_name} / awards={evaluated.get('n_awarded')}", related=rfq_id)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gen_decision_log_markdown(scenario_name: str, rfq_id: str = "",
                              decided_by: str = "",
                              retention_years: int = 7) -> str:
    """Markdown version of the decision log. Designed to be pasted into
    Microsoft 365 Copilot (which is IT-approved at Andersen) to ask
    follow-up questions, generate executive summaries, draft email replies
    to suppliers, etc. — without ever putting the data through a
    third-party AI from inside this app."""
    if not decided_by:
        decided_by = (get_user_profile().get("name") or "(operator name not set)")
    sc = _get_scenarios().get(scenario_name)
    if not sc:
        raise ValueError(f"No scenario named {scenario_name!r}")
    evaluated = evaluate_award_scenario(scenario_name)
    matrix = compute_comparison_matrix(included_keys=sc.get("included_keys"))
    rows_by_key = {r["rfq_key"]: r for r in matrix["rows"]}
    th = get_thresholds()
    if not rfq_id:
        rfq_id = f"RFQ-{datetime.now().strftime('%Y-%m')}-001"
    generated = datetime.now()
    retention_until = datetime(generated.year + retention_years, generated.month, generated.day)

    out = []
    out.append(f"# Decision Log — {rfq_id}")
    out.append("")
    out.append("> **How to use this file with Copilot**")
    out.append(">")
    out.append("> This markdown file mirrors the contents of the decision log xlsx. ")
    out.append("> M365 Copilot is approved at Andersen, so you can paste the relevant section ")
    out.append("> into Copilot Chat to ask follow-up questions, draft an email summary for a ")
    out.append("> supplier, generate an executive briefing, or stress-test the rationale. ")
    out.append("> Examples of useful Copilot prompts:")
    out.append("> - *\"Summarize this decision log into a 5-bullet executive briefing.\"*")
    out.append("> - *\"Draft a polite reply to {supplier} explaining why they were not awarded.\"*")
    out.append("> - *\"List any items where the rationale relies on a single bid — these need follow-up.\"*")
    out.append("> - *\"Extract every push-back recommendation as a checklist for next round.\"*")
    out.append("")
    out.append("## Provenance")
    out.append("")
    out.append(f"- **RFQ ID:** {rfq_id}")
    out.append(f"- **Scenario:** {scenario_name} ({sc.get('strategy')})")
    out.append(f"- **Decided by:** {decided_by}")
    out.append(f"- **Decision date:** {generated.strftime('%Y-%m-%d %H:%M')}")
    out.append(f"- **Retain until at least:** {retention_until.strftime('%Y-%m-%d')}  ({retention_years}-year retention)")
    out.append("")
    out.append("## Headline numbers")
    out.append("")
    out.append(f"- Items in scenario: **{evaluated.get('n_items'):,}**")
    out.append(f"- Items awarded: **{evaluated.get('n_awarded'):,}**")
    out.append(f"- Items not awarded: {evaluated.get('n_no_award'):,}")
    out.append(f"- Manual overrides: {evaluated.get('n_manual_overrides'):,}")
    out.append(f"- Carve-outs: {evaluated.get('n_carved', 0):,}")
    out.append(f"- Total awarded value (annualized): **${evaluated.get('award_total'):,.2f}**")
    out.append(f"- Historical baseline (annualized): ${evaluated.get('historical_total'):,.2f}")
    out.append(f"- Estimated annual savings: **${evaluated.get('savings_total'):,.2f}** (~{evaluated.get('savings_pct',0):.1f}%)")
    out.append("")
    out.append("## Thresholds at decision time")
    out.append("")
    out.append("| Threshold | Value |")
    out.append("| --- | --- |")
    for k, v in th.items():
        out.append(f"| {k} | {v} |")
    out.append("")
    out.append("## Per-item award rationale")
    out.append("")
    awards_sorted = sorted(evaluated.get("awards", []),
                           key=lambda a: a.get("awarded_value") or 0, reverse=True)
    for a in awards_sorted:
        comp_row = rows_by_key.get(a.get("rfq_key"), {})
        rationale = generate_award_rationale(a, comp_row, sc)
        item = a.get("item_num", "?")
        sup = a.get("awarded_supplier") or "(no award)"
        price = a.get("awarded_price")
        price_str = f"${price:.2f}" if price is not None else "—"
        annual = a.get("awarded_value") or 0
        out.append(f"### {item} — awarded to **{sup}** at {price_str}")
        out.append("")
        out.append(f"*Annual award value: ${annual:,.2f}  ·  Decision basis: `{a.get('decision_basis') or ''}`*")
        out.append("")
        out.append("**Primary reason:**  ")
        out.append(rationale["primary_reason"])
        out.append("")
        if rationale["alternatives_considered"]:
            out.append("**Alternatives considered:**")
            for alt in rationale["alternatives_considered"]:
                out.append(f"- {alt}")
            out.append("")
        if rationale["expected_outcome"]:
            out.append("**Expected outcome:**  ")
            out.append(rationale["expected_outcome"])
            out.append("")
        if rationale["risk_factors"]:
            out.append("**Risk factors / follow-up actions:**")
            for r in rationale["risk_factors"]:
                out.append(f"- {r}")
            out.append("")
        out.append("---")
        out.append("")
    out.append("")
    out.append("## Methodology")
    out.append("")
    methodology_paragraphs = [
        "Item identity: Items are identified by an internal key derived from the Andersen Item Number, falling back to EAM Part Number, then to the supplier's catalog Part Number.",
        "Window aggregations: Quantities and spend are aggregated across rolling 12-month, 24-month, and 36-month windows anchored to the most recent order date in the source data.",
        "Last paid price: The unit price of the most recent priced order line. No medians or smoothing in displayed RFQ pricing.",
        "Recommendation engine: Each item with bids gets a deterministic 5-tier recommendation (ACCEPT / PUSH_BACK / ASK_CLARIFICATION / EXCLUDE / MANUAL_REVIEW) with a concrete reason string.",
        "Award scenarios: Awards generated under one of five strategies — lowest_price, lowest_qualified, consolidate_to (with a named consolidation supplier), incumbent_preferred, or manual.",
        "UOM-discrepancy guard: Carve-outs flagged with a UOM warning are excluded from counted savings until the supplier confirms.",
        "Cross-supplier isolation: Every supplier-bound output filters to a single supplier's data; defensive cell-level scan refuses export if any other supplier's name appears.",
    ]
    for p in methodology_paragraphs:
        out.append(f"- {p}")
    log_event("gen_decision_log_md", f"scenario={scenario_name}", related=rfq_id)
    return "\n".join(out)


def gen_internal_award_summary_xlsx(scenario_name: str, rfq_id: str = "") -> bytes:
    """Internal-audience summary of the full award scenario. Includes the
    cross-supplier comparison + per-supplier totals + every award decision.

    Banner: 'INTERNAL — NEVER FORWARD'. Contains pricing across all suppliers
    side-by-side; explicitly NOT supplier-bound.
    """
    evaluated = evaluate_award_scenario(scenario_name)
    sc = _get_scenarios().get(scenario_name) or {}

    if not rfq_id:
        rfq_id = f"RFQ-{datetime.now().strftime('%Y-%m')}-001"

    wb = Workbook()
    HEADER_FILL = PatternFill("solid", fgColor="0a0e1a")
    BANNER_FILL = PatternFill("solid", fgColor="ff4d6d")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BANNER_FONT = Font(bold=True, color="FFFFFF", size=16)
    LABEL_FONT = Font(bold=True, color="000000", size=11)

    # ---- TAB 1: Internal Summary ----
    ws = wb.active
    ws.title = "INTERNAL Summary"
    ws.append(["INTERNAL — NEVER FORWARD TO ANY SUPPLIER"])
    ws["A1"].font = BANNER_FONT
    ws["A1"].fill = BANNER_FILL
    ws.append([f"Scenario: {scenario_name} · {sc.get('strategy', '')}"])
    ws.append([f"RFQ ID: {rfq_id}"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])
    ws.append(["Headline"])
    ws[f"A{ws.max_row}"].font = LABEL_FONT
    ws.append(["Items in scenario", evaluated["n_items"]])
    ws.append(["Items awarded", evaluated["n_awarded"]])
    ws.append(["Items with no award", evaluated["n_no_award"]])
    ws.append(["Items switched (vs incumbent)", evaluated["items_switched"]])
    ws.append(["Items kept with incumbent", evaluated["incumbent_retained"]])
    ws.append(["Manual overrides", evaluated["n_manual_overrides"]])
    ws.append(["Carve-outs (consolidate strategy only)", evaluated.get("n_carved", 0)])
    ws.append([])
    ws.append(["Award total (annualized)", evaluated["award_total"]])
    ws.append(["Historical baseline (annualized)", evaluated["historical_total"]])
    ws.append(["Estimated annual savings", evaluated["savings_total"]])
    ws.append(["Savings %", f"{evaluated['savings_pct']:.1f}%"])
    for r in (8, 9, 10, 11, 12, 13, 14):
        c = ws.cell(row=r, column=2)
        c.number_format = "#,##0"
    for r in (16, 17, 18):
        c = ws.cell(row=r, column=2)
        c.number_format = "$#,##0.00"
    autosize(ws)

    # ---- TAB 2: Award by Supplier ----
    ws2 = wb.create_sheet("Award by Supplier")
    ws2.append(["Supplier", "Items awarded", "Total annual value"])
    for c in ws2[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    by_sup = evaluated.get("award_by_supplier", {})
    counts = {}
    for a in evaluated.get("awards", []):
        s = a.get("awarded_supplier")
        if s:
            counts[s] = counts.get(s, 0) + 1
    for sup in sorted(by_sup.keys(), key=lambda s: -by_sup[s]):
        ws2.append([sup, counts.get(sup, 0), by_sup[sup]])
    for r in ws2.iter_rows(min_row=2, min_col=3, max_col=3):
        for c in r: c.number_format = "$#,##0.00"
    autosize(ws2)

    # ---- TAB 3: All Awards (full detail, all suppliers visible) ----
    ws3 = wb.create_sheet("All Awards")
    ws3.append([
        "Item #", "Description", "Annual Qty", "Awarded Supplier",
        "Awarded Unit Price", "Awarded Annual Value",
        "Historical Unit Price", "Historical Annual Value",
        "Annual Savings", "Decision basis",
    ])
    for c in ws3[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for a in sorted(evaluated.get("awards", []), key=lambda x: x.get("awarded_value") or 0, reverse=True):
        ws3.append([
            a.get("item_num"), a.get("description"), a.get("qty_24mo"),
            a.get("awarded_supplier") or "(no award)",
            a.get("awarded_price"), a.get("awarded_value") or 0,
            a.get("historical_price") or 0, a.get("historical_value") or 0,
            a.get("savings_value") or 0,
            a.get("decision_basis"),
        ])
    for r in ws3.iter_rows(min_row=2, min_col=3, max_col=3):
        for c in r: c.number_format = "#,##0"
    for r in ws3.iter_rows(min_row=2, min_col=5, max_col=9):
        for c in r: c.number_format = "$#,##0.00"
    autosize(ws3)
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Round 2 / Rn focused-RFQ generator + intake.
#
# After Round 1 bids come back, the analyst typically wants to push back on
# specific items where one supplier was uncompetitive vs another. The R2
# flow lets them:
#   1. Select a small set of items in the comparison matrix.
#   2. Generate per-supplier focused xlsx files — read-only context + R1
#      echo (their prior price as gray context) + a Reference Price
#      computed from our cleaned trend (linear regression on priced order
#      lines, outliers excluded, projected to anchor date) + a small
#      supplier-input column block (8 fields vs 17 in R1).
#   3. Drop the returned R2 file into the same step-4 dropzone — parser
#      detects the round=2 hidden marker and OVERWRITES that supplier's
#      R1 bids for items that came back with new prices, leaving items
#      they didn't re-quote untouched.
#
# State:
#   _STATE["round2_selection"]  list of item_num the analyst has marked for R2
#   _STATE["current_round"]     int (defaults to 1; R2 generator bumps to 2;
#                                    Rn generator passes round_num explicitly)
#   _STATE["bids"][supplier]["bids"][i]["round_history"]  list of prior
#       (round_num, price, status, ts) snapshots so the audit trail shows
#       the negotiation arc per item.
# ---------------------------------------------------------------------------

REFERENCE_PRICE_BANNER = (
    "REFERENCE PRICE COLUMN — what it is and how to use it.  "
    "Based on our internal purchase history for each item, the Reference Price "
    "is what our trend math (least-squares linear regression on priced order "
    "lines, with analyst-confirmed outlier orders excluded, projected to today's "
    "anchor date) suggests a competitive bid should be in line with. This same "
    "reference is shared with every supplier participating in this round — it's "
    "our read of the data, not a supplier-specific number. IF YOUR QUOTE DIVERGES "
    "SIGNIFICANTLY FROM THIS REFERENCE, FLAG IT IN THE NOTES COLUMN — especially "
    "when the divergence is from a unit-of-measure mismatch we may have missed "
    "(e.g., we may list 'EA' while your catalog sells in boxes of 10). Catching "
    "those early avoids confusion in the comparison stage. Reference is left "
    "blank for items where our trend confidence is low (R-squared < 0.2 or fewer "
    "than 3 priced orders)."
)


def list_round2_selection() -> list:
    """Return the currently-selected item_nums for the next R2/Rn batch."""
    return list(_STATE.get("round2_selection", []) or [])


def set_round2_selection(item_nums) -> dict:
    """Persist the analyst's R2-selection set. Idempotent + dedup'd."""
    cleaned = sorted({str(n) for n in (item_nums or []) if n})
    _STATE["round2_selection"] = cleaned
    return {"n_selected": len(cleaned), "selection": cleaned}


def clear_round2_selection() -> dict:
    _STATE["round2_selection"] = []
    return {"n_selected": 0, "selection": []}


def _compute_reference_price(item_num: str):
    """Reference price for one item — uses the cleaned-set trend + anchor.

    Returns ``(price, confidence, reason)`` where price is float|None,
    confidence is "high"|"medium"|"low"|None, and reason is a short
    explanation suitable for the bid intake parser to surface in
    `notes` when the supplier's bid diverges from this reference.

    Mirrors get_item_history's trend computation but skips the spike /
    bid-overlay machinery — we only need the projected price + a
    confidence label here.
    """
    if not item_num:
        return (None, None, "no item")
    key = norm_pn(item_num)
    po_lines = _STATE.get("po_lines_by_key", {}).get(key, [])
    if not po_lines:
        return (None, None, "no priced order history")
    sorted_lines = sorted(po_lines, key=lambda r: r[0] or "")
    excluded = set(_STATE.get("item_exclusions", {}).get(item_num, []) or [])
    cleaned = [ln for idx, ln in enumerate(sorted_lines) if idx not in excluded]
    first_dt = None
    xs, ys = [], []
    for ln in cleaned:
        if not ln[0] or ln[2] is None:
            continue
        try:
            d = datetime.fromisoformat(ln[0])
        except ValueError:
            continue
        if first_dt is None:
            first_dt = d
        xs.append((d - first_dt).days)
        ys.append(float(ln[2]))
    if len(xs) < 3:
        return (None, "low", f"only {len(xs)} priced orders after exclusions")
    slope, intercept, r2 = _linear_fit(xs, ys)
    if slope is None or r2 is None:
        return (None, "low", "trend not fittable")
    if r2 < 0.2:
        return (None, "low", f"R² {r2:.2f} — too noisy to project confidently")
    confidence = "high" if r2 >= 0.6 else "medium"
    anchor = _STATE.get("data_anchor_date")
    expected = None
    if anchor and first_dt:
        try:
            anchor_dt = datetime.fromisoformat(anchor)
            anchor_x = (anchor_dt - first_dt).days
            expected = slope * anchor_x + intercept
        except ValueError:
            return (None, confidence, "could not parse anchor date")
    if expected is None or expected <= 0:
        # Trend extrapolated to a zero/negative price is meaningless.
        return (None, confidence, "trend projects an invalid price")
    return (float(expected), confidence, f"R² {r2:.2f}")


def gen_round2_rfq_xlsx(
    supplier_name: str,
    item_nums,
    round_num: int = 2,
    rfq_id: str = "",
    response_due_date: str = "",
    contact_name: str = "",
    contact_email: str = "",
    include_r1_echo: bool = True,
    include_reference_price: bool = True,
) -> bytes:
    """Build a focused R2/R3/Rn xlsx for one supplier.

    Strict isolation: the file contains only this supplier's prior bid data
    (R1 echo) + Andersen's own historical-trend reference. NO cross-supplier
    bid data leaks. The Reference Price is per-item Andersen-internal
    analytics, safe to share with every bidder participating in this round
    (the banner makes that explicit so the supplier isn't confused).

    Args:
        supplier_name: exact name (must match a key in _STATE["bids"] for
            R1-echo data to populate; if absent, R1-echo cells stay blank).
        item_nums: list of display item numbers to include.
        round_num: 2 by default; pass 3 for R3, etc.
        rfq_id, response_due_date, contact_name, contact_email: passed
            through into the Instructions tab; same shape as gen_outbound_rfq_xlsx.
        include_r1_echo: when True (default), surface R1 quote price + UOM
            + notes as a gray read-only column block. False = no echo.
        include_reference_price: when True (default), surface the per-item
            Reference Price + the explanatory banner. False = no reference.

    Returns: xlsx bytes. Filename convention (caller decides):
        Round{round_num}_RFQ_{supplier_safe}_{rfq_id}.xlsx
    """
    if not supplier_name:
        raise ValueError("supplier_name required")
    if not item_nums:
        raise ValueError("at least one item_num required")
    if round_num < 2:
        raise ValueError(f"round_num must be ≥ 2 (got {round_num})")

    items_state = _STATE.get("items", []) or []
    items_by_num = {it.get("item_num"): it for it in items_state}
    selected_items = [items_by_num[n] for n in item_nums if n in items_by_num]
    if not selected_items:
        raise ValueError("no selected item_nums match the current item list")

    # Pull the supplier's R1 bid lookup once.
    parsed = (_STATE.get("bids", {}) or {}).get(supplier_name) or {}
    r1_by_key = {}
    for b in parsed.get("bids", []) or []:
        rk = b.get("rfq_key")
        if not rk:
            continue
        prev = r1_by_key.get(rk)
        # If the supplier had multiple lines for the same item (qty break /
        # alt SKU), pick the canonical one — same priority used in the
        # per-item modal overlay (PRICED > UOM_DISC > SUBSTITUTE), then
        # lowest effective_price within ties.
        prio = {"PRICED": 0, "UOM_DISC": 1, "SUBSTITUTE": 2,
                "NEED_INFO": 3, "NO_BID": 4}.get(b.get("status") or "", 9)
        eff = b.get("effective_price") if (b.get("effective_price") and b["effective_price"] > 0) else 1e12
        if prev is None:
            r1_by_key[rk] = (prio, eff, b)
        else:
            if (prio, eff) < (prev[0], prev[1]):
                r1_by_key[rk] = (prio, eff, b)
    r1_resolved = {k: tup[2] for k, tup in r1_by_key.items()}

    wb = Workbook()

    # ---- TAB 1: Instructions ----
    ws = wb.active
    ws.title = "Instructions"
    rfq_id_for_display = rfq_id or _STATE.get("rfq_id") or ""
    _write_banner_row(ws, 1,
                      f"REQUEST FOR QUOTATION  ·  ROUND {round_num}  ·  {rfq_id_for_display}",
                      span_cols=4, font_size=20, height=46)
    ws.append([])
    rows = [
        ("Supplier",          supplier_name),
        ("Round",             f"{round_num} (focused list — {len(selected_items)} item(s))"),
        ("Response due",      response_due_date or "—"),
        ("Contact (Andersen)",contact_name or "—"),
        ("Contact email",     contact_email or "—"),
        ("Round-trip data",   "Hidden columns 'item_key' + 'rfq_line_id' + 'round' on the response template are used to match your file back to our items. Do not modify or delete these columns."),
        ("Confidentiality",   "Pricing in this RFQ and your responses are confidential between Andersen and your company."),
        ("Validity",          "Quotes should remain firm for at least 90 days unless otherwise noted in the Valid Through column."),
    ]
    for label, value in rows:
        ws.append([label, value])
        cell_label = ws.cell(row=ws.max_row, column=1)
        cell_label.font = Font(bold=True, color=COLOR_TEXT_INK)
        cell_label.fill = _fill(COLOR_REFERENCE_GRAY)
        cell_label.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        cell_val = ws.cell(row=ws.max_row, column=2)
        cell_val.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    # Reference-price banner — only when include_reference_price is on.
    if include_reference_price:
        ws.append([])
        ref_row = ws.max_row + 1
        ws.cell(row=ref_row, column=1, value=REFERENCE_PRICE_BANNER)
        ws.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=2)
        ref_cell = ws.cell(row=ref_row, column=1)
        ref_cell.font = Font(bold=False, color=COLOR_TEXT_INK, size=11)
        ref_cell.fill = _fill(COLOR_INFO_BLUE)
        ref_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        ws.row_dimensions[ref_row].height = 160

    # ---- TAB 2: Response Template (the one supplier fills out) ----
    ws2 = wb.create_sheet("Round 2 Response")

    # Build the headers dynamically based on toggles.
    ref_cols = [
        "Andersen Item #",          # A
        "EAM Part #",               # B
        "Manufacturer Part #",      # C
        "Manufacturer",             # D
        "Description",              # E
        "Annual Qty",               # F
        "Our UOM",                  # G
    ]
    r1_cols = []
    if include_r1_echo:
        r1_cols = [
            f"R{round_num - 1} echo: Quote Price",   # H
            f"R{round_num - 1} echo: Quote UOM",     # I
            f"R{round_num - 1} echo: Notes",          # J
        ]
    ref_price_cols = []
    if include_reference_price:
        ref_price_cols = [
            "Reference Price (Andersen-projected)",  # K
            "Reference confidence",                  # L
        ]
    response_cols = [
        f"R{round_num} Quote Price",          # supplier yellow
        f"R{round_num} Quote UOM",
        f"R{round_num} Your Part #",
        f"R{round_num} Lead Time Days",
        f"R{round_num} No Bid",
        f"R{round_num} No Bid Reason",
        f"R{round_num} Notes",
        f"R{round_num} Valid Through Date",
    ]
    hidden_cols = ["item_key", "rfq_line_id", "round"]

    headers = ref_cols + r1_cols + ref_price_cols + response_cols + hidden_cols
    n_total = len(headers)
    n_ref = len(ref_cols)
    n_r1 = len(r1_cols)
    n_refprice = len(ref_price_cols)
    n_response = len(response_cols)
    n_hidden = len(hidden_cols)

    # Banner with reference-price explanation embedded for at-a-glance reading.
    _write_banner_row(ws2, 1,
                      f"RFQ ROUND {round_num}  ·  {len(selected_items)} ITEM(S)  ·  {supplier_name}",
                      span_cols=n_total, font_size=18, height=42)
    ws2.append([])

    # Reference banner repeated above the table (more useful here than buried in instructions).
    if include_reference_price:
        ws2.cell(row=ws2.max_row + 1, column=1, value=REFERENCE_PRICE_BANNER)
        merged_row = ws2.max_row
        ws2.merge_cells(start_row=merged_row, start_column=1, end_row=merged_row, end_column=n_total)
        ref_cell = ws2.cell(row=merged_row, column=1)
        ref_cell.font = Font(bold=False, color=COLOR_TEXT_INK, size=11)
        ref_cell.fill = _fill(COLOR_INFO_BLUE)
        ref_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        ws2.row_dimensions[merged_row].height = 130
        ws2.append([])

    # Header row
    ws2.append(headers)
    header_row_idx = ws2.max_row
    border = _border()
    for c_idx in range(1, n_total + 1):
        cell = ws2.cell(row=header_row_idx, column=c_idx)
        cell.font = Font(bold=True, color=COLOR_TEXT_LIGHT, size=10)
        if c_idx <= n_ref:
            cell.fill = _fill(COLOR_BRAND_DARK)
        elif c_idx <= n_ref + n_r1:
            cell.fill = _fill(COLOR_BAND_DARK)  # R1 echo — secondary band
        elif c_idx <= n_ref + n_r1 + n_refprice:
            cell.fill = _fill(COLOR_BRAND_DARK)
            cell.font = Font(bold=True, color=COLOR_BRAND_AMBER, size=10)  # amber-on-dark for the reference
        elif c_idx <= n_ref + n_r1 + n_refprice + n_response:
            cell.fill = _fill(COLOR_BRAND_AMBER)
            cell.font = Font(bold=True, color=COLOR_TEXT_INK, size=10)
        else:
            cell.fill = _fill("CCCCCC")  # hidden round-trip
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        cell.border = border
    ws2.row_dimensions[header_row_idx].height = 38

    # Data rows
    for i, it in enumerate(selected_items, start=1):
        rfq_line_id = f"{rfq_id_for_display or 'rfq'}-r{round_num}-{i:05d}"
        rk = it.get("key") or norm_pn(it.get("item_num") or "")
        b = r1_resolved.get(rk) if include_r1_echo else None
        r1_price = b.get("effective_price") if b else None
        r1_uom = (b.get("uom") if b else "") or ""
        r1_notes = (b.get("notes") if b else "") or ""

        ref_price = ref_conf = ref_reason = None
        if include_reference_price:
            ref_price, ref_conf, ref_reason = _compute_reference_price(it.get("item_num"))

        row_vals = [
            it.get("item_num"),                                                  # A
            it.get("eam_pn") or "",                                              # B
            it.get("mfg_pn") or "",                                              # C
            it.get("mfg_name") or "",                                            # D
            it.get("description") or "",                                         # E
            it.get("qty_24mo") or 0,                                             # F
            it.get("uom") or "",                                                 # G
        ]
        if include_r1_echo:
            row_vals += [r1_price, r1_uom, r1_notes]
        if include_reference_price:
            row_vals += [ref_price, ref_conf or "—"]
        # Yellow response cells (8 of them) start blank
        row_vals += [None] * n_response
        row_vals += [it.get("key") or "", rfq_line_id, round_num]
        ws2.append(row_vals)

    # Color-code data rows
    n_data_rows = len(selected_items)
    REF_FILL = _fill(COLOR_REFERENCE_GRAY)
    YELLOW_FILL = _fill(COLOR_RESPONSE_YELLOW)
    R1_ECHO_FILL = _fill("D7DBE3")     # slightly darker gray to distinguish R1 echo
    REFPRICE_FILL = _fill("FFF8E1")    # very pale amber — calls attention but stays in palette
    data_start = header_row_idx + 1
    for r_idx in range(data_start, data_start + n_data_rows):
        # Reference cols (1..n_ref) — read-only gray
        for c_idx in range(1, n_ref + 1):
            c = ws2.cell(row=r_idx, column=c_idx)
            c.fill = REF_FILL
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        # R1 echo (next n_r1 cols) — slightly darker gray
        for c_idx in range(n_ref + 1, n_ref + n_r1 + 1):
            c = ws2.cell(row=r_idx, column=c_idx)
            c.fill = R1_ECHO_FILL
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        # Reference price (next n_refprice cols) — pale amber (read-only)
        for c_idx in range(n_ref + n_r1 + 1, n_ref + n_r1 + n_refprice + 1):
            c = ws2.cell(row=r_idx, column=c_idx)
            c.fill = REFPRICE_FILL
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        # Response cells (next n_response cols) — supplier-input yellow
        for c_idx in range(n_ref + n_r1 + n_refprice + 1, n_ref + n_r1 + n_refprice + n_response + 1):
            c = ws2.cell(row=r_idx, column=c_idx)
            c.fill = YELLOW_FILL
            c.border = border

    # Number formats
    # Annual Qty (col F, idx 6)
    for r in ws2.iter_rows(min_row=data_start, min_col=6, max_col=6):
        for c in r: c.number_format = "#,##0"
    # R1 echo price (col H, idx 8) when present
    if include_r1_echo:
        for r in ws2.iter_rows(min_row=data_start, min_col=n_ref + 1, max_col=n_ref + 1):
            for c in r: c.number_format = "$#,##0.00"
    # Reference price (col K) when present
    if include_reference_price:
        ref_col_idx = n_ref + n_r1 + 1
        for r in ws2.iter_rows(min_row=data_start, min_col=ref_col_idx, max_col=ref_col_idx):
            for c in r:
                if c.value is None:
                    continue
                c.number_format = "$#,##0.00"
    # R2 quote price (first response col)
    rprice_col_idx = n_ref + n_r1 + n_refprice + 1
    for r in ws2.iter_rows(min_row=data_start, min_col=rprice_col_idx, max_col=rprice_col_idx):
        for c in r: c.number_format = "$#,##0.00"

    # Hide round-trip cols
    for c_idx in range(n_total - n_hidden + 1, n_total + 1):
        ws2.column_dimensions[get_column_letter(c_idx)].hidden = True

    # Autosize visible cols
    autosize(ws2, min_w=10, max_w=44)
    ws2.freeze_panes = ws2.cell(row=data_start, column=n_ref + 1).coordinate

    log_event(
        f"round_{round_num}_rfq_generated",
        f"{supplier_name}: {len(selected_items)} items"
        + (" + ref-price" if include_reference_price else "")
        + (" + R1 echo" if include_r1_echo else ""),
        related=supplier_name,
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def gen_round2_rfqs_for_selection(
    selected_item_nums,
    suppliers,
    round_num: int = 2,
    rfq_id: str = "",
    response_due_date: str = "",
    contact_name: str = "",
    contact_email: str = "",
    include_r1_echo: bool = True,
    include_reference_price: bool = True,
) -> dict:
    """Batch wrapper — produces one xlsx per supplier for the selection.

    Returns ``{supplier: bytes}`` dict so the JS caller can write all of
    them at once. Errors per-supplier are caught and surfaced as an
    ``errors`` dict so a single bad supplier doesn't kill the batch.
    """
    out = {}
    errors = {}
    for sup in suppliers:
        try:
            out[sup] = gen_round2_rfq_xlsx(
                supplier_name=sup,
                item_nums=selected_item_nums,
                round_num=round_num,
                rfq_id=rfq_id,
                response_due_date=response_due_date,
                contact_name=contact_name,
                contact_email=contact_email,
                include_r1_echo=include_r1_echo,
                include_reference_price=include_reference_price,
            )
        except Exception as e:
            errors[sup] = str(e)
    return {"files": out, "errors": errors,
            "n_items": len(selected_item_nums or []),
            "n_suppliers": len(out)}


def parse_round2_supplier_bid(file_bytes, supplier_name: str = None) -> dict:
    """Parse a returned R2/Rn xlsx. Detects round_num from the hidden
    'round' column on the response template.

    Returns ``{round, supplier, items: [{rfq_key, item_num, price, uom,
    status, notes, ...}], n_repriced, n_no_bid, n_blank}``. The caller
    (typically ingest_round2_supplier_bid) uses this to overwrite the
    R1 bids in place.
    """
    if not isinstance(file_bytes, (bytes, bytearray)):
        file_bytes = bytes(file_bytes)
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    # Look for the "Round 2 Response" tab; fall back to the first sheet.
    target_ws = None
    for name in wb.sheetnames:
        if name.lower().startswith("round") and "response" in name.lower():
            target_ws = wb[name]
            break
    if target_ws is None:
        target_ws = wb[wb.sheetnames[0]]

    # Find the header row — it has "Andersen Item #" in the first cell of a row.
    header_row_idx = None
    headers = None
    for r_idx in range(1, 25):
        try:
            row = next(target_ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True), None)
        except Exception:
            break
        if row is None:
            continue
        if row and row[0] and "andersen item" in str(row[0]).lower():
            header_row_idx = r_idx
            headers = [str(c) if c is not None else "" for c in row]
            break
    if not header_row_idx or not headers:
        wb.close()
        return {"error": "Could not locate the response-template header row.",
                "supplier": supplier_name, "round": None}

    # Map header → column index.
    cols = {}
    for i, h in enumerate(headers):
        hl = h.lower().strip()
        if "item_key" in hl or hl == "item_key":
            cols["item_key"] = i
        elif "rfq_line_id" in hl:
            cols["rfq_line_id"] = i
        elif hl == "round":
            cols["round"] = i
        elif "andersen item" in hl:
            cols["item_num"] = i
        elif "quote price" in hl and "echo" not in hl:
            cols["quote_price"] = i
        elif "quote uom" in hl and "echo" not in hl:
            cols["quote_uom"] = i
        elif "your part" in hl and "echo" not in hl:
            cols["your_part"] = i
        elif "lead time" in hl:
            cols["lead_time"] = i
        elif "no bid" in hl and "reason" not in hl:
            cols["no_bid"] = i
        elif "no bid reason" in hl:
            cols["no_bid_reason"] = i
        elif "valid through" in hl:
            cols["valid_through"] = i
        elif hl.endswith("notes") and "echo" not in hl:
            cols["notes"] = i

    items_out = []
    n_repriced = n_no_bid = n_blank = 0
    detected_round = None

    for row in target_ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row is None:
            continue
        # Skip obvious blank rows
        non_blank = sum(1 for v in row if v is not None and str(v).strip())
        if non_blank == 0:
            continue
        item_num_raw = row[cols["item_num"]] if "item_num" in cols and cols["item_num"] < len(row) else None
        if not item_num_raw:
            continue
        item_key_raw = row[cols["item_key"]] if "item_key" in cols and cols["item_key"] < len(row) else None
        rfq_key = norm_pn(item_key_raw or item_num_raw)
        if "round" in cols and cols["round"] < len(row):
            try:
                rnd = int(row[cols["round"]]) if row[cols["round"]] is not None else None
                if rnd is not None and detected_round is None:
                    detected_round = rnd
            except (ValueError, TypeError):
                pass
        price = safe_float(row[cols["quote_price"]]) if "quote_price" in cols and cols["quote_price"] < len(row) else None
        uom = norm_text(row[cols["quote_uom"]]) if "quote_uom" in cols and cols["quote_uom"] < len(row) else ""
        no_bid = row[cols["no_bid"]] if "no_bid" in cols and cols["no_bid"] < len(row) else None
        no_bid_truthy = bool(no_bid) and str(no_bid).strip().lower() in ("y", "yes", "true", "1", "x", "no bid")
        notes = norm_text(row[cols["notes"]]) if "notes" in cols and cols["notes"] < len(row) else ""

        if no_bid_truthy:
            status = BID_STATUS_NO_BID
            n_no_bid += 1
        elif price is not None and price > 0:
            status = BID_STATUS_PRICED
            n_repriced += 1
        else:
            # Blank in R2 = leave R1 alone — don't accidentally erase a good R1 bid
            n_blank += 1
            continue

        items_out.append({
            "rfq_key": rfq_key,
            "item_num": str(item_num_raw),
            "quote_price": price,
            "quote_uom": uom,
            "status": status,
            "your_part": norm_text(row[cols["your_part"]]) if "your_part" in cols and cols["your_part"] < len(row) else "",
            "lead_time_days": safe_float(row[cols["lead_time"]]) if "lead_time" in cols and cols["lead_time"] < len(row) else None,
            "no_bid_reason": norm_text(row[cols["no_bid_reason"]]) if "no_bid_reason" in cols and cols["no_bid_reason"] < len(row) else "",
            "notes": notes,
            "valid_through": norm_text(row[cols["valid_through"]]) if "valid_through" in cols and cols["valid_through"] < len(row) else "",
        })

    wb.close()
    return {
        "supplier": supplier_name,
        "round": detected_round or 2,
        "items": items_out,
        "n_repriced": n_repriced,
        "n_no_bid": n_no_bid,
        "n_blank": n_blank,
    }


def ingest_round2_supplier_bid(file_bytes, supplier_name: str) -> dict:
    """Parse + overwrite the supplier's R1 bids in place for items the R2
    file came back with new prices. Items the supplier left blank in R2
    keep their R1 bids untouched. NO_BID in R2 explicitly overwrites the
    R1 bid with a NO_BID record (analyst can see they actively declined).

    Each item that gets overwritten has the prior bid's price + status
    appended to ``round_history`` on the new bid record, so the
    negotiation arc is auditable per item.

    Args:
        file_bytes: the returned R2 xlsx bytes.
        supplier_name: explicit supplier (since the file itself is the
            same shape regardless of who returned it). Required.

    Returns:
        ``{supplier, round, n_repriced, n_unchanged_r1, n_no_bid_overwrites,
           n_new_items, items: [...]}``
    """
    if not supplier_name:
        return {"error": "supplier_name required"}
    parsed = parse_round2_supplier_bid(file_bytes, supplier_name=supplier_name)
    if "error" in parsed:
        return parsed
    round_num = parsed.get("round") or 2
    items = parsed.get("items", []) or []

    bids_by_supplier = _STATE.setdefault("bids", {})
    if supplier_name not in bids_by_supplier:
        # Treat as a fresh supplier whose first bid happened to be Rn — store
        # the bids list directly.
        bids_by_supplier[supplier_name] = {
            "supplier": supplier_name,
            "bids": [],
            "summary": {"n_lines": 0, "n_priced": 0, "n_no_bid": 0,
                        "n_need_info": 0, "n_uom_disc": 0, "n_substitute": 0,
                        "total_quoted_value": 0.0},
            "format": f"round_{round_num}",
        }
    parsed_state = bids_by_supplier[supplier_name]
    existing = parsed_state.get("bids", []) or []
    existing_by_key = {b.get("rfq_key"): b for b in existing if b.get("rfq_key")}

    n_repriced = 0
    n_no_bid_overwrites = 0
    n_new_items = 0

    items_state = _STATE.get("items", []) or []
    items_by_key = {it.get("key"): it for it in items_state}

    log = _STATE.setdefault("exclusion_log", [])
    ts = datetime.now().isoformat()

    for r2 in items:
        rk = r2.get("rfq_key")
        prior = existing_by_key.get(rk)
        prior_price = prior.get("effective_price") if prior else None
        prior_status = prior.get("status") if prior else None

        item_dict = items_by_key.get(rk) or {}
        qty = item_dict.get("qty_24mo") or 0

        new_price = r2.get("quote_price")
        new_status = r2.get("status")

        # Build the new bid record (mirrors the shape parse_supplier_bid emits).
        history = list((prior or {}).get("round_history") or [])
        if prior:
            history.append({
                "round": (prior.get("round") or 1),
                "price": prior_price,
                "status": prior_status,
                "timestamp": (prior.get("recorded_at") or ""),
            })

        new_bid = {
            "rfq_key": rk,
            "item_num": r2.get("item_num"),
            "part_number": (prior.get("part_number") if prior else "") or "",
            "eam_pn": (prior.get("eam_pn") if prior else "") or "",
            "mfg_name": item_dict.get("mfg_name") or (prior.get("mfg_name") if prior else "") or "",
            "description": item_dict.get("description") or (prior.get("description") if prior else "") or "",
            "commodity": item_dict.get("commodity") or "",
            "qty": qty,
            "uom": r2.get("quote_uom") or (item_dict.get("uom") or ""),
            "quoted_price": new_price,
            "verified_price": None,
            "effective_price": new_price if new_status == BID_STATUS_PRICED else None,
            "notes": r2.get("notes") or "",
            "status": new_status,
            "alt_part": r2.get("your_part") or "",
            "has_substitute": False,
            "round": round_num,
            "round_history": history,
            "recorded_at": ts,
        }

        if prior:
            # Overwrite in place (preserving list order so row-by-row UIs stay stable).
            for i, b in enumerate(existing):
                if b.get("rfq_key") == rk:
                    existing[i] = new_bid
                    break
            if new_status == BID_STATUS_NO_BID:
                n_no_bid_overwrites += 1
            else:
                n_repriced += 1
        else:
            existing.append(new_bid)
            n_new_items += 1

        # Master data-quality log: capture the negotiation outcome.
        log.append({
            "timestamp": ts,
            "app_source": "auto-rfq-banana",
            "event_type": f"round_{round_num}_overwrite",
            "rfq_id": _STATE.get("rfq_id") or "",
            "supplier_name": supplier_name,
            "item_num": r2.get("item_num"),
            "description": item_dict.get("description") or "",
            "mfg_name": item_dict.get("mfg_name") or "",
            "mfg_pn": item_dict.get("mfg_pn") or "",
            "uom": item_dict.get("uom") or "",
            "line_idx": None,
            "line_date": "",
            "line_qty": qty,
            "line_unit_price": new_price,
            "line_total": (new_price or 0) * (qty or 0) if (new_price and qty) else None,
            "line_po": "",
            "line_uom": r2.get("quote_uom") or "",
            "median_before": prior_price,   # the supplier's prior round price (for delta calc)
            "avg_before": None,
            "n_other_lines_before": None,
            "ratio_to_median": (new_price / prior_price) if (prior_price and new_price and prior_price > 0) else None,
            "pct_diff_median": (((new_price / prior_price) - 1.0) * 100.0) if (prior_price and new_price and prior_price > 0) else None,
            "pct_diff_avg": None,
            "notes": f"R{round_num} overwrite: prior {prior_status or 'NEW'} ${prior_price or 0:.2f} → {new_status} ${new_price or 0:.2f}",
        })

    # Refresh aggregate summary on the supplier's parsed-state.
    n_priced = sum(1 for b in existing if b.get("status") == BID_STATUS_PRICED)
    total_value = 0.0
    for b in existing:
        ep = b.get("effective_price")
        q = b.get("qty")
        if ep and q:
            total_value += float(ep) * float(q)
    parsed_state["summary"] = {
        "n_lines": len(existing),
        "n_priced": n_priced,
        "n_no_bid": sum(1 for b in existing if b.get("status") == BID_STATUS_NO_BID),
        "n_need_info": sum(1 for b in existing if b.get("status") == BID_STATUS_NEED_INFO),
        "n_uom_disc": sum(1 for b in existing if b.get("status") == BID_STATUS_UOM_DISC),
        "n_substitute": sum(1 for b in existing if b.get("status") == BID_STATUS_SUBSTITUTE),
        "total_quoted_value": total_value,
    }
    parsed_state["bids"] = existing

    # Bump current_round counter for the dialog to default the next batch.
    _STATE["current_round"] = max(_STATE.get("current_round", 1) or 1, round_num)

    n_unchanged = max(0, len(existing) - n_repriced - n_no_bid_overwrites - n_new_items)
    log_event(
        f"round_{round_num}_bids_loaded",
        f"{supplier_name}: {n_repriced} repriced, {n_no_bid_overwrites} declined, {n_new_items} new, {n_unchanged} unchanged",
        related=supplier_name,
    )

    return {
        "supplier": supplier_name,
        "round": round_num,
        "n_repriced": n_repriced,
        "n_no_bid_overwrites": n_no_bid_overwrites,
        "n_new_items": n_new_items,
        "n_unchanged_r1": n_unchanged,
        "items": items,
    }


# ---------------------------------------------------------------------------
# Decision Summary — the legal-hold narrative companion to the per-supplier
# award letters and the audit / exclusion / data-quality logs.
#
# What this file is for:
#   The user / Ryan asked for a single document that captures, for any RFQ:
#     1. A written prose narrative of what was done and what was decided
#     2. Every threshold + setting active when the decision was made
#     3. Every analyst action (locks, exclusions, UOM resolutions, scenario
#        picks) — quantifies "amount of work done by the user"
#     4. Every system flag (outliers, UOM_DISC, partial coverage, etc.) —
#        quantifies "amount of work the app did showing them what to look at"
#     5. Cost avoidance (vs historical paid prices) AND savings (vs the
#        no-touch auto-recommendation baseline) tracked SEPARATELY so it
#        can feed both KPIs in leadership reporting
#     6. A markable "Items Needing Follow-Up" list — SKUs the analyst
#        manually flagged for double-check after award
#
# Output: an internal-audience xlsx with 7 tabs, banner row 1
# "INTERNAL — NEVER FORWARD". Per-RFQ; designed to be retained for years
# alongside the Decision Log + audit log + exclusion log + award letters.
# ---------------------------------------------------------------------------


def flag_item_for_follow_up(item_num: str, note: str = "") -> dict:
    """Manually mark a SKU for post-award follow-up. Use case: "this bid
    looks suspicious — let's award based on the system rec but verify
    after the first PO arrives." Each flag carries a free-form analyst
    note + timestamp; the Decision Summary's Tab 6 surfaces the list.
    """
    if not item_num:
        raise ValueError("item_num required")
    flags = _STATE.setdefault("follow_up_flags", {})
    flags[item_num] = {
        "item_num": item_num,
        "note": note or "",
        "flagged_at": datetime.now().isoformat(),
        "resolved": False,
        "resolved_at": None,
        "resolved_note": "",
    }
    log_event("flag_follow_up", f"item={item_num} note={note[:80]}")
    return flags[item_num]


def resolve_item_follow_up(item_num: str, note: str = "") -> dict | None:
    """Mark a follow-up flag as resolved (kept in the record, not deleted)."""
    flags = _STATE.get("follow_up_flags", {}) or {}
    if item_num not in flags:
        return None
    flags[item_num]["resolved"] = True
    flags[item_num]["resolved_at"] = datetime.now().isoformat()
    flags[item_num]["resolved_note"] = note or ""
    log_event("resolve_follow_up", f"item={item_num} note={note[:80]}")
    return flags[item_num]


def list_follow_up_flags() -> list:
    flags = _STATE.get("follow_up_flags", {}) or {}
    return list(flags.values())


def compute_decision_summary_metrics(scenario_name: str = None) -> dict:
    """Aggregate every number the Decision Summary needs into one payload.
    Pure function — no state mutation. Returns:

      {
        "rfq_id":                str,
        "supplier_name":         str,    # the source-data supplier (e.g. McMaster)
        "n_items":               int,
        "historical_baseline":   float,  # qty × last_unit_price across all items
        "auto_recommendation": {
          "strategy":              str,
          "supplier_primary":      str,
          "award_total":           float,
          "savings_vs_history":    float,
          "savings_pct":           float,
        },
        "active_award": {
          "strategy":              str,    # the saved/active scenario
          "supplier_primary":      str,
          "award_total":           float,
          "savings_vs_history":    float,  # cost avoidance
          "savings_vs_auto":       float,  # uplift from manual curation
          "savings_pct":           float,
        },
        "analyst_work": {
          "n_locks":               int,
          "n_outlier_exclusions":  int,    # items with ≥1 excluded line
          "n_excluded_lines":      int,    # total individual lines excluded
          "n_uom_resolutions":     int,
          "n_scenarios_saved":     int,
          "n_round2_selected":     int,
          "n_follow_up_flags":     int,
        },
        "system_work": {
          "n_recommendations":     dict,   # {ACCEPT, PUSH_BACK, ASK_CLARIFICATION, EXCLUDE, MANUAL_REVIEW}
          "n_outliers_flagged":    int,
          "n_uom_disc_flagged":    int,
          "n_substitutes_flagged": int,
          "n_partial_coverage":    int,    # items with <full coverage
          "n_difficulty_signals":  int,
        },
        "thresholds":              dict,   # all 11 active thresholds
        "scenarios":               list,   # all saved scenarios with totals
        "follow_up_flags":         list,
        "audit_events":            list,   # full audit_log
        "exclusion_log_summary":   dict,
      }
    """
    items = _STATE.get("items", []) or []
    bids_by_sup = _STATE.get("bids", {}) or {}
    item_locks = _STATE.get("item_locks", {}) or {}
    item_exclusions = _STATE.get("item_exclusions", {}) or {}
    uom_annotations = _STATE.get("uom_annotations", {}) or {}
    scenarios = _STATE.get("scenarios", {}) or {}
    follow_ups = _STATE.get("follow_up_flags", {}) or {}
    round2_sel = _STATE.get("round2_selection", []) or []
    audit_log = _STATE.get("audit_log", []) or []

    historical_baseline = sum(
        (it.get("last_unit_price") or 0) * (it.get("qty_24mo") or 0)
        for it in items
    )

    # The auto-recommendation = lowest_qualified with NO manual overrides.
    # We compute it by temporarily swapping out item_locks / item_exclusions
    # and running _evaluate_scenario, then restoring. Cleaner approach: the
    # current state's `lowest_qualified` IS the auto rec when no manual
    # changes are present. When the analyst HAS made manual changes, the
    # "auto" baseline is what lowest_qualified WOULD produce if we cleared
    # them. For now, we capture the current lowest_qualified summary as
    # the "current-state auto" — leadership reporting wants this number to
    # reflect the bid landscape at decision time (including any UOM
    # resolutions the analyst has confirmed), not a hypothetical
    # zero-curation baseline.
    auto_eval = _evaluate_scenario("lowest_qualified", {}, {})
    auto_summary = _summarize_eval_for_headline(auto_eval)

    # Active award: if a scenario name is provided, evaluate that. Else, fall
    # back to the most-recent saved scenario, else the current consolidate_to.
    active_eval = None
    active_strategy = None
    if scenario_name and scenario_name in scenarios:
        sc = scenarios[scenario_name]
        active_eval = _evaluate_scenario(
            sc["strategy"], sc.get("parameters") or {},
            sc.get("overrides") or {}, sc.get("included_keys"),
        )
        active_strategy = sc["strategy"]
    elif scenarios:
        most_recent = sorted(scenarios.values(), key=lambda s: s.get("saved_at") or "", reverse=True)[0]
        active_eval = _evaluate_scenario(
            most_recent["strategy"], most_recent.get("parameters") or {},
            most_recent.get("overrides") or {}, most_recent.get("included_keys"),
        )
        active_strategy = most_recent["strategy"]
        scenario_name = most_recent.get("name")
    else:
        # No saved scenario yet — use current consolidate_to with default supplier
        consol = compute_consolidation_analysis()
        cands = consol.get("candidates") or []
        target = cands[0]["supplier"] if cands else None
        if target:
            active_eval = _evaluate_scenario("consolidate_to", {"supplier": target}, {})
            active_strategy = "consolidate_to"
        else:
            active_eval = auto_eval
            active_strategy = "lowest_qualified"

    active_summary = _summarize_eval_for_headline(active_eval)

    # Counts of system-flagged work
    matrix = compute_comparison_matrix() or {}
    rec_counts = {"ACCEPT": 0, "PUSH_BACK": 0, "ASK_CLARIFICATION": 0, "EXCLUDE": 0, "MANUAL_REVIEW": 0}
    n_outliers = 0
    n_uom_disc = 0
    n_subs = 0
    n_partial = 0
    for r in (matrix.get("rows") or []):
        # recommendation is a string label (ACCEPT / PUSH_BACK / ...)
        # When the engine returns a richer object, it carries `recommendation`
        # as a top-level string + `recommendation_reason` as the reason text.
        rec = r.get("recommendation") or ""
        if isinstance(rec, dict):
            rec = rec.get("level") or rec.get("recommendation") or ""
        if rec in rec_counts:
            rec_counts[rec] += 1
        flags = r.get("flags") or []
        if any("OUTLIER" in f or "BIG_SAVINGS" in f or "ALL_BIDS_HIGH" in f for f in flags):
            n_outliers += 1
        # supplier_bids may be a list or a dict — normalize
        sbids = r.get("supplier_bids") or {}
        if isinstance(sbids, list):
            sbids_iter = sbids
        else:
            sbids_iter = list(sbids.values())
        n_priced = sum(1 for s in sbids_iter if s and s.get("status") == BID_STATUS_PRICED)
        if n_priced < len(bids_by_sup):
            n_partial += 1
    for sup, parsed in bids_by_sup.items():
        for b in parsed.get("bids", []) or []:
            st = b.get("status")
            if st == BID_STATUS_UOM_DISC:
                n_uom_disc += 1
            elif st == BID_STATUS_SUBSTITUTE:
                n_subs += 1

    diff = _STATE.get("difficulty") or {}
    n_diff_signals = len((diff.get("signals") or []))

    n_excluded_lines = sum(len(v) for v in item_exclusions.values() if v)
    n_excluded_items = sum(1 for v in item_exclusions.values() if v)

    return {
        "rfq_id": _STATE.get("rfq_id") or "",
        "supplier_name": _STATE.get("supplier_name") or "",
        "n_items": len(items),
        "historical_baseline": historical_baseline,
        "auto_recommendation": {
            "strategy": "lowest_qualified",
            "supplier_primary": auto_summary["supplier_primary"],
            "award_total": auto_summary["award_total"],
            "savings_vs_history": auto_summary["savings_total"],
            "savings_pct": auto_summary["savings_pct"],
        },
        "active_award": {
            "scenario_name": scenario_name,
            "strategy": active_strategy,
            "supplier_primary": active_summary["supplier_primary"],
            "award_total": active_summary["award_total"],
            "savings_vs_history": active_summary["savings_total"],
            "savings_vs_auto": (auto_summary["award_total"] - active_summary["award_total"]),
            "savings_pct": active_summary["savings_pct"],
            "n_carved": active_summary["n_carved"],
        },
        "analyst_work": {
            "n_locks": len(item_locks),
            "n_outlier_exclusions": n_excluded_items,
            "n_excluded_lines": n_excluded_lines,
            "n_uom_resolutions": len([k for k, v in uom_annotations.items() if v and v.get("factor") is not None]),
            "n_scenarios_saved": len(scenarios),
            "n_round2_selected": len(round2_sel),
            "n_follow_up_flags": len(follow_ups),
            "n_follow_up_unresolved": len([f for f in follow_ups.values() if not f.get("resolved")]),
        },
        "system_work": {
            "n_recommendations": rec_counts,
            "n_outliers_flagged": n_outliers,
            "n_uom_disc_flagged": n_uom_disc,
            "n_substitutes_flagged": n_subs,
            "n_partial_coverage": n_partial,
            "n_difficulty_signals": n_diff_signals,
        },
        "thresholds": get_thresholds(),
        "scenarios": list_award_scenarios(),
        "follow_up_flags": list(follow_ups.values()),
        "audit_events": audit_log,
        "exclusion_log_summary": get_exclusion_log_summary(),
    }


def _build_decision_narrative(metrics: dict) -> str:
    """Generate the prose summary text. Reads from the metrics dict; pure
    string formatting. Multi-paragraph, ~10 sentences, captures both
    cost avoidance vs savings + the volume of work done by both sides.
    """
    m = metrics
    aw = m["analyst_work"]
    sw = m["system_work"]
    auto = m["auto_recommendation"]
    active = m["active_award"]
    n_items = m["n_items"]
    hist = m["historical_baseline"]

    # Helpers
    def _money(v):
        if v is None:
            return "$0"
        sign = "−$" if v < 0 else "$"
        return sign + f"{abs(v):,.0f}"
    def _pct(v):
        if v is None:
            return "0%"
        return f"{v:+.1f}%"

    n_suppliers = len(_STATE.get("bids", {}) or {})
    n_priced_total = 0
    for sup_data in (_STATE.get("bids", {}) or {}).values():
        n_priced_total += sum(1 for b in (sup_data.get("bids") or []) if b.get("status") == BID_STATUS_PRICED)

    p1 = (
        f"This RFQ analysis covered {n_items:,} items with a historical baseline of "
        f"{_money(hist)} (Σ qty_24mo × last-paid price). "
        f"{n_suppliers} supplier(s) responded with {n_priced_total:,} priced bids in total."
    )

    rec_counts = sw["n_recommendations"]
    p2 = (
        f"The system surfaced {rec_counts.get('ACCEPT', 0):,} ACCEPT recommendations, "
        f"{rec_counts.get('PUSH_BACK', 0):,} PUSH_BACK candidates, "
        f"{rec_counts.get('ASK_CLARIFICATION', 0):,} clarification asks, "
        f"{rec_counts.get('EXCLUDE', 0):,} excludes, and "
        f"{rec_counts.get('MANUAL_REVIEW', 0):,} items needing manual review. "
        f"It flagged {sw['n_outliers_flagged']:,} outlier-priced items, "
        f"{sw['n_uom_disc_flagged']:,} UOM-discrepancy bids, "
        f"{sw['n_substitutes_flagged']:,} substitute-part offers, and "
        f"{sw['n_partial_coverage']:,} items with partial supplier coverage."
    )

    if any([aw['n_locks'], aw['n_outlier_exclusions'], aw['n_uom_resolutions'],
            aw['n_round2_selected'], aw['n_follow_up_flags'], aw['n_scenarios_saved']]):
        analyst_actions = []
        if aw['n_outlier_exclusions']:
            analyst_actions.append(
                f"excluded {aw['n_excluded_lines']:,} order line(s) across "
                f"{aw['n_outlier_exclusions']:,} item(s) as confirmed outliers"
            )
        if aw['n_locks']:
            analyst_actions.append(f"locked {aw['n_locks']:,} item(s) to specific suppliers after audit")
        if aw['n_uom_resolutions']:
            analyst_actions.append(f"resolved {aw['n_uom_resolutions']:,} UOM mismatch(es)")
        if aw['n_round2_selected']:
            analyst_actions.append(f"selected {aw['n_round2_selected']:,} item(s) for a Round 2 push-back")
        if aw['n_follow_up_flags']:
            analyst_actions.append(f"flagged {aw['n_follow_up_flags']:,} SKU(s) for post-award follow-up")
        if aw['n_scenarios_saved']:
            analyst_actions.append(f"saved {aw['n_scenarios_saved']:,} award scenario(s)")
        p3 = "The analyst's manual curation: " + "; ".join(analyst_actions) + "."
    else:
        p3 = "The analyst applied no manual curation — the award reflects the system's auto-recommendation as-is."

    p4 = (
        f"Award strategy chosen: {active['strategy']}. "
        f"Primary award supplier: {active['supplier_primary'] or '(none)'}."
    )
    if active.get("n_carved"):
        p4 += f" {active['n_carved']:,} item(s) were carved out to other suppliers per the dual-threshold rule "
        p4 += f"(≥{int(round(m['thresholds'].get('carve_out_min_savings_pct', 0)*100))}% savings OR "
        p4 += f"≥${m['thresholds'].get('carve_out_min_savings_annual_dollar', 0):,.0f}/yr)."

    # Cost avoidance vs savings — the two reporting numbers
    cost_avoid = active['savings_vs_history']
    uplift = active['savings_vs_auto']
    p5 = (
        f"COST AVOIDANCE (active award vs historical paid baseline): {_money(cost_avoid)} "
        f"({_pct(active['savings_pct'])}). "
    )
    if abs(uplift) >= 1.0:
        sign = "saved an additional" if uplift > 0 else "cost an additional"
        p5 += (
            f"SAVINGS UPLIFT FROM MANUAL CURATION (active vs no-touch auto baseline of "
            f"{_money(auto['award_total'])}): {sign} {_money(abs(uplift))} versus simply "
            f"taking the lowest-qualified bid as-is."
        )
    else:
        p5 += (
            f"The active award matches the no-touch auto baseline within rounding — "
            f"manual curation did not change the headline number, but the curation work "
            f"is recorded for audit and may surface follow-up issues post-PO."
        )

    return "\n\n".join([p1, p2, p3, p4, p5])


def gen_decision_summary_xlsx(scenario_name: str = None, rfq_id: str = "") -> bytes:
    """Build the multi-tab Decision Summary xlsx. Internal-audience banner.
    Tabs:
      1. Executive Summary    (prose narrative + headline numbers)
      2. Settings & Thresholds (every threshold + scenario params snapshot)
      3. Analyst Actions       (locks, exclusions, UOM, scenarios, follow-ups)
      4. System Flags          (recommendations + outliers + UOM_DISC + ...)
      5. Cost Avoidance vs Savings (the two-number reporting table)
      6. Items Needing Follow-Up (markable list — flagged SKUs)
      7. Decision Log Timeline (audit_log unfiltered)

    Always includes "INTERNAL — NEVER FORWARD" banner row 1.
    """
    metrics = compute_decision_summary_metrics(scenario_name)
    if rfq_id:
        metrics["rfq_id"] = rfq_id
    narrative = _build_decision_narrative(metrics)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    BANNER_FILL = PatternFill("solid", fgColor="FFF1B341")
    BANNER_FONT = Font(name="Consolas", size=11, bold=True, color="000000")
    H = Font(name="Calibri", size=11, bold=True)
    MONO = Font(name="Consolas", size=10)
    THIN = Side(border_style="thin", color="999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    BANNER_TXT = "INTERNAL — NEVER FORWARD — RETAIN PER LEGAL HOLD"

    def _banner(ws, ncols=8):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=BANNER_TXT)
        c.fill = BANNER_FILL
        c.font = BANNER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    # ==== TAB 1: Executive Summary ====
    ws = wb.active
    ws.title = "1_Executive_Summary"
    _banner(ws, ncols=4)
    ws.cell(row=2, column=1, value="DECISION SUMMARY").font = Font(name="Calibri", size=18, bold=True)
    ws.cell(row=3, column=1, value=f"RFQ: {metrics['rfq_id'] or '(unset)'}    Source supplier: {metrics['supplier_name'] or '(unset)'}    Generated: {datetime.now().isoformat(timespec='seconds')}").font = MONO
    ws.row_dimensions[2].height = 28
    ws.cell(row=5, column=1, value="HEADLINE").font = H
    rows = [
        ("Items in RFQ", f"{metrics['n_items']:,}"),
        ("Historical baseline (qty × last-paid)", f"${metrics['historical_baseline']:,.2f}"),
        ("Award strategy", metrics['active_award']['strategy']),
        ("Award supplier (primary)", metrics['active_award']['supplier_primary'] or "(none)"),
        ("Award total", f"${metrics['active_award']['award_total']:,.2f}"),
        ("Cost avoidance vs historical", f"${metrics['active_award']['savings_vs_history']:,.2f}  ({metrics['active_award']['savings_pct']:+.1f}%)"),
        ("Savings vs no-touch auto baseline", f"${metrics['active_award']['savings_vs_auto']:,.2f}"),
        ("Carve-outs in active award", f"{metrics['active_award']['n_carved']:,}"),
    ]
    for i, (k, v) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=k).font = MONO
        ws.cell(row=i, column=2, value=v).font = MONO
    ws.cell(row=15, column=1, value="NARRATIVE").font = H
    # Narrative split into rows by paragraph for readability
    paras = narrative.split("\n\n")
    cur = 16
    for p in paras:
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=4)
        c = ws.cell(row=cur, column=1, value=p)
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[cur].height = max(60, 18 * (1 + len(p)//90))
        cur += 1
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25

    # ==== TAB 2: Settings & Thresholds ====
    ws2 = wb.create_sheet("2_Settings_Thresholds")
    _banner(ws2, ncols=3)
    ws2.cell(row=2, column=1, value="ENGINE THRESHOLDS ACTIVE AT DECISION TIME").font = H
    ws2.cell(row=3, column=1, value="Threshold").font = H
    ws2.cell(row=3, column=2, value="Value").font = H
    ws2.cell(row=3, column=3, value="Default").font = H
    th = metrics["thresholds"]
    r = 4
    for k in sorted(th.keys()):
        ws2.cell(row=r, column=1, value=k).font = MONO
        ws2.cell(row=r, column=2, value=th[k]).font = MONO
        ws2.cell(row=r, column=3, value=DEFAULT_THRESHOLDS.get(k)).font = MONO
        r += 1
    r += 2
    ws2.cell(row=r, column=1, value="SAVED SCENARIOS").font = H; r += 1
    ws2.cell(row=r, column=1, value="Name").font = H
    ws2.cell(row=r, column=2, value="Strategy").font = H
    ws2.cell(row=r, column=3, value="Parameters").font = H
    r += 1
    for s in metrics["scenarios"]:
        ws2.cell(row=r, column=1, value=s.get("name")).font = MONO
        ws2.cell(row=r, column=2, value=s.get("strategy")).font = MONO
        ws2.cell(row=r, column=3, value=str(s.get("parameters") or {})).font = MONO
        r += 1
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 60

    # ==== TAB 3: Analyst Actions ====
    ws3 = wb.create_sheet("3_Analyst_Actions")
    _banner(ws3, ncols=5)
    ws3.cell(row=2, column=1, value="ANALYST WORK PERFORMED — quantifies manual curation effort").font = H
    aw = metrics["analyst_work"]
    counts = [
        ("Item locks", aw["n_locks"]),
        ("Items with outlier exclusions", aw["n_outlier_exclusions"]),
        ("Total order lines excluded", aw["n_excluded_lines"]),
        ("UOM resolutions applied", aw["n_uom_resolutions"]),
        ("Award scenarios saved", aw["n_scenarios_saved"]),
        ("Round 2 / Rn items selected", aw["n_round2_selected"]),
        ("Follow-up flags placed", aw["n_follow_up_flags"]),
        ("Follow-up flags unresolved", aw["n_follow_up_unresolved"]),
    ]
    for i, (k, v) in enumerate(counts, start=4):
        ws3.cell(row=i, column=1, value=k).font = MONO
        ws3.cell(row=i, column=2, value=v).font = MONO
    r = 4 + len(counts) + 2
    # Lock detail
    ws3.cell(row=r, column=1, value="ITEM LOCKS").font = H; r += 1
    locks = _STATE.get("item_locks", {}) or {}
    if locks:
        for h_idx, h_lbl in enumerate(["Item", "Locked supplier", "Reason", "Locked at"], start=1):
            ws3.cell(row=r, column=h_idx, value=h_lbl).font = H
        r += 1
        for it_num, lk in locks.items():
            ws3.cell(row=r, column=1, value=it_num).font = MONO
            ws3.cell(row=r, column=2, value=lk.get("supplier")).font = MONO
            ws3.cell(row=r, column=3, value=lk.get("reason")).font = MONO
            ws3.cell(row=r, column=4, value=lk.get("locked_at")).font = MONO
            r += 1
    else:
        ws3.cell(row=r, column=1, value="(no locks)").font = MONO; r += 1
    r += 2
    # Exclusion detail
    ws3.cell(row=r, column=1, value="OUTLIER EXCLUSIONS").font = H; r += 1
    exc = _STATE.get("item_exclusions", {}) or {}
    nonempty = {k: v for k, v in exc.items() if v}
    if nonempty:
        for h_idx, h_lbl in enumerate(["Item", "Excluded line indices", "N excluded"], start=1):
            ws3.cell(row=r, column=h_idx, value=h_lbl).font = H
        r += 1
        for it_num, idxs in nonempty.items():
            ws3.cell(row=r, column=1, value=it_num).font = MONO
            ws3.cell(row=r, column=2, value=", ".join(str(i) for i in idxs)).font = MONO
            ws3.cell(row=r, column=3, value=len(idxs)).font = MONO
            r += 1
    else:
        ws3.cell(row=r, column=1, value="(no outlier exclusions)").font = MONO; r += 1
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 36
    ws3.column_dimensions["D"].width = 22

    # ==== TAB 4: System Flags ====
    ws4 = wb.create_sheet("4_System_Flags")
    _banner(ws4, ncols=3)
    ws4.cell(row=2, column=1, value="SYSTEM-FLAGGED WORK — quantifies what the engine surfaced").font = H
    sw = metrics["system_work"]
    rc = sw["n_recommendations"]
    rows4 = [
        ("Recommendation: ACCEPT", rc.get("ACCEPT", 0)),
        ("Recommendation: PUSH_BACK", rc.get("PUSH_BACK", 0)),
        ("Recommendation: ASK_CLARIFICATION", rc.get("ASK_CLARIFICATION", 0)),
        ("Recommendation: EXCLUDE", rc.get("EXCLUDE", 0)),
        ("Recommendation: MANUAL_REVIEW", rc.get("MANUAL_REVIEW", 0)),
        ("Outlier-priced items flagged", sw["n_outliers_flagged"]),
        ("UOM_DISC bids flagged", sw["n_uom_disc_flagged"]),
        ("Substitute-part bids flagged", sw["n_substitutes_flagged"]),
        ("Items with partial coverage", sw["n_partial_coverage"]),
        ("Difficulty signals on the source data", sw["n_difficulty_signals"]),
    ]
    for i, (k, v) in enumerate(rows4, start=4):
        ws4.cell(row=i, column=1, value=k).font = MONO
        ws4.cell(row=i, column=2, value=v).font = MONO
    ws4.column_dimensions["A"].width = 44
    ws4.column_dimensions["B"].width = 16

    # ==== TAB 5: Cost Avoidance vs Savings ====
    ws5 = wb.create_sheet("5_CostAvoid_vs_Savings")
    _banner(ws5, ncols=4)
    ws5.cell(row=2, column=1, value="THE TWO REPORTING NUMBERS — cost avoidance and savings tracked separately").font = H
    ws5.cell(row=3, column=1, value="Cost Avoidance (CA) = historical_baseline − active_award_total. Used for leadership reporting against budgeted spend.").font = MONO
    ws5.cell(row=4, column=1, value="Savings (S) = no-touch_auto_award_total − active_award_total. Quantifies the lift from manual curation.").font = MONO
    headers5 = ["Metric", "Auto recommendation", "Active award", "Delta"]
    for i, h in enumerate(headers5, start=1):
        ws5.cell(row=6, column=i, value=h).font = H
    auto = metrics["auto_recommendation"]; active = metrics["active_award"]
    table5 = [
        ("Strategy", auto["strategy"], active["strategy"], "—"),
        ("Primary supplier", auto["supplier_primary"], active["supplier_primary"], "—"),
        ("Award total", auto["award_total"], active["award_total"], active["award_total"] - auto["award_total"]),
        ("Cost avoidance vs historical", auto["savings_vs_history"], active["savings_vs_history"], active["savings_vs_history"] - auto["savings_vs_history"]),
        ("Savings vs auto baseline", 0.0, active["savings_vs_auto"], active["savings_vs_auto"]),
    ]
    for i, row in enumerate(table5, start=7):
        for j, v in enumerate(row, start=1):
            c = ws5.cell(row=i, column=j, value=v)
            c.font = MONO
            if isinstance(v, (int, float)) and j > 1:
                c.number_format = "$#,##0.00"
    ws5.column_dimensions["A"].width = 36
    for col in ("B", "C", "D"):
        ws5.column_dimensions[col].width = 24

    # ==== TAB 6: Items Needing Follow-Up ====
    ws6 = wb.create_sheet("6_FollowUp_Items")
    _banner(ws6, ncols=6)
    ws6.cell(row=2, column=1, value="MARKED FOR FOLLOW-UP — analyst-flagged SKUs to double-check post-award").font = H
    headers6 = ["Item", "Note", "Flagged at", "Resolved", "Resolved at", "Resolution note"]
    for i, h in enumerate(headers6, start=1):
        ws6.cell(row=3, column=i, value=h).font = H
    flags = metrics["follow_up_flags"]
    if flags:
        for i, f in enumerate(flags, start=4):
            ws6.cell(row=i, column=1, value=f.get("item_num")).font = MONO
            ws6.cell(row=i, column=2, value=f.get("note")).font = MONO
            ws6.cell(row=i, column=3, value=f.get("flagged_at")).font = MONO
            ws6.cell(row=i, column=4, value="YES" if f.get("resolved") else "no").font = MONO
            ws6.cell(row=i, column=5, value=f.get("resolved_at")).font = MONO
            ws6.cell(row=i, column=6, value=f.get("resolved_note")).font = MONO
    else:
        ws6.cell(row=4, column=1, value="(no items flagged for follow-up)").font = MONO
    ws6.column_dimensions["A"].width = 18
    ws6.column_dimensions["B"].width = 50
    ws6.column_dimensions["C"].width = 22
    ws6.column_dimensions["D"].width = 12
    ws6.column_dimensions["E"].width = 22
    ws6.column_dimensions["F"].width = 50

    # ==== TAB 7: Decision Log Timeline ====
    ws7 = wb.create_sheet("7_Decision_Log_Timeline")
    _banner(ws7, ncols=4)
    ws7.cell(row=2, column=1, value="EVERY DISCRETE EVENT — full audit trail").font = H
    headers7 = ["Timestamp", "Action", "Detail", "Related"]
    for i, h in enumerate(headers7, start=1):
        ws7.cell(row=3, column=i, value=h).font = H
    events = metrics["audit_events"]
    if events:
        for i, ev in enumerate(events, start=4):
            ws7.cell(row=i, column=1, value=ev.get("timestamp")).font = MONO
            ws7.cell(row=i, column=2, value=ev.get("action")).font = MONO
            ws7.cell(row=i, column=3, value=ev.get("detail")).font = MONO
            ws7.cell(row=i, column=4, value=ev.get("related")).font = MONO
    else:
        ws7.cell(row=4, column=1, value="(no audit events)").font = MONO
    ws7.column_dimensions["A"].width = 22
    ws7.column_dimensions["B"].width = 28
    ws7.column_dimensions["C"].width = 80
    ws7.column_dimensions["D"].width = 30

    log_event(
        "gen_decision_summary_xlsx",
        f"strategy={metrics['active_award']['strategy']} CA=${metrics['active_award']['savings_vs_history']:,.0f} uplift=${metrics['active_award']['savings_vs_auto']:,.0f}",
        related=scenario_name,
    )

    import io
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Register module so app.js can `from app_engine import ...`
# ---------------------------------------------------------------------------
sys.modules["app_engine"] = sys.modules[__name__]
