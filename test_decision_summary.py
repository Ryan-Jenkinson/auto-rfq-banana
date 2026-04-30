"""
Self-test for compute_decision_summary_metrics + gen_decision_summary_xlsx.

Builds the same synthetic 6-item / 3-supplier RFQ as test_headline_strategies,
then asserts:
  - cost avoidance + savings tracked separately and the math is consistent
  - manual-work counters reflect every action class (locks, exclusions, UOM,
    follow-ups, scenarios)
  - system-work counters surface flags from the comparison matrix
  - flag_item_for_follow_up appends to the markable list and resolve marks
    it without removing
  - gen_decision_summary_xlsx returns bytes that openpyxl can re-load and
    contains the 7 expected tabs + the INTERNAL banner
"""
import sys, io
sys.path.insert(0, ".")
import app
from app import (
    _STATE, BID_STATUS_PRICED, compute_decision_summary_metrics,
    gen_decision_summary_xlsx, _build_decision_narrative,
    flag_item_for_follow_up, resolve_item_follow_up, list_follow_up_flags,
    save_award_scenario, set_item_lock, set_uom_annotation, set_thresholds,
)
from openpyxl import load_workbook


def _reset_state():
    _STATE["items"] = []
    _STATE["bids"] = {}
    _STATE["item_locks"] = {}
    _STATE["item_exclusions"] = {}
    _STATE["uom_annotations"] = {}
    _STATE["scenarios"] = {}
    _STATE["audit_log"] = []
    _STATE["thresholds"] = {}
    _STATE["follow_up_flags"] = {}
    _STATE["round2_selection"] = []
    _STATE["supplier_name"] = "INCUMBENT"
    _STATE["rfq_id"] = ""


def _mk(key, num, desc, qty, price):
    return {
        "key": key, "item_num": num, "description": desc,
        "qty_24mo": qty, "qty_12mo": qty/2, "qty_36mo": qty,
        "spend_24mo": price*qty, "spend_12mo": price*qty/2, "spend_36mo": price*qty,
        "spend_24mo_actual": price*qty, "spend_12mo_actual": price*qty/2, "spend_36mo_actual": price*qty,
        "last_unit_price": price, "uom": "EA", "uom_mixed": False,
        "mfg_name": "", "mfg_pn": "", "commodity": "MRO",
        "po_lines": [], "supplier": "INCUMBENT",
        "first_order": "2024-01-01", "last_order": "2026-04-01", "po_count": 1,
        "score": 100, "tier": "STRONG",
    }


def _seed():
    _STATE["items"] = [
        _mk("ITEM_A", "A", "RED dominates", 50000, 12.0),
        _mk("ITEM_B", "B", "PCT carve",     20,    11.0),
        _mk("ITEM_C", "C", "DOLLAR carve",  20000, 10.0),
        _mk("ITEM_D", "D", "BOTH carve",    15000, 10.0),
        _mk("ITEM_E", "E", "no carve",      200,   10.0),
        _mk("ITEM_F", "F", "RED no-quote",  100,   5.0),
    ]
    _STATE["bids"] = {
        "RED": {"bids": [
            {"rfq_key": "ITEM_A", "status": BID_STATUS_PRICED, "effective_price": 2.0,  "notes": ""},
            {"rfq_key": "ITEM_B", "status": BID_STATUS_PRICED, "effective_price": 10.0, "notes": ""},
            {"rfq_key": "ITEM_C", "status": BID_STATUS_PRICED, "effective_price": 10.0, "notes": ""},
            {"rfq_key": "ITEM_D", "status": BID_STATUS_PRICED, "effective_price": 10.0, "notes": ""},
            {"rfq_key": "ITEM_E", "status": BID_STATUS_PRICED, "effective_price": 10.0, "notes": ""},
        ]},
        "BLUE": {"bids": [
            {"rfq_key": "ITEM_A", "status": BID_STATUS_PRICED, "effective_price": 10.0, "notes": ""},
            {"rfq_key": "ITEM_B", "status": BID_STATUS_PRICED, "effective_price": 7.00, "notes": ""},
            {"rfq_key": "ITEM_C", "status": BID_STATUS_PRICED, "effective_price": 9.50, "notes": ""},
            {"rfq_key": "ITEM_D", "status": BID_STATUS_PRICED, "effective_price": 7.00, "notes": ""},
            {"rfq_key": "ITEM_E", "status": BID_STATUS_PRICED, "effective_price": 9.80, "notes": ""},
            {"rfq_key": "ITEM_F", "status": BID_STATUS_PRICED, "effective_price": 5.00, "notes": ""},
        ]},
        "GREEN": {"bids": [
            {"rfq_key": "ITEM_F", "status": BID_STATUS_PRICED, "effective_price": 4.80, "notes": ""},
        ]},
    }


def test_metrics_clean_baseline():
    """No manual work — analyst_work counters all zero, system_work surfaces
    the recommendation distribution + outlier counts."""
    _reset_state(); _seed()
    m = compute_decision_summary_metrics()
    assert m["n_items"] == 6
    aw = m["analyst_work"]
    assert aw["n_locks"] == 0 and aw["n_outlier_exclusions"] == 0
    assert aw["n_excluded_lines"] == 0 and aw["n_uom_resolutions"] == 0
    assert aw["n_follow_up_flags"] == 0 and aw["n_scenarios_saved"] == 0
    sw = m["system_work"]
    rc = sw["n_recommendations"]
    # rec_counts must sum to n_items (each item gets exactly one recommendation)
    assert sum(rc.values()) == 6, f"rec counts sum to {sum(rc.values())}, expected 6"
    print(f"PASS  metrics_clean_baseline  (rec_counts={rc})")


def test_metrics_after_manual_work():
    _reset_state(); _seed()
    set_item_lock("A", supplier="RED", reason="audit-confirmed")
    _STATE["item_exclusions"]["B"] = [0]   # no real po_lines, but counter logic still works
    set_uom_annotation("ITEM_C", "BLUE", factor=2.0, hist_uom="EA", bid_uom="BX")
    flag_item_for_follow_up("D", "verify with site lead")
    save_award_scenario("April Q1", "consolidate_to", {"supplier": "RED"})

    m = compute_decision_summary_metrics()
    aw = m["analyst_work"]
    assert aw["n_locks"] == 1, f"n_locks={aw['n_locks']}"
    assert aw["n_outlier_exclusions"] == 1
    assert aw["n_excluded_lines"] == 1
    assert aw["n_uom_resolutions"] == 1
    assert aw["n_follow_up_flags"] == 1
    assert aw["n_follow_up_unresolved"] == 1
    assert aw["n_scenarios_saved"] == 1
    print(f"PASS  metrics_after_manual_work  ({aw})")


def test_cost_avoidance_vs_savings():
    """The two reporting numbers must be tracked separately. Without manual
    overrides, the active award (consolidate_to default supplier) and the
    auto recommendation (lowest_qualified) can BOTH have a cost-avoidance
    number — but the savings-vs-auto number is the DELTA between them."""
    _reset_state(); _seed()
    set_thresholds({"carve_out_min_savings_pct": 0.20, "carve_out_min_savings_annual_dollar": 3000.0})
    m = compute_decision_summary_metrics()
    auto = m["auto_recommendation"]
    active = m["active_award"]
    # cost avoidance vs history must equal historical_baseline − award_total
    assert abs(auto["savings_vs_history"] - (m["historical_baseline"] - auto["award_total"])) < 0.01
    assert abs(active["savings_vs_history"] - (m["historical_baseline"] - active["award_total"])) < 0.01
    # savings_vs_auto is the difference of award totals
    assert abs(active["savings_vs_auto"] - (auto["award_total"] - active["award_total"])) < 0.01
    print(f"PASS  cost_avoidance_vs_savings  (auto_award=${auto['award_total']:,.0f}, active_award=${active['award_total']:,.0f}, CA=${active['savings_vs_history']:,.0f}, uplift=${active['savings_vs_auto']:,.0f})")


def test_follow_up_lifecycle():
    """Flag → list shows it → resolve → still in list, marked resolved."""
    _reset_state(); _seed()
    flag_item_for_follow_up("A", "weird $2 quote — verify")
    flag_item_for_follow_up("D", "BOTH carve — confirm with engineering")
    flags = list_follow_up_flags()
    assert len(flags) == 2
    assert all(not f["resolved"] for f in flags)

    resolve_item_follow_up("A", "verified — RED was running a clearance bin")
    flags2 = list_follow_up_flags()
    assert len(flags2) == 2  # not deleted
    a = next(f for f in flags2 if f["item_num"] == "A")
    d = next(f for f in flags2 if f["item_num"] == "D")
    assert a["resolved"] is True and a["resolved_note"]
    assert d["resolved"] is False
    print("PASS  follow_up_lifecycle")


def test_narrative_text_present():
    _reset_state(); _seed()
    set_item_lock("A", supplier="RED", reason="t1")
    flag_item_for_follow_up("B", "verify post-PO")
    m = compute_decision_summary_metrics()
    narrative = _build_decision_narrative(m)
    assert "RFQ analysis covered" in narrative
    assert "supplier(s) responded" in narrative
    assert "manual curation" in narrative or "no manual curation" in narrative
    assert "COST AVOIDANCE" in narrative
    print("PASS  narrative_text_present")


def test_xlsx_round_trips():
    _reset_state(); _seed()
    flag_item_for_follow_up("A", "verify")
    set_item_lock("B", supplier="RED", reason="audit")
    save_award_scenario("April Q1", "consolidate_to", {"supplier": "RED"})

    raw = gen_decision_summary_xlsx(scenario_name="April Q1", rfq_id="RFQ-2026-04")
    assert isinstance(raw, (bytes, bytearray))
    assert len(raw) > 4000

    wb = load_workbook(io.BytesIO(raw))
    expected_tabs = [
        "1_Executive_Summary",
        "2_Settings_Thresholds",
        "3_Analyst_Actions",
        "4_System_Flags",
        "5_CostAvoid_vs_Savings",
        "6_FollowUp_Items",
        "7_Decision_Log_Timeline",
    ]
    assert wb.sheetnames == expected_tabs, f"got {wb.sheetnames}"

    # Banner check on every tab
    for name in expected_tabs:
        ws = wb[name]
        a1 = str(ws.cell(row=1, column=1).value or "")
        assert "INTERNAL" in a1 and "NEVER FORWARD" in a1, f"tab {name} missing banner: {a1!r}"

    # Tab 1 must have RFQ id + the narrative
    ws1 = wb["1_Executive_Summary"]
    found_rfq = False
    for row in ws1.iter_rows(values_only=True):
        for v in row:
            if v and "RFQ-2026-04" in str(v):
                found_rfq = True; break
        if found_rfq: break
    assert found_rfq, "Tab 1 missing RFQ id"

    # Tab 6 must have the flagged item
    ws6 = wb["6_FollowUp_Items"]
    found_flag = False
    for row in ws6.iter_rows(values_only=True):
        if row and row[0] == "A":
            found_flag = True; break
    assert found_flag, "Tab 6 missing flagged item"

    print(f"PASS  xlsx_round_trips  ({len(raw):,} bytes, {len(expected_tabs)} tabs)")


if __name__ == "__main__":
    test_metrics_clean_baseline()
    test_metrics_after_manual_work()
    test_cost_avoidance_vs_savings()
    test_follow_up_lifecycle()
    test_narrative_text_present()
    test_xlsx_round_trips()
    print("\nALL DECISION-SUMMARY TESTS PASS")
