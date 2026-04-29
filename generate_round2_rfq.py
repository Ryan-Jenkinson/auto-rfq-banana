"""ONE-OFF — Round 2 per-supplier RFQ generator.

Why this exists:
    Round 1 went out with the full Coupa item universe (way too big to price).
    Round 2 cuts to just the auto-rfq-banana RECOMMENDED items, pre-fills each
    supplier's own Round 1 quote so they can confirm or revise rather than
    re-key, and leaves blank rows for items they didn't price the first time.

Per-supplier isolation:
    Each output xlsx contains EXACTLY ONE supplier's Round 1 data. No
    cross-supplier benchmark numbers, no Andersen-internal targets, no
    historical paid prices. Read-only fields are limited to item identity +
    annual qty so the supplier can locate the part — that's it.

Input field trim (from the original polished template):
    Round 1 asked the supplier for 24 columns. That overhead is part of why
    they couldn't get through the list. Round 2 asks for 8:
      Quote Price  ·  Quote UOM  ·  Your Part #  ·  Lead Time (days)
      No Bid? (Y/N)  ·  No Bid Reason  ·  Supplier Notes  ·  Valid Through Date

Run: python3 generate_round2_rfq.py
Output: ~/Desktop/round2/Round2_<Supplier>_<YYYY-MM-DD>.xlsx (one per supplier)
"""
from __future__ import annotations
import os, re, datetime as dt
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HOME = Path.home()
DESKTOP = HOME / "Desktop"
DOWNLOADS = HOME / "Downloads"
OUT_DIR = DESKTOP / "round2"
OUT_DIR.mkdir(exist_ok=True)

CANDIDATE_LIST = DESKTOP / "RFQ_candidate_list_McMaster Coupa.xlsx"

SUPPLIERS = [
    {
        "name":    "Fastenal",
        "file":    DOWNLOADS / "rfq" / "Fastenal" / "Fastenal 4.23.xlsx",
        "sheet":   "Fastenal",
        "header":  7,
        "cols":    {  # 1-indexed column numbers in the Round 1 file
            "commodity": 2, "description": 3, "eam_pn": 4, "part_number": 5,
            "item_num":  6, "mfg_name":   7, "qty": 8, "uom": 9,
            "quoted_price": 10, "verified_price": 11, "notes": 12,
        },
    },
    {
        "name":    "Grainger",
        "file":    DOWNLOADS / "rfq" / "grainger" / "Andersen Grainger RFQ 4.10.26 - McMaster Items.xlsx",
        "sheet":   "MainItemList",
        "header":  7,
        "cols":    {
            "commodity": 2, "description": 3, "eam_pn": 4, "part_number": 5,
            "item_num":  6, "mfg_name":   7, "qty": 8, "uom": 9,
            "quoted_price": 10, "notes": 11,
        },
    },
    {
        "name":    "MSC",
        "file":    DOWNLOADS / "rfq" / "msc" / "Anderson RFQ - McMaster Items_updated 4-8-26.xlsx",
        "sheet":   "MainItemList",
        "header":  7,
        "cols":    {
            "commodity": 2, "description": 3, "eam_pn": 4, "part_number": 5,
            "item_num":  6, "mfg_name":   7, "qty": 8, "uom": 9,
            "quoted_price": 10, "notes": 11,
        },
    },
]

# Markers in the Notes col that signal supplier explicitly declined to bid.
# Word-boundary matched so we don't false-positive on "fasteNAl".
NO_BID_MARKERS = re.compile(
    r"\b(?:no\s+bid|no\s+similar|not\s+available|phased\s+out|obsolete|"
    r"discontinued|n/?a|tbd|cannot\s+source|cannot\s+quote|won'?t\s+bid|"
    r"non[\s-]?stocked)\b",
    re.I,
)
NEED_INFO_MARKERS = re.compile(r"\bneed\s+more\s+information\b", re.I)


def norm_key(s) -> str:
    """Normalize an item key for cross-file lookup. Strip whitespace + uppercase."""
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).upper().strip()


def parse_supplier_round1(spec: dict) -> dict:
    """Read one supplier's Round 1 file and return key -> bid dict.

    Tries multiple identifier columns (Item #, Part Number, EAM) so we still
    match even when the supplier filled out one but not the others.
    """
    wb = openpyxl.load_workbook(spec["file"], data_only=True, read_only=True)
    ws = wb[spec["sheet"]]
    cols = spec["cols"]
    bids = {}
    for ridx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if ridx <= spec["header"]:
            continue
        if not row or all(c is None for c in row):
            continue
        def g(field):
            idx = cols.get(field)
            if not idx or idx > len(row):
                return None
            v = row[idx - 1]
            if isinstance(v, str):
                v = v.strip().replace("\xa0", "")
                if v == "":
                    return None
            return v
        item_num = g("item_num")
        part_num = g("part_number")
        eam_pn   = g("eam_pn")
        if not (item_num or part_num or eam_pn):
            continue
        price = g("quoted_price")
        try:
            price = float(price) if price not in (None, "") else None
        except (TypeError, ValueError):
            price = None
        notes = g("notes")
        notes_str = str(notes) if notes else ""

        if price is not None and price > 0:
            status = "PRICED"
        elif notes_str and NO_BID_MARKERS.search(notes_str):
            status = "NO_BID"
        elif notes_str and NEED_INFO_MARKERS.search(notes_str):
            status = "NEED_INFO"
        elif price == 0 and notes_str:
            status = "NEED_INFO"
        else:
            # Row returned but no price and no usable note — supplier echoed
            # the item back without engaging. Distinct from "never seen".
            status = "BLANK_ROW"

        bid = {
            "round1_price": price,
            "round1_uom":   g("uom"),
            "round1_notes": notes_str,
            "round1_status": status,
            "supplier_part": None,
        }
        # Verified price overrides quoted price when supplier returned an updated value.
        verified = g("verified_price") if "verified_price" in cols else None
        if verified is not None:
            try:
                vfloat = float(verified)
                if vfloat > 0:
                    bid["round1_price"] = vfloat
            except (TypeError, ValueError):
                pass
        for k in (item_num, part_num, eam_pn):
            if k:
                nk = norm_key(k)
                if nk and nk not in bids:
                    bids[nk] = bid
    wb.close()
    return bids


def load_candidate_items() -> list:
    """Read the auto-rfq-banana RECOMMENDED list (already saved to Desktop)."""
    wb = openpyxl.load_workbook(CANDIDATE_LIST, data_only=True, read_only=True)
    ws = wb["RFQ Candidate List"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip() for h in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}
    out = []
    for r in rows[1:]:
        if not r or r[idx["Item #"]] in (None, ""):
            continue
        out.append({
            "item_num":     r[idx["Item #"]],
            "description":  r[idx["Description"]],
            "manufacturer": r[idx["Manufacturer"]],
            "mfg_part":     r[idx["MFG Part #"]],
            "commodity":    r[idx["Commodity"]],
            "uom":          r[idx["UOM"]],
            "annual_qty":   r[idx["12mo qty"]] or 0,
        })
    wb.close()
    return out


# ---- Workbook generation -----------------------------------------------------

NAVY      = "1F2A44"
GOLD      = "C99A2E"
GRAY_BG   = "F0F0F0"
GRAY_TEXT = "606060"
YELLOW_BG = "FFF4C2"
GREEN_BG  = "E6F4E6"
RED_BG    = "FCE4E4"
BLUE_BG   = "E4ECF7"
WHITE     = "FFFFFF"

THIN = Side(style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def make_workbook(supplier_name: str, items: list, bid_lookup: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    today = dt.date.today().isoformat()

    def status_of(it):
        b = bid_lookup.get(norm_key(it["item_num"]))
        return b["round1_status"] if b else "ABSENT"
    counts = {k: 0 for k in ("PRICED", "NO_BID", "NEED_INFO", "BLANK_ROW", "ABSENT")}
    for it in items:
        counts[status_of(it)] += 1
    priced     = counts["PRICED"]
    no_bid     = counts["NO_BID"]
    need_info  = counts["NEED_INFO"]
    blank_row  = counts["BLANK_ROW"]
    absent     = counts["ABSENT"]

    # === Tab 1: Instructions ==================================================
    ws = wb.create_sheet("Instructions")
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"REQUEST FOR QUOTATION  ·  ROUND 2  ·  RFQ-2026-04-MCMASTER-002"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)
    ws["A2"] = f"Issued to: {supplier_name}"
    ws["A2"].font = Font(name="Calibri", size=12, bold=True, color=GRAY_TEXT)
    ws["A3"] = f"Issue date: {today}"
    ws["A3"].font = Font(name="Calibri", size=10, color=GRAY_TEXT)

    ws["A5"] = "Why we're sending a Round 2"
    ws["A5"].font = Font(bold=True, size=12, color=NAVY)
    ws.merge_cells("A6:F11")
    ws["A6"] = (
        "Round 1 included our full Coupa item universe — too many SKUs for any "
        "supplier to price comprehensively. Thank you for what you returned.\n\n"
        "Round 2 is the focused subset our analysis identified as the most "
        "important items to lock in pricing for. To save you time, the rows "
        f"already priced in your Round 1 response are PRE-FILLED below. {supplier_name} only — "
        "no other supplier's data appears in this file.\n\n"
        "Please confirm or revise the pre-filled rows, and quote the blank rows "
        "where possible. If a row is a no-bid, mark it No Bid (Y) and give us a "
        "short reason."
    )
    ws["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A6"].font = Font(size=10)

    ws["A13"] = "Coverage of your Round 1 vs this Round 2 list"
    ws["A13"].font = Font(bold=True, size=12, color=NAVY)
    cov = [
        ("Round 2 items requested",                    len(items)),
        ("Already priced in your Round 1",             priced),
        ("Round 1 declined / no-bid",                  no_bid),
        ("Round 1 needed more info",                   need_info),
        ("Round 1 echoed back blank — no price/notes", blank_row),
        ("Not seen in your Round 1 response",          absent),
    ]
    for i, (label, val) in enumerate(cov, 14):
        ws.cell(i, 1, label).font = Font(size=10)
        ws.cell(i, 2, val).font = Font(size=10, bold=True)
        ws.cell(i, 2).alignment = Alignment(horizontal="right")

    ws["A21"] = "What we're asking you to fill in (8 fields per row)"
    ws["A21"].font = Font(bold=True, size=12, color=NAVY)
    fields = [
        ("Quote Price",   "USD per Quote UOM. If pre-filled and unchanged, leave as-is."),
        ("Quote UOM",     "What unit your price is for (Each / Pack / Foot / etc). Critical — UOM mismatches caused most Round 1 issues."),
        ("Your Part #",   "Your distributor SKU so we can place the order with you."),
        ("Lead Time (days)", "Typical lead time from PO to dock. Best estimate is fine."),
        ("No Bid?",       "Y if you cannot quote this item. Otherwise leave blank."),
        ("No Bid Reason", "Short reason if No Bid = Y."),
        ("Supplier Notes", "Anything else relevant — substitutions, pack-size differences, etc."),
        ("Valid Through Date", "Date your quoted price holds firm through (YYYY-MM-DD)."),
    ]
    for i, (f, desc) in enumerate(fields, 22):
        ws.cell(i, 1, f).font = Font(bold=True, size=10, color=NAVY)
        ws.cell(i, 2, desc).font = Font(size=10)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 75
    for c in "CDEF":
        ws.column_dimensions[c].width = 10

    # === Tab 2: Round 2 Items =================================================
    items_ws = wb.create_sheet("Round 2 Items")
    items_ws.sheet_view.showGridLines = False
    items_ws.freeze_panes = "A4"

    items_ws["A1"] = f"ROUND 2 ITEMS  ·  {supplier_name}  ·  {len(items)} items"
    items_ws["A1"].font = Font(size=14, bold=True, color=NAVY)
    items_ws["A2"] = (
        "GRAY cells = read-only reference (do not modify).  "
        "BLUE cells = your Round 1 echo, please confirm or revise.  "
        "YELLOW cells = your input."
    )
    items_ws["A2"].font = Font(size=9, italic=True, color=GRAY_TEXT)

    headers = [
        # Read-only Andersen context (gray)
        ("Status",          "ro_status"),
        ("Item #",          "ro"),
        ("Description",     "ro"),
        ("Manufacturer",    "ro"),
        ("MFG Part #",      "ro"),
        ("Commodity",       "ro"),
        ("Annual Qty",      "ro"),
        ("Our UOM",         "ro"),
        # Round 1 echo (blue, supplier should confirm/revise)
        ("Round 1 — Your Price", "echo"),
        ("Round 1 — Your UOM",   "echo"),
        ("Round 1 — Your Notes", "echo"),
        # Round 2 inputs (yellow)
        ("Quote Price",        "input"),
        ("Quote UOM",          "input"),
        ("Your Part #",        "input"),
        ("Lead Time (days)",   "input"),
        ("No Bid?",            "input"),
        ("No Bid Reason",      "input"),
        ("Supplier Notes",     "input"),
        ("Valid Through Date", "input"),
    ]
    for ci, (label, kind) in enumerate(headers, 1):
        c = items_ws.cell(3, ci, label)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    items_ws.row_dimensions[3].height = 32

    widths = [28, 14, 50, 22, 18, 22, 12, 10, 14, 12, 36, 14, 12, 16, 14, 10, 28, 32, 16]
    for i, w in enumerate(widths, 1):
        items_ws.column_dimensions[get_column_letter(i)].width = w

    # Status text + color per row
    status_disp = {
        "PRICED":     ("Already quoted — confirm/revise",        GREEN_BG),
        "NO_BID":     ("Round 1 no-bid — please reconsider",      RED_BG),
        "NEED_INFO":  ("Round 1 needed info — please quote",      YELLOW_BG),
        "BLANK_ROW":  ("Round 1 returned blank — please quote",   YELLOW_BG),
        "ABSENT":     ("Not seen in Round 1 — please quote",      BLUE_BG),
    }

    for ri, it in enumerate(items, start=4):
        bid = bid_lookup.get(norm_key(it["item_num"]))
        st = bid["round1_status"] if bid else "ABSENT"
        st_text, st_color = status_disp[st]

        row_vals = [
            st_text,
            it["item_num"],
            it["description"],
            it["manufacturer"],
            it["mfg_part"],
            it["commodity"],
            it["annual_qty"],
            it["uom"],
            bid["round1_price"] if bid else None,
            bid["round1_uom"]   if bid else None,
            bid["round1_notes"] if bid else None,
            None, None, None, None, None, None, None, None,  # 8 input cols
        ]
        for ci, v in enumerate(row_vals, 1):
            c = items_ws.cell(ri, ci, v)
            c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
            kind = headers[ci - 1][1]
            if ci == 1:
                c.fill = PatternFill("solid", fgColor=st_color)
                c.font = Font(size=9, bold=True)
            elif kind == "ro":
                c.fill = PatternFill("solid", fgColor=GRAY_BG)
                c.font = Font(size=9, color=GRAY_TEXT)
            elif kind == "echo":
                c.fill = PatternFill("solid", fgColor=BLUE_BG)
            elif kind == "input":
                c.fill = PatternFill("solid", fgColor=YELLOW_BG)
            if ci in (7, 9, 12):  # numeric cols
                c.number_format = "#,##0.0000" if ci != 7 else "#,##0"

    # Auto-filter on header
    items_ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3+len(items)}"

    return wb


def main():
    items = load_candidate_items()
    print(f"Recommended items loaded: {len(items)}")
    today = dt.date.today().isoformat()

    summary_rows = []
    for spec in SUPPLIERS:
        name = spec["name"]
        if not spec["file"].exists():
            print(f"  ! {name}: bid file not found at {spec['file']}")
            continue
        bids = parse_supplier_round1(spec)
        wb = make_workbook(name, items, bids)
        out = OUT_DIR / f"Round2_{name}_{today}.xlsx"
        wb.save(out)
        cnt = {k: 0 for k in ("PRICED","NO_BID","NEED_INFO","BLANK_ROW","ABSENT")}
        for it in items:
            b = bids.get(norm_key(it["item_num"]))
            cnt[b["round1_status"] if b else "ABSENT"] += 1
        print(f"  ✓ {name}: priced={cnt['PRICED']}  no-bid={cnt['NO_BID']}  "
              f"need-info={cnt['NEED_INFO']}  blank-row={cnt['BLANK_ROW']}  "
              f"absent={cnt['ABSENT']}  →  {out}")
        summary_rows.append((name, cnt, str(out)))

    print("\nDone.")
    for name, cnt, path in summary_rows:
        print(f"  {name}  →  {path}")
        for label, key in [
            ("priced (confirm/revise)", "PRICED"),
            ("no-bid (reconsider)",     "NO_BID"),
            ("need info (please quote)","NEED_INFO"),
            ("returned blank",          "BLANK_ROW"),
            ("not in round 1",          "ABSENT"),
        ]:
            print(f"     {cnt[key]:5d}  {label}")


if __name__ == "__main__":
    main()
